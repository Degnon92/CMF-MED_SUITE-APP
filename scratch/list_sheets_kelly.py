import openpyxl
import os

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")
wb = openpyxl.load_workbook(excel_path, read_only=True)
print(f"Total sheets in PROFORMA CHIRURGIE Exemplaire Proforma: {len(wb.sheetnames)}")
print("Matching sheets:")
for name in wb.sheetnames:
    if "KELLY" in name.upper():
        print(name)
