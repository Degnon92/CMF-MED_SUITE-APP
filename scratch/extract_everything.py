import os
import re
import glob
import json
import openpyxl
import docx
from docx.oxml.ns import qn

workspace = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
app_dir = os.path.join(workspace, "MercyFiatMedSuiteDesktop")

print("Démarrage de l'extraction complète du plan de travail...")

# -------------------------------------------------------------
# 1. FONCTIONS DE NETTOYAGE & NORMALISATION (reproduisant le JS)
# -------------------------------------------------------------
def clean_patient_name(name):
    if not name or not isinstance(name, str):
        return ""
    clean = name.strip()
    
    prefix_pattern = re.compile(
        r'^(?:CERTIFICAT\s+DE\s+MARIAGE|CERTIFICAT\s+DE\s+NON\s+BEGAIEMENT|CERTIFICAT\s+MEDICAL\s+INITIAL\s+DE\s+CONSTATATION\s+DE\s+COUPS\s+ET\s+BLESSURES|CERTIFICAT\s+MEDICAL\s+POUR\s+COUPS\s+ET\s+BLESSURES|CERTIFICAT\s+MEDICAL\s+DE\s+L[’\']ETAT\s+ACTUEL|CERTIFICAT\s+MEDICAL\s+INITIAL|CERTIFICAT\s+MEDICAL|CERTIFICAT\s+MED\s+INITIAL|CERTIFICAT\s+DE\s+REPOS|CERTIFICAT\s+DE\s+REPRISE|CERTIFICAT\s+DE\s+GUERISON|CERTIFICAT\s+DE\s+GUÉRISON|CERTIFICAT\s+DE|CERTIFICAT|RAPPORT\s+MEDICAL|RAPPORT\s+DE\s+MONSIEUR|RAPPORT\s+DE\s+MME|RAPPORT\s+DE|RAPPORT\s+D\'HOSPI|RAPPORT\s+D\'HOSPITALISATION|RAPPORT\s+DE\s+CONSULTATION|RAPPORT|CRO\s+MODELE|CRO|CMI|MEDICAL|MED\s+INITIAL|GUERISON\s+DE\s+MONSIEUR|GUERISON\s+DE\s+MME|GUERISON\s+DE|GUERISON|GUÉRISON|DECES\s+DE\s+MONSIEUR|DECES\s+DE\s+MME|DECES\s+DE|DECES|DÉCÈS|D\'HOSPI\s+TYPE|D\'HOSPI|DHOSPI|ATTESTATION\s+DE\s+GUERISON|ATTESTATION\s+DE\s+GUÉRISON|ATTESTATION\s+DE|ATTESTATION)\s+',
        re.IGNORECASE
    )
    
    old_clean = ""
    while clean != old_clean:
        old_clean = clean
        clean = prefix_pattern.sub("", clean).strip()
        
    split_pattern = re.compile(
        r'(?:CERTIFICAT|JE\s+SOUSSIGN|JE\s+SOUSSIGNE|RAPPORT|DOSSIER|N°|NO\s+DOSSIER|CMI|CRO|CLINIQUE|MÉDECINE|MEDECINE|DIAGNOSTIC|INTERVENTION|CLIENT|ASSURANCE|AFRICAINE|NSIA|ALLIANZ|SUNU|AROO|SAAR|CORIS|FEDAS|MUTUELLE|SÉJOUR|SEJOUR|DATE|OPÉRATOIRE|OPERATOIRE|CERTFICAT|CERTIF|PATIENT|PATIENTE|COTONOU|RUE\s+PAVILLON|TEL\s*\:|E\-MAIL|EMAIL|E\s+MAIL|GUERISON|GUÉRISON|DECES|DÉCÈS)',
        re.IGNORECASE
    )
    
    match = split_pattern.search(clean)
    if match:
        clean = clean[:match.start()].strip()
        
    clean = re.sub(r'(?:\s+|:)\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    if clean.upper().endswith("AGE") and len(clean) > 5:
        clean = clean[:-3].strip()
        
    clean = re.sub(r'[\s\-\.\,\:\_]+$', '', clean).strip()
    clean = re.sub(r'\s+', ' ', clean)
    clean = re.sub(r'\s+\(?(?:ok|copie|2|3|4|1)\)?$', '', clean, flags=re.IGNORECASE).strip()
    return clean

def split_nom_prenom(full_name):
    cleaned = clean_patient_name(full_name)
    if not cleaned:
        return "", ""
    parts = cleaned.split(' ')
    if len(parts) == 1:
        return cleaned.upper(), ""
    nom_parts = []
    prenom_parts = []
    for part in parts:
        if part.isupper() and len(part) > 1:
            nom_parts.append(part)
        else:
            prenom_parts.append(part)
    if not nom_parts:
        nom = parts[0].upper()
        prenom = " ".join(parts[1:])
    else:
        nom = " ".join(nom_parts).upper()
        prenom = " ".join(prenom_parts)
    if not prenom and len(parts) > 1:
        nom = parts[0].upper()
        prenom = " ".join(parts[1:])
    return nom.strip(), prenom.strip()

def clean_clinical_term(term):
    if not term or not isinstance(term, str):
        return ""
    clean = term.strip()
    clean = re.sub(r'^[\s\-\.\,\:\_\•\*\t\?\/]+', '', clean).strip()
    clean = re.sub(r'[\s\-\.\,\:\_\?\/]+$', '', clean).strip()
    # Retirer les retours à la ligne
    clean = clean.replace('\n', ' ').replace('\r', ' ')
    clean = re.sub(r'\s+', ' ', clean).strip()
    
    # Exclure les termes trop longs ou administratifs
    if len(clean) < 5 or len(clean) > 120:
        return ""
    upper = clean.upper()
    forbidden = ["JE SOUSSIGN", "CLINIQUE", "TEL :", "EMAIL", "IFU", "RCCM", "ORABANK", "SEME AGUE", "COMPTE RENDU", "RAPPORT DE"]
    if any(f in upper for f in forbidden):
        return ""
    return clean

def map_insurance(text):
    if not text:
        return "PRIVE"
    t = text.lower().strip()
    if "sanlam" in t or "saham" in t:
        return "SANLAM"
    elif "ascoma" in t:
        return "ASCOMA"
    elif "sunu" in t:
        return "SUNU"
    elif "nsia" in t:
        return "NSIA"
    elif "atlantique" in t:
        return "ATLANTIQUE"
    elif "africaine" in t:
        if "sinistre" in t or "accident" in t:
            return "AFRICAINE_SINISTRE"
        return "AFG"
    elif "lotto" in t or "loto" in t:
        return "LOTTO_FOOTBALL_CLUB"
    elif "coton" in t:
        return "COTON_SPORT"
    elif "nobila" in t:
        return "NOBILA"
    elif "gras" in t or "savoye" in t:
        return "GRAS SAVOYE"
    elif "olea" in t:
        return "OLEA"
    elif "fonds" in t or "garantie" in t or "fga" in t:
        return "FONDS_GARANTIE_AUTO"
    elif "transvie" in t:
        return "TRANSVIE"
    elif "sobremap" in t or "sobemap" in t:
        return "SOBEMAP"
    elif "port" in t or "pac" in t:
        return "PORT_AUTONOME_COTONOU"
    return "PRIVE"

# -------------------------------------------------------------
# 2. ANALYSE RECURSIVE DE TOUT LE DOSSIER
# -------------------------------------------------------------
extracted_patients = {}
extracted_diagnoses = set()
extracted_interventions = set()

def get_textbox_paragraphs(doc):
    tb_texts = []
    root = doc.element
    for el in root.iter():
        if el.tag.endswith('txbxContent'):
            for p in el.findall(qn('w:p')):
                p_obj = docx.text.paragraph.Paragraph(p, doc)
                if p_obj.text.strip():
                    tb_texts.append(p_obj.text.strip())
    return tb_texts

def add_patient(name, diagnosis, intervention, kCode, age, insurer):
    if not name:
        return
    fullname = name.strip().upper()
    if len(fullname) < 3 or len(fullname) > 50:
        return
    
    # Exclure les mots administratifs du nom
    forbidden_names = ["RAPPORT", "CLINIQUE", "CERTIFICAT", "EXEMPLAIRE", "FACTURE", "PROFORMA", "DOSSIER", "SUIVI"]
    if any(fn in fullname for fn in forbidden_names):
        return
    
    diag_clean = clean_clinical_term(diagnosis)
    interv_clean = clean_clinical_term(intervention)
    
    if diag_clean:
        extracted_diagnoses.add(diag_clean)
    if interv_clean:
        extracted_interventions.add(interv_clean)
        
    if fullname not in extracted_patients:
        extracted_patients[fullname] = {
            "name": fullname,
            "diagnosis": diag_clean or "Bilan clinique",
            "intervention": interv_clean,
            "kCode": kCode or "",
            "age": age or "35 ans",
            "matricule": "",
            "insurer": insurer or "PRIVE",
            "priseEnCharge": 100 if insurer in ["LOTTO_FOOTBALL_CLUB", "COTON_SPORT", "ENERGIE_BASKET_BALL", "AFRICAINE_SINISTRE", "FONDS_GARANTIE_AUTO"] else (80 if insurer != "PRIVE" else 0)
        }
    else:
        # Fusionner
        p = extracted_patients[fullname]
        if diag_clean and (p["diagnosis"] == "Bilan clinique" or not p["diagnosis"]):
            p["diagnosis"] = diag_clean
        if interv_clean and not p["intervention"]:
            p["intervention"] = interv_clean
        if kCode and not p["kCode"]:
            p["kCode"] = kCode
        if age and (p["age"] == "35 ans" or not p["age"]):
            p["age"] = age
        if insurer and insurer != "PRIVE" and p["insurer"] == "PRIVE":
            p["insurer"] = insurer
            p["priseEnCharge"] = 100 if insurer in ["LOTTO_FOOTBALL_CLUB", "COTON_SPORT", "ENERGIE_BASKET_BALL", "AFRICAINE_SINISTRE", "FONDS_GARANTIE_AUTO"] else 80

# A. SCAN DOSSIERS POUR DOCX
print("Scan de tous les fichiers Word (.docx)...")
docx_files = []
for r, d, fs in os.walk(workspace):
    if "MercyFiatMedSuiteDesktop" in r or "node_modules" in r or ".git" in r:
        continue
    for file in fs:
        if file.endswith(".docx") and not file.startswith("~$"):
            docx_files.append(os.path.join(r, file))

for f_path in docx_files:
    try:
        doc = docx.Document(f_path)
        body_text = "\n".join([p.text for p in doc.paragraphs])
        tb_text = "\n".join(get_textbox_paragraphs(doc))
        full_text = body_text + "\n" + tb_text
        
        # 1. Rechercher des blocs "Patient:"
        # Certains documents contiennent des informations structurées
        segments = re.split(r"patient\s*:", full_text, flags=re.IGNORECASE)
        if len(segments) > 1:
            for seg in segments[1:]:
                # Premier paragraphe du segment contient le nom
                seg_lines = [l.strip() for l in seg.split("\n") if l.strip()]
                if not seg_lines:
                    continue
                p_name = clean_patient_name(seg_lines[0])
                if not p_name:
                    continue
                    
                # Analyser le reste du segment
                seg_content = "\n".join(seg_lines[1:])
                diagnosis = ""
                intervention = ""
                age = ""
                insurer = "PRIVE"
                
                diag_m = re.search(r'(?:diagnostic|motif|bilan\s+lésionnel)\s*:\s*([^\n\.]+)', seg_content, re.IGNORECASE)
                if diag_m:
                    diagnosis = diag_m.group(1).strip()
                    
                int_m = re.search(r'(?:intervention|geste|acte|traitement)\s*:\s*([^\n\.]+)', seg_content, re.IGNORECASE)
                if int_m:
                    intervention = int_m.group(1).strip()
                    
                age_m = re.search(r'(?:age|âge)\s*:\s*([^\n\.]+)', seg_content, re.IGNORECASE)
                if age_m:
                    age = age_m.group(1).strip()
                    
                ins_m = re.search(r'(?:client|assurance|organisme)\s*:\s*([^\n\.]+)', seg_content, re.IGNORECASE)
                if ins_m:
                    insurer = map_insurance(ins_m.group(1))
                    
                add_patient(p_name, diagnosis, intervention, "", age, insurer)
        else:
            # Document unique sans tag "Patient:"
            # Extraire les infos à partir du titre ou du texte global
            # Chercher des expressions comme "Je soussigné, ..., certifie que l'état de santé de ..."
            m_souss = re.search(r"(?:de\s+m\.|de\s+mme|de\s+l\’enfant|patient\s+|de\s+)(?:monsieur|madame|m\.|mme\s+)?([A-Z\s]{4,35})", full_text, re.IGNORECASE)
            p_name = ""
            if m_souss:
                p_name = clean_patient_name(m_souss.group(1))
            
            if not p_name:
                # Essayer à partir du nom du fichier
                fn = os.path.basename(f_path).replace(".docx", "")
                p_name = clean_patient_name(fn)
                
            if p_name:
                diagnosis = ""
                intervention = ""
                age = ""
                insurer = "PRIVE"
                
                # Diagnostic
                diag_m = re.search(r'(?:diagnostic|motif|bilan\s+lésionnel)\s*:\s*([^\n\.]+)', full_text, re.IGNORECASE)
                if diag_m:
                    diagnosis = diag_m.group(1).strip()
                
                # Intervention
                int_m = re.search(r'(?:intervention|geste|acte|traitement)\s*:\s*([^\n\.]+)', full_text, re.IGNORECASE)
                if int_m:
                    intervention = int_m.group(1).strip()
                
                # Age
                age_m = re.search(r'(\d+)\s*(?:ans|mois)', full_text, re.IGNORECASE)
                if age_m:
                    age = age_m.group(0).strip()
                
                # Insurer
                insurer = map_insurance(full_text)
                
                add_patient(p_name, diagnosis, intervention, "", age, insurer)
    except Exception as e:
        pass

# B. SCAN DOSSIERS POUR XLSX (FACTURES/PROFORMAS)
print("Scan de tous les fichiers Excel (.xlsx)...")
xlsx_files = []
for r, d, fs in os.walk(workspace):
    if "MercyFiatMedSuiteDesktop" in r or "node_modules" in r or ".git" in r:
        continue
    for file in fs:
        if file.endswith(".xlsx") and not file.startswith("~$"):
            xlsx_files.append(os.path.join(r, file))

for f_path in xlsx_files:
    try:
        wb = openpyxl.load_workbook(f_path, data_only=True)
        for sheetname in wb.sheetnames:
            if "feuil" in sheetname.lower() or "sheet" in sheetname.lower():
                continue
            sheet = wb[sheetname]
            patient_name = ""
            intervention = ""
            diagnosis = ""
            insurer = "PRIVE"
            kCode = ""
            age = "35 ans"
            
            # Balayer les 40 premières lignes et 10 premières colonnes
            for r in range(1, 40):
                for c in range(1, 10):
                    val = sheet.cell(r, c).value
                    if not val or not isinstance(val, str):
                        continue
                    val_str = val.strip()
                    val_lower = val_str.lower()
                    
                    if val_lower.startswith("patient:") or val_lower.startswith("patient :"):
                        patient_name = val_str.split(":", 1)[1].strip()
                    elif val_lower.startswith("client:") or val_lower.startswith("client :"):
                        insurer = map_insurance(val_str.split(":", 1)[1].strip())
                    elif val_lower.startswith("intervention:") or val_lower.startswith("intervention :"):
                        intervention = val_str.split(":", 1)[1].strip()
                    elif val_lower.startswith("diagnostic:") or val_lower.startswith("diagnostic :"):
                        diagnosis = val_str.split(":", 1)[1].strip()
                    
                    # Chercher un code K
                    k_m = re.search(r'\b(K\d+)\b', val_str)
                    if k_m:
                        kCode = k_m.group(1)
            
            if not patient_name and len(sheetname) > 3:
                patient_name = sheetname
                
            p_clean = clean_patient_name(patient_name)
            if p_clean:
                add_patient(p_clean, diagnosis, intervention, kCode, age, insurer)
                
            # Extraire également toutes les désignations d'actes dans la colonne A (1) comme interventions potentielles !
            for r in range(8, 50):
                val = sheet.cell(r, 1).value
                if val and isinstance(val, str) and len(val.strip()) > 5:
                    v_str = val.strip()
                    # Exclure le total, la caisse, etc.
                    if not any(kw in v_str.lower() for kw in ["total", "caisse", "option", "chambre", "hébergement", "hebergement", "repas", "soins", "acte"]):
                        clean_acte = clean_clinical_term(v_str)
                        if clean_acte:
                            extracted_interventions.add(clean_acte)
    except Exception as e:
        pass

print(f"Extraction terminée !")
print(f"Nombre total de patients recensés : {len(extracted_patients)}")
print(f"Nombre total de diagnostics uniques : {len(extracted_diagnoses)}")
print(f"Nombre total de interventions uniques : {len(extracted_interventions)}")

# -------------------------------------------------------------
# 3. CHARGER LES DONNÉES DE L'APPLICATION EXISTANTE & FUSIONNER
# -------------------------------------------------------------
existing_patients_file = os.path.join(app_dir, "patients_db.json")
existing_patients = []
if os.path.exists(existing_patients_file):
    try:
        with open(existing_patients_file, "r", encoding="utf-8") as f:
            existing_patients = json.load(f)
        print(f"Patients existants dans l'app : {len(existing_patients)}")
    except Exception as e:
        print("Erreur lecture patients existants :", e)

# Fusionner les patients
for p in existing_patients:
    add_patient(p["name"], p.get("diagnosis", ""), p.get("intervention", ""), p.get("kCode", ""), p.get("age", ""), p.get("insurer", "PRIVE"))

final_patients_list = list(extracted_patients.values())
print(f"Patients après fusion complète : {len(final_patients_list)}")

# Écrire dans patients_db.json
with open(existing_patients_file, "w", encoding="utf-8") as f:
    json.dump(final_patients_list, f, indent=4, ensure_ascii=False)
print("Fichier patients_db.json mis à jour avec succès !")

# -------------------------------------------------------------
# 4. METTRE À JOUR database.js AVEC LES DIAGNOSTICS & INTERVENTIONS
# -------------------------------------------------------------
database_js_path = os.path.join(app_dir, "database.js")
if os.path.exists(database_js_path):
    with open(database_js_path, "r", encoding="utf-8") as f:
        db_js_content = f.read()
        
    # Extraire les listes d'origines de database.js (en secours)
    orig_diagnoses = []
    orig_interventions = []
    
    # regex pour capturer DIAGNOSES: [ ... ]
    diag_match = re.search(r'DIAGNOSES\s*:\s*\[(.*?)\]', db_js_content, re.DOTALL)
    if diag_match:
        items = re.findall(r'"(.*?)"', diag_match.group(1))
        orig_diagnoses = [it for it in items if it.strip()]
        
    interv_match = re.search(r'INTERVENTIONS\s*:\s*\[(.*?)\]', db_js_content, re.DOTALL)
    if interv_match:
        items = re.findall(r'"(.*?)"', interv_match.group(1))
        orig_interventions = [it for it in items if it.strip()]
        
    # Fusionner et dédupliquer
    all_diags = sorted(list(extracted_diagnoses.union(set(orig_diagnoses))))
    all_intervs = sorted(list(extracted_interventions.union(set(orig_interventions))))
    
    # Nettoyer
    all_diags = [d for d in all_diags if clean_clinical_term(d)]
    all_intervs = [i for i in all_intervs if clean_clinical_term(i)]
    
    # Remplacer dans le fichier
    new_diag_str = "DIAGNOSES: [\n        " + ",\n        ".join(f'"{d}"' for d in all_diags) + "\n    ]"
    new_interv_str = "INTERVENTIONS: [\n        " + ",\n        ".join(f'"{i}"' for i in all_intervs) + "\n    ]"
    
    # Effectuer le remplacement
    updated_content = re.sub(r'DIAGNOSES\s*:\s*\[.*?\]', new_diag_str, db_js_content, flags=re.DOTALL)
    updated_content = re.sub(r'INTERVENTIONS\s*:\s*\[.*?\]', new_interv_str, updated_content, flags=re.DOTALL)
    
    with open(database_js_path, "w", encoding="utf-8") as f:
        f.write(updated_content)
    print("Fichier database.js mis à jour avec les diagnostics et interventions fusionnés !")
else:
    print("Fichier database.js introuvable !")
