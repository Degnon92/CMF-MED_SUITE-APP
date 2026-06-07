import openpyxl
import os
import re

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")

wb = openpyxl.load_workbook(excel_path, data_only=True)

feuil_sheets = [s for s in wb.sheetnames if "feuil" in s.lower() and s.lower() not in ["feuil1", "feuil2", "feuil4"]]
print(f"Found {len(feuil_sheets)} Feuil sheets.")

for sheetname in feuil_sheets[:15]:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    for r in range(1, 8):
        row_vals = []
        for c in range(1, 5):
            val = sheet.cell(r, c).value
            if val is not None:
                row_vals.append(f"R{r}C{c}: {repr(val)}")
        if row_vals:
            print(", ".join(row_vals))
