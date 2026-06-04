import openpyxl, re
wb = openpyxl.load_workbook(r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx", read_only=True)

has_patient_colon = 0
has_patient_any = 0
total_sheets = 0

for sheet_name in wb.sheetnames:
    total_sheets += 1
    ws = wb[sheet_name]
    found_colon = False
    found_any = False
    for r in ws.iter_rows(values_only=True):
        for c in r:
            if c and isinstance(c, str):
                if re.search(r'patient\s*:\s*.+', c, re.I):
                    found_colon = True
                if 'patient' in c.lower():
                    found_any = True
    if found_colon:
        has_patient_colon += 1
    if found_any:
        has_patient_any += 1

print(f"Total sheets: {total_sheets}")
print(f"Sheets with 'PATIENT :' (colon): {has_patient_colon}")
print(f"Sheets with 'patient' (anywhere): {has_patient_any}")
