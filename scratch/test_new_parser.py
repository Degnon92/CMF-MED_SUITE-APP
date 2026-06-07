import openpyxl
import os
import re

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_main = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")

REJECT_WORDS = {
    # Clinic info
    'clinique', 'mercy', 'fiat', 'medecine', 'générale', 'generale', 'specialites', 'spécialités', 'médicales', 'medicales', 'chirurgicales', 'laboratoire',
    # Medical terms
    'chirurgien', 'anesthesiste', 'anesthésiste', 'infirmier', 'kinésithérapeute', 'kinesitherapeute', 'radiologue', 'medecin', 'docteur',
    'forfait', 'materiel', 'matériels', 'implant', 'implants', 'amos', 'clou', 'gamma', 'plaque', 'vis', 'broche', 'sonde',
    'chirurgical', 'chirurgicale', 'ostheosynthese', 'osteosynthese', 'nephrolithotomie', 'ureteroscopie', 'arthroscopie', 'osteotomie',
    'ostheosynthèse', 'percutanee', 'percutanée', 'nlpc',
    'unitaire', 'quantite', 'quantité', 'designation', 'désignation', 'montant', 'total', 'principal', 'complement', 'complément',
    'acte', 'actes', 'cs', 'echographie', 'scanner', 'radio',
    'femur', 'fémur', 'tibia', 'humerus', 'humérus', 'cubitus', 'ulna', 'fibula', 'cheville', 'genou', 'hanche', 'épaule', 'poignet',
    'tibiale', 'fémorale', 'femorales',
    'athemy', 'kinésithérapies', 'kado', 'seance', 'séance', 'intervention', 'diagnostic', 'assurance', 'labo', 'biologie', 'client',
    'patient', 'nom', 'prenom', 'page', 'feuil', 'sheet', 'caisse', 'loto', 'lotto', 'football', 'club', 'sport', 'coton',
    'proforma', 'définitif', 'definitif', 'détail', 'detail', 'assur', 'prive', 'ascoma', 'sanlam', 'sunu', 'nsia', 'olea', 'gras', 'savoye',
    # Additional items
    'médicaments', 'medicaments', 'consommables', 'consommable', 'pharmacie', 'chambre', 'hospitalisation', 'hospit',
    # Placeholders
    'nom..', 'prenom..', 'age..', 'diagnostic..', 'horaires..', 'date..', 'séances..', 'seances..', 'seances'
}

def clean_name(s):
    if not s: return ""
    # Remove anything in parentheses
    s = re.sub(r'\(.*?\)', '', s)
    # Remove multiple spaces
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def is_real_patient_name(s):
    s = clean_name(s)
    if not s or len(s) < 5 or len(s) > 55: return False
    
    words = s.split()
    if len(words) < 2: return False
    
    for w in words:
        wl = w.lower().strip('():,-..')
        wl = re.sub(r'\.+$', '', wl)
        if not wl: continue
        if wl in REJECT_WORDS: return False
        if len(wl) == 1 and wl not in 'abcdefghijklmnopqrstuvwxyz': return False
        
    if 'intervention:' in s.lower(): return False
    if ':' in s and not re.match(r'^[A-Z][a-z]+:', s): return False
    
    for w in words:
        alpha = sum(1 for c in w if c.isalpha())
        if alpha < 2: return False
        
    alpha_total = sum(1 for c in s if c.isalpha() or c in ' -\'')
    if alpha_total / len(s) < 0.70: return False
    
    return True

def find_patient_name_strict(ws):
    all_rows = list(ws.iter_rows(values_only=True, max_row=80))
    
    # Priority 1: Labels
    for ri, row in enumerate(all_rows[:20]):
        for ci, val in enumerate(row):
            if not val or not isinstance(val, str): continue
            s = val.strip()
            sl = s.lower()
            
            if re.match(r'patient\s*:', sl, re.I):
                candidate = s.split(':', 1)[1].strip()
                candidate = candidate.split('\n')[0].strip()
                cleaned = clean_name(candidate)
                if is_real_patient_name(cleaned):
                    return cleaned
            
            if re.match(r'^(?:patient|nom\s+du\s+patient|nom\s*(?:et\s+pr[eé]nom)?)\s*:?\s*$', sl):
                for next_ci in range(ci+1, min(ci+4, len(row))):
                    nv = row[next_ci]
                    if nv and isinstance(nv, str):
                        cleaned = clean_name(nv)
                        if is_real_patient_name(cleaned):
                            return cleaned
                for nri in range(ri+1, min(ri+3, len(all_rows))):
                    if len(all_rows[nri]) > ci:
                        nv = all_rows[nri][ci]
                        if nv and isinstance(nv, str):
                            cleaned = clean_name(nv)
                            if is_real_patient_name(cleaned):
                                return cleaned
    
    # Priority 2: Uppercase/Mixed-case cell check in the first 8 rows
    for ri, row in enumerate(all_rows[:8]):
        for val in row:
            if not val or not isinstance(val, str): continue
            s = val.strip()
            cleaned = clean_name(s)
            if is_real_patient_name(cleaned):
                return cleaned
    
    return None

wb = openpyxl.load_workbook(excel_main, data_only=True)
feuil_sheets = [s for s in wb.sheetnames if re.match(r'^feuil\d+(\s*\(.*\))?$', s.lower().strip())]

results = []
for sheetname in feuil_sheets:
    ws = wb[sheetname]
    name = find_patient_name_strict(ws)
    results.append((sheetname, name))

print("Results with 8-row limit fallback:")
print("-" * 50)
found_count = 0
for sheetname, name in results:
    if name:
        print(f"{sheetname} -> {name}")
        found_count += 1
    else:
        # Check if the sheet has any data at all
        ws = wb[sheetname]
        non_empty = sum(1 for row in ws.iter_rows(values_only=True) for val in row if val is not None)
        if non_empty == 0:
            print(f"{sheetname} -> [EMPTY SHEET]")
        else:
            print(f"{sheetname} -> [NOT FOUND]")

print("-" * 50)
print(f"Total found: {found_count} / {len(feuil_sheets)}")
