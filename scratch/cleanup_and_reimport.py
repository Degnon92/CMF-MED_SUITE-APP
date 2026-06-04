"""
NETTOYAGE ET RÉ-IMPORTATION ROBUSTE DES FEUILLES Feuil2(X)
1. Supprime les factures avec de faux noms (comme 'CLINIQUE MERCY FIAT') de bills_db.json.
2. Utilise des filtres de nom plus précis (excluant la clinique et les termes médicaux).
3. Lit les vrais noms de patients (avec support du casse mixte et nettoyage des parenthèses).
4. Sauvegarde la base de données propre.
"""
import json
import sys
import os
import re
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_main = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")
bills_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "bills_db.json")

with open(bills_db_path, encoding='utf-8') as f:
    bills_db = json.load(f)

print(f"DB actuelle: {len(bills_db)} factures")

# ─────────────────────────────────────────────────────────────────────────────
# MOTS REJETÉS ET FILTRAGE DE NOMS
# ─────────────────────────────────────────────────────────────────────────────
REJECT_WORDS = {
    # Infos Clinique / En-têtes
    'clinique', 'mercy', 'fiat', 'medecine', 'générale', 'generale', 'specialites', 'spécialités', 'médicales', 'medicales', 'chirurgicales', 'laboratoire',
    # Actes médicaux
    'chirurgien', 'anesthesiste', 'anesthésiste', 'infirmier', 'kinésithérapeute', 'kinesitherapeute', 'radiologue', 'medecin', 'docteur',
    # Fournitures / forfaits
    'forfait', 'materiel', 'matériels', 'implant', 'implants', 'amos', 'clou', 'gamma', 'plaque', 'vis', 'broche', 'sonde',
    # Actes chirurgicaux
    'chirurgical', 'chirurgicale', 'ostheosynthese', 'osteosynthese', 'nephrolithotomie', 'ureteroscopie', 'arthroscopie', 'osteotomie',
    'ostheosynthèse', 'percutanee', 'percutanée', 'nlpc',
    # Colonnes / Mots-clés de facturation
    'unitaire', 'quantite', 'quantité', 'designation', 'désignation', 'montant', 'total', 'principal', 'complement', 'complément',
    'acte', 'actes', 'cs', 'echographie', 'scanner', 'radio',
    # Anatomie
    'femur', 'fémur', 'tibia', 'humerus', 'humérus', 'cubitus', 'ulna', 'fibula', 'cheville', 'genou', 'hanche', 'épaule', 'poignet',
    'tibiale', 'fémorale', 'femorales',
    # Divers
    'athemy', 'kinésithérapies', 'kado', 'seance', 'séance', 'intervention', 'diagnostic', 'assurance', 'labo', 'biologie', 'client',
    'patient', 'nom', 'prenom', 'page', 'feuil', 'sheet', 'caisse', 'loto', 'lotto', 'football', 'club', 'sport', 'coton',
    'proforma', 'définitif', 'definitif', 'détail', 'detail', 'assur', 'prive', 'ascoma', 'sanlam', 'sunu', 'nsia', 'olea', 'gras', 'savoye',
    # Médicaments / Hospit
    'médicaments', 'medicaments', 'consommables', 'consommable', 'pharmacie', 'chambre', 'hospitalisation', 'hospit',
    # Placeholders / Templates
    'nom..', 'prenom..', 'age..', 'diagnostic..', 'horaires..', 'date..', 'séances..', 'seances..', 'seances'
}

def clean_name(s):
    if not s: return ""
    # Enlever le contenu entre parenthèses, ex: "(ASSUR SANLAM)", "(2)"
    s = re.sub(r'\(.*?\)', '', s)
    # Enlever les espaces doubles
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def is_real_patient_name(s):
    """Filtre strict pour valider si une chaîne est un vrai nom de patient."""
    s = clean_name(s)
    if not s or len(s) < 5 or len(s) > 55: return False
    
    words = s.split()
    if len(words) < 2: return False
    
    # Vérifier chaque mot
    for w in words:
        wl = w.lower().strip('():,-..')
        wl = re.sub(r'\.+$', '', wl) # Supprimer les points de fin
        if not wl: continue
        if wl in REJECT_WORDS: return False
        if len(wl) == 1 and wl not in 'abcdefghijklmnopqrstuvwxyz': return False
    
    # Pas de colonnes d'intervention
    if 'intervention:' in s.lower(): return False
    if ':' in s and not re.match(r'^[A-Z][a-z]+:', s): return False
    
    # Lettres dans chaque mot
    for w in words:
        alpha = sum(1 for c in w if c.isalpha())
        if alpha < 2: return False
        
    # Ratio d'alphabétique
    alpha_total = sum(1 for c in s if c.isalpha() or c in ' -\'')
    if alpha_total / len(s) < 0.70: return False
    
    return True

# ─────────────────────────────────────────────────────────────────────────────
# ÉTAPE 1 : NETTOYER LA BASE DES FAUX NOMS PRÉCÉDEMMENT IMPORTÉS
# ─────────────────────────────────────────────────────────────────────────────
# On enlève CLINIQUE MERCY FIAT et les autres mauvais imports détectés.
# On conserve bien sûr TADOGBE CLAUDE et SEDJAME FREJUS qui sont des vrais patients.
FAUX_NOMS = {
    'CHIRURGIEN', 'ANESTHESISTE', 'FORFAIT', 'ACTE', 'ECHOGRAPHIE',
    'PERCUTANEE', 'CLOU', 'TIBIALE', 'ULNA', 'FEMUR', 'VALGISATION',
    'GAMMA', 'NLPC', 'CS', 'ATHEMY', 'P', 'SEANCES', 'CLINIQUE',
    'OSTHEOSYNTHESE', 'OSTEOTOMIE', 'NEPHROLITHOTOMIE', 'FEUIL', 'FEUIL2', 'UNITEVA'
}

CONFIRMED_BAD = {
    'CHIRURGIEN PRINCIPAL', 'P UNITAIRE', 'CLINIQUE MERCY FIAT',
    'ACTE CHIRURGICAL', 'ECHOGRAPHIE ABDOMINALE', 'CS ANESTHESISTE',
    'PERCUTANEE DROITE', 'CLOU GAMMA STANDARD', 'ULNA OU CUBITUS',
    'DU FEMUR', 'DE VALGISATION K120', 'ATHEMY KINESITHERAPIES KADO',
    'FORFAIT MATERIELS UROLOGIQUES', 'AMOS CLOU FEMUR',
    'N SEANCES'
}

cleaned = []
removed = []
for b in bills_db:
    nom = (b.get('patientNom') or '').upper().strip()
    prenom = (b.get('patientPrenom') or '').upper().strip()
    full = f"{nom} {prenom}".strip()
    
    bad = False
    if nom in FAUX_NOMS or nom.startswith('FEUIL') or 'FEUIL' in full or 'UNITEVA' in full:
        bad = True
    for bad_name in CONFIRMED_BAD:
        if bad_name in full:
            bad = True
            break
    if 'intervention' in nom.lower() or 'intervention' in prenom.lower():
        bad = True
        
    if bad:
        removed.append(f"{b['id']} | {full} ({b.get('date')})")
    else:
        cleaned.append(b)

print(f"\n🗑️  Factures avec faux noms supprimées: {len(removed)}")
for r in removed:
    print(f"   - {r}")
print(f"DB après nettoyage: {len(cleaned)} factures")

# ─────────────────────────────────────────────────────────────────────────────
# ÉTAPE 2 : RE-IMPORTATION DES ONGLETS AVEC LE NOUVEAU PARSER
# ─────────────────────────────────────────────────────────────────────────────
def cell_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return int(val)
    try: return int(float(str(val).replace(' ','').replace('\xa0','').replace(',','.')))
    except: return 0

def parse_date_val(val):
    if not val: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2: y = '20' + y
        try: return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except: return None
    return None

def map_insurance(text):
    if not text: return "PRIVE"
    t = text.lower()
    if "sanlam" in t: return "SANLAM"
    if "ascoma" in t: return "ASCOMA"
    if "nsia" in t: return "NSIA"
    if "africaine" in t or "afg" in t: return "AFG"
    if "loto" in t or "lotto" in t: return "LOTTO_FOOTBALL_CLUB"
    if "coton" in t: return "COTON_SPORT"
    if "olea" in t: return "OLEA"
    if "gras" in t or "savoye" in t: return "GRAS SAVOYE"
    return "PRIVE"

def split_nom_prenom(full):
    full = re.sub(r'\s+', ' ', full.strip())
    parts = full.split(' ')
    if len(parts) == 1: return full.upper(), ""
    nom_parts, prenom_parts = [], []
    for p in parts:
        if p.isupper() and len(p) > 1: nom_parts.append(p)
        else: prenom_parts.append(p)
    if not nom_parts: return parts[0].upper(), ' '.join(parts[1:])
    if not prenom_parts: return parts[0].upper(), ' '.join(parts[1:])
    return ' '.join(nom_parts).upper(), ' '.join(prenom_parts)

def find_patient_name_strict(ws):
    """Recherche le nom de façon précise et robuste."""
    all_rows = list(ws.iter_rows(values_only=True, max_row=80))
    
    # PRIORITÉ 1 : Cellule contenant explicitement "Patient:" ou "Nom:"
    for ri, row in enumerate(all_rows[:20]):
        for ci, val in enumerate(row):
            if not val or not isinstance(val, str): continue
            s = val.strip()
            sl = s.lower()
            
            if re.match(r'patient\s*:', sl, re.I):
                candidate = s.split(':', 1)[1].strip()
                candidate = candidate.split('\n')[0].strip()
                cleaned_cand = clean_name(candidate)
                if is_real_patient_name(cleaned_cand):
                    return cleaned_cand
            
            if re.match(r'^(?:patient|nom\s+du\s+patient|nom\s*(?:et\s+pr[eé]nom)?)\s*:?\s*$', sl):
                for next_ci in range(ci+1, min(ci+4, len(row))):
                    nv = row[next_ci]
                    if nv and isinstance(nv, str):
                        cleaned_cand = clean_name(nv)
                        if is_real_patient_name(cleaned_cand):
                            return cleaned_cand
                for nri in range(ri+1, min(ri+3, len(all_rows))):
                    if len(all_rows[nri]) > ci:
                        nv = all_rows[nri][ci]
                        if nv and isinstance(nv, str):
                            cleaned_cand = clean_name(nv)
                            if is_real_patient_name(cleaned_cand):
                                return cleaned_cand
    
    # PRIORITÉ 2 : Recherche de toute cellule de texte valide dans les 8 premières lignes
    # (limité à 8 lignes pour éviter de lire les désignations de prestations plus bas)
    for ri, row in enumerate(all_rows[:8]):
        for val in row:
            if not val or not isinstance(val, str): continue
            cleaned_cand = clean_name(val)
            if is_real_patient_name(cleaned_cand):
                return cleaned_cand
                
    return None

# Indexation DB pour les doublons
def normalize(s):
    return re.sub(r'\s+', ' ', str(s).strip().upper())

db_by_nom = {}
for b in cleaned:
    n = normalize(b.get('patientNom',''))
    p = normalize(b.get('patientPrenom',''))
    full = f"{n} {p}".strip()
    db_by_nom.setdefault(full, []).append(b)

def already_in_db(nom, prenom, date):
    n = normalize(nom); p = normalize(prenom)
    full = f"{n} {p}".strip()
    if full in db_by_nom:
        for b in db_by_nom[full]:
            if b.get('date') == date: return True
    parts = [w for w in n.split() if len(w) > 2]
    for key in db_by_nom:
        if len(parts) >= 2 and all(w in key for w in parts[:2]):
            for b in db_by_nom[key]:
                if b.get('date') == date: return True
    return False

# Trouver l'ID max
max_id = 0
for b in cleaned:
    bid = b.get('id', '')
    m = re.search(r'(\d+)$', bid)
    if m: max_id = max(max_id, int(m.group(1)))
next_id = max_id + 1

print(f"\n--- Ré-importation des feuilles Feuil(X) ---\n")
wb = openpyxl.load_workbook(excel_main, data_only=True)
feuil_sheets = [s for s in wb.sheetnames if re.match(r'^feuil\d+(\s*\(.*\))?$', s.lower().strip())]

new_bills = []
cant_find = []

for sheetname in feuil_sheets:
    try:
        ws = wb[sheetname]
        all_rows = list(ws.iter_rows(values_only=True, max_row=100))
        if not all_rows: continue
        
        patient_name = find_patient_name_strict(ws)
        
        if not patient_name:
            # Vérifier si la feuille est vide
            non_empty_cells = sum(1 for r in all_rows for cell in r if cell is not None)
            if non_empty_cells > 0:
                cells_preview = []
                for row in all_rows[:12]:
                    for val in row:
                        if val and isinstance(val, str) and len(val.strip()) > 3:
                            cells_preview.append(val.strip()[:60])
                            if len(cells_preview) >= 4: break
                    if len(cells_preview) >= 4: break
                cant_find.append({'sheet': sheetname, 'preview': cells_preview})
            continue
            
        # Extraire date, prestations et montant
        date_source = None
        items = []
        total_source = 0
        insurance_txt = ""
        intervention = ""
        diagnostic = ""
        header_row = -1
        bill_type = "PROFORMA"
        k_code = ""
        
        # Parcourir pour en-têtes
        for ri, row in enumerate(all_rows[:60]):
            for ci, val in enumerate(row):
                if isinstance(val, datetime):
                    if not date_source: date_source = val.strftime('%Y-%m-%d')
                    continue
                if not val or not isinstance(val, str): continue
                s = val.strip(); sl = s.lower()
                if re.match(r'(client|assurance)\s*:', sl, re.I): insurance_txt = s.split(':',1)[1].strip()
                if re.match(r'intervention\s*:', sl, re.I): intervention = s.split(':',1)[1].strip()
                if re.match(r'diagnostic\s*:', sl, re.I): diagnostic = s.split(':',1)[1].strip()
                if 'cotonou' in sl and not date_source:
                    d = parse_date_val(s)
                    if d: date_source = d
                if any(k in sl for k in ['définitif','definitif']): bill_type = "DEFINITIF"
                elif any(k in sl for k in ['détail assur','detail assur']): bill_type = "DETAIL_ASSUR"
                km = re.search(r'\b(KC?\s*\d+)\b', s)
                if km: k_code = km.group(1)
                if any(k in sl for k in ['désignation','designation','actes','libellé','prestation']):
                    if header_row == -1: header_row = ri

        if not date_source:
            for row in all_rows[:60]:
                for val in row:
                    if isinstance(val, datetime): date_source = val.strftime('%Y-%m-%d'); break
                if date_source: break
                
        # Items
        if header_row >= 0:
            for row in all_rows[header_row+1: header_row+60]:
                if not row or row[0] is None: continue
                d0 = str(row[0]).strip()
                if not d0: continue
                if any(k in d0.lower() for k in ['total','caisse','à payer','a payer','net à','reste','solde']):
                    for v in row:
                        n = cell_num(v)
                        if n > 10000: total_source = max(total_source, n)
                    break
                price = cell_num(row[2] if len(row) > 2 else None)
                qty = cell_num(row[1] if len(row) > 1 else None) or 1
                subtot = cell_num(row[3] if len(row) > 3 else None) or (price * qty)
                if price > 0 and len(d0) > 2:
                    items.append({'name': d0[:100], 'price': price, 'qty': qty, 'subtotal': subtot})
                    if subtot: total_source += subtot
                    
        # Total direct
        for row in all_rows:
            for ci, val in enumerate(row):
                if not val or not isinstance(val, str): continue
                if any(k in val.lower() for k in ['total général','total general','à payer','a payer','net à']):
                    for v in row[ci:]:
                        n = cell_num(v)
                        if n > 10000: total_source = n; break
                        
        if not items and total_source == 0: continue
        
        nom, prenom = split_nom_prenom(patient_name)
        date = date_source or '?'
        
        if already_in_db(nom, prenom, date):
            print(f"⏭️  [{sheetname}] → {nom} {prenom} ({date}) déjà en DB")
            continue
            
        insurance = map_insurance(insurance_txt)
        coverage = 80 if insurance != "PRIVE" else 0
        part_assurance = int(total_source * coverage / 100) if insurance != "PRIVE" else 0
        part_patient = total_source - part_assurance
        
        print(f"✅ [{sheetname}] → {nom} {prenom} | {total_source:,} FCFA | {date}")
        
        items_final = items if items else [{'name': 'Frais cliniques', 'price': total_source, 'qty': 1, 'subtotal': total_source}]
        new_bills.append({
            "id": f"BILL-CMF-IMP-{next_id:04d}",
            "reference": f"MF-{bill_type[:3]}-IMP-{next_id:04d}",
            "patientNom": nom, "patientPrenom": prenom,
            "type": bill_type, "customTitle": "",
            "insurance": insurance, "coverage": coverage, "matricule": "",
            "diagnostic": diagnostic, "intervention": intervention, "kCode": k_code,
            "showDiag": bool(diagnostic), "showInterv": bool(intervention),
            "showSig": True, "showCachet": True, "useSplit": insurance != "PRIVE",
            "items": items_final, "grossTotal": total_source,
            "discountPct": 0, "reductionAmount": 0, "discountedTotal": total_source,
            "discountType": "PERCENT", "discountValue": 0,
            "partAssurance": part_assurance, "partPatient": part_patient,
            "paymentMethod": "TIERS_PAYANT" if insurance != "PRIVE" else "CASH",
            "amountPaidPatient": part_patient, "balancePatient": 0,
            "status": "RÉGLÉ", "date": date, "patientAge": "", "diagnosis": diagnostic
        })
        next_id += 1

    except Exception as e:
        print(f"❌ [{sheetname}] ERREUR: {e}")

# SAUVEGARDER
all_bills = cleaned + new_bills
with open(bills_db_path, 'w', encoding='utf-8') as f:
    json.dump(all_bills, f, ensure_ascii=False, indent=2)

print(f"\n{'='*70}")
print(f"RÉSULTAT FINAL")
print(f"{'='*70}")
print(f"🗑️  Faux noms supprimés    : {len(removed)}")
print(f"✅ Nouvelles factures      : {len(new_bills)}")
print(f"DB: {len(bills_db)} → {len(all_bills)} factures")

if cant_find:
    print(f"\n❓ Feuilles non-vides sans nom trouvable ({len(cant_find)}):")
    for e in cant_find:
        print(f"  [{e['sheet']}] → {e['preview']}")
print("\n✅ Base de données sauvegardée et nettoyée.")
