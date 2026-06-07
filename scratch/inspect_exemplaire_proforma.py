"""
inspect_exemplaire_proforma.py
Inspecte la structure de l'EXEMPLAIRE PROFORMA principal
pour comprendre tous les onglets et colonnes.
"""
import sys
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from pathlib import Path

EXCEL_FILE = Path(r"C:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx")

print(f"\n=== Inspection de : {EXCEL_FILE.name} ===\n")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"Onglets : {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"ONGLET : '{sheet_name}' ({ws.max_row} lignes x {ws.max_column} colonnes)")
    print('='*70)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), 1):
        if any(v is not None for v in row):
            vals = []
            for v in row:
                if v is not None:
                    s = str(v)
                    if len(s) > 35:
                        s = s[:35] + '...'
                    vals.append(s)
                else:
                    vals.append('')
            print(f"  L{row_idx:2d}: {' | '.join(v for v in vals)}")
