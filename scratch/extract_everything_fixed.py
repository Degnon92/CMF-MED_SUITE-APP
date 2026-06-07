import os
import re
import glob
import json
import openpyxl
import docx
from docx.oxml.ns import qn

workspace = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
app_dir = os.path.join(workspace, "MercyFiatMedSuiteDesktop")

print("Démarrage de l'extraction de secours de la base de données...")

# -------------------------------------------------------------
# 1. ORIGINAUX DE CONFIANCE (HARDCODÉS)
# -------------------------------------------------------------
orig_diagnoses = [
    "CAL VICIEUX ARTICULAIRE DES PLATEAUX TIBIAUX GAUCHES",
    "DOULEUR CHRONIQUE ,TENDINITE",
    "DOULEUR CHRONIQUE,TENDINITE",
    "Disjonction Acromio -clavi",
    "Douleur Méniscale du genou droit",
    "ENTORSE GRAVE DU FEMUR",
    "EVENTRATION MEDIANE SUS  ET SOUS OMBILICALE",
    "Embrochage du 5e Doigt Droit",
    "FR JAMBE",
    "FRACTURE SOUS TROCHANTERO DIAPHYSAIRE",
    "Fracture du col du fémur",
    "Fracture du quart latéral de la clavicule droite",
    "Fracture ouverte de l'olecrane droite + Luxation tête radiale",
    "GONALGIE GAUCHE/BLOCAGE A LA FLEXION EXTENSION",
    "GONARTHROSE  AVEC DESTRUCTION DE CONDYLE INTERNE",
    "Genouthrose bilat",
    "Hypertrophie prostatique bénigne",
    "INFECTION DU SITE OPERATOIRE TARDIVE SUR CLOU GAMMA LONG DU FEMUR GAUCHE",
    "LAXITE INTERNE DU GENOU GAUCHE",
    "LUXATION NEGLIGEE DU COUDE DROIT",
    "LUXATION RECIDIVANTE DE L'EPAULE",
    "Lésion du ménisque latéral du genou droit",
    "Lésion du ménisque latéral du genou gauche",
    "MALPOSITION D'IMPLANT SUR CLE",
    "Macrolithiase rénale droite obstructive",
    "ONGLE INCARNE DES DEUX HALLUX",
    "OSTEOSYNTHESE DE LA CLAVICULE GAUCHE",
    "OSTEOTOMIE DU FEMUR",
    "PARAGE ET TENORAPHIE  EXTENSEUR DU GROS  ORTEIL",
    "PERITONITE PAR PERFORATION SIGMOIDIENNE",
    "Plaie de pied",
    "Rupture partielle du ligament croisé antérieur ;",
    "SENSATION DE BLOCAGE DU GENOU DROIT:SUSPICION DE LESION MENISCALE",
    "SUBLUXATION NEGLIGEE MP DU POUCE DROIT",
    "Syndrome de blocage du genou gauche",
    "TRAUMATISME DU MEMBRE PELVIEN GAUCHE",
    "entorse chronique du ligament collatéral interne du genou droit",
    "entorse du ligament latéral externe de la cheville droite",
    "entorse grave du genou droit",
    "fracture ancienne de la malléole latérale gauche",
    "fracture associant une fracture fermée du fémur droit, une fracture du bassin et de multiples der...",
    "fracture bifocale du fémur gauche",
    "fracture comminutive médio-diaphysaire ouverte type 2 de GUSTILLO du fémur gauche avec perte de s...",
    "fracture de la palette humérale droite",
    "fracture de l’avant-bras droit",
    "fracture de l’extrémité distale de la jambe, avec séquestre osseux au niveau du foyer de fracture",
    "fracture des deux os de la jambe droite et une fracture du poignet gauche",
    "fracture des deux os de la jambe gauche et un traumatisme oculaire gauche grave",
    "fracture des deux os de l’avant-bras gauche",
    "fracture des plateaux tibiaux droits et une fracture de la mandibule",
    "fracture des épines tibiales",
    "fracture du col du 5ème métacarpien droit",
    "fracture du col du 5ème métacarpien droit le 18 avril 2025",
    "fracture du fémur distal droit",
    "fracture du genou droit suite à un accident de la circulation survenu le 21 janvier 2024",
    "fracture déplacée de la clavicule gauche survenue à la suite d’un accident de la circulation le 1...",
    "fracture fermée de l’extrémité supérieure du fémur gauche, une fracture fermée des deux (02) os d...",
    "fracture fermée du fémur gauche",
    "fracture fermée du plateau tibial médial gauche, traitée avec une ostéosynthèse par plaque vissée...",
    "fracture isolée du radius gauche et une fracture céphalo-tubérositaire de l’humérus droit",
    "fracture ostéo-chondrale de la patella gauche",
    "fracture ouverte de la cheville gauche type 2 de GUSTILLO ET ANDERSON",
    "fracture ouverte de la jambe droite",
    "fracture ouverte de type II du 1/3 distal des deux os de la jambe gauche",
    "fracture ouverte supracondylienne du fémur gauche et une fracture du tibia proximal gauche",
    "fracture ouverte type IIIA des deux os de la jambe droite",
    "fracture pertrochantérienne",
    "fracture pertrochantérienne droite, une facture de la diaphyse fémorale sous-jacente et une disjo...",
    "fracture sous-trochantéro-diaphysaire double spiroïde déterminant trois segments s’étendant jusqu...",
    "fracture sus et intercondylienne gauche",
    "fracture très déplacée de l’extrémité distale du radius et de l’ulna gauche",
    "fracture équivalent bimalléolaire droite",
    "lésion méniscale latérale gauche",
    "lésion ostéochondrale fémorale gauche",
    "lésion traumatique (confère compte rendu)",
    "rupture ancienne de ce ligament",
    "rupture complexe transfixante de la corne postérieure du ménisque latéral ;",
    "rupture complète du LCA et une fissure du ménisque médial",
    "rupture complète du ligament croisé antérieur et fissure du ménisque médial",
    "rupture complète du ligament croisé antérieur et une fissure du ménisque médial",
    "rupture complète du ligament croisé antérieur plus fissure corne postérieure ménisque latéral sur...",
    "rupture du Ligament Croisé antérieur du genou",
    "rupture du Ligament Croisé antérieur du genou droit",
    "rupture du clou dans le fragment distal",
    "rupture du ligament croisé antérieur",
    "rupture du ligament croisé antérieur droit associée à une fissure de la corne méniscale postérieu...",
    "rupture du ligament croisé antérieur et une fissure de la corne postérieure du ménisque médial du...",
    "rupture du ligament croisé antérieur plus fissure verticale transfixiante de la corne postérieure...",
    "rupture du ménisque médial droit",
    "rupture du ménisque médial et du ménisque latéral, associée à une rupture incomplète du ligament ...",
    "rupture du tendon patellaire gauche",
    "rupture en anse de seau de la corne antérieure du ménisque latéral, ainsi qu’une lésion ostéochon...",
    "rupture incomplète du ligament croisé antérieur ;",
    "rupture incomplète du ligament croisé antérieur plus fissure verticale transfixiante de la corne ...",
    "rupture méniscale médiale instable",
    "rupture quasi complète du ligament croisé antérieur",
    "rupture totale du 1/3 moyen du ligament croisé antérieur ;",
    "rupture verticale de la corne postérieure du ménisque latéral ;",
    "rupture verticale à la partie moyenne du ménisque latéral du genou gauche grade 3 de Stoller",
    "syndrome douloureux du ménisque médial droit",
    "syndrome infectieux",
    "tendinite d’insertion tibiale du fascia lata",
    "tendinite d’insertion tibiale du tendon patellaire droit",
    "tendinite patellaire"
]

orig_interventions = [
    "ABLATION CLOU FEMUR GAUCHE",
    "ABLATION CLOU GAMMA LONG DROIT",
    "ABLATION DE MATERIEL + OSTEOSYNTHESE PAR VIS + TIGE + CAGE L3 L4 L5",
    "ABLATION DE MATERIEL DE  CHEVILLE",
    "ABLATION DE MATERIEL FEMUR DISTAL 0,6 K150",
    "ABLATION MATERIEL D'OSTEOSYNTHESE DE L'HUMERUS DROIT",
    "AMOS (K80)",
    "AMOS + ALESAGE DE PROPRETE + FISTULECTOMIE + LAVAGE CHIRURGICAL PULSE + ANTIBIOTIQUE",
    "AMOS CLOU GAMMA LONG K90",
    "ARTHRODESE DU GENOU",
    "Ablation  Osteosynthese Humerus proximal",
    "Ablation de Matériel d'Osthéosynthèse Fémur",
    "Ablation de Sonde double J par urétéroscopie",
    "Ablation de clou tibial  droit",
    "Ablation de matériel d'ostéosynthèse (AMOS)",
    "Ablation du clou fémoral droit",
    "Arthroscopie",
    "Arthroscopie  Ménisectomie Genou (D)",
    "Arthroscopie diagnostique & debridement du genou",
    "Arthroscopie diagnostique , thérapeutique du genou droit",
    "Arthroscopie diagnostique et Ménisectomie",
    "Arthroscopie diagnostique et Ménisectomie du genou (D)",
    "Arthroscopie diagnostique et ménisectomie des 2 genoux",
    "Arthroscopie diagnostique et therapeutique du genou",
    "Arthroscopie diagnostique et therapeutique du genou (G)",
    "Arthroscopie diagnostique et therapeutique du genou Droit",
    "Arthroscopie diagnostique et thérapeutique du genou",
    "Arthroscopie diagnostique et thérapeutique du genou droit",
    "Arthroscopie diagnostique reconstruction  LCA et Ménisectomie",
    "Arthroscopie diagnostique reconstruction LCA et Ménisectomie",
    "Arthroscopie diagnostique;therapeutique du genou",
    "Arthroscopie du genou et Méniscectomie médiale",
    "Arthrose de la cheville en deux K180",
    "CHANGEMENT PROTHESE TOTALE DE LA HANCHE K220",
    "CIRCONCISION",
    "CLOU GAMMA LONG K180",
    "CURE DE VARICOCELE BILATERALE",
    "CURE HERNIAIRE GAUCHE",
    "CURE HERNIAIRE GAUCHE + CURE DE VARICOCELE BILATERALE",
    "CURE PAR PROTHESE RETROMUSCULAIRE",
    "CURETAGE + LAVAGE + PLASTIE CUTANEE DE RECOUVREMENT",
    "Correction Orthopédique sous Sédation",
    "Cure Hernie Ombilicale",
    "Cure d'hydrocèle ou de varicocèle",
    "DEPOSE PIH DROITE",
    "ENCLOUAGE DU FEMUR +HAUBANAGE DE LA PATELLA",
    "EPIPHYSIODESE TIBIALE PROXIMALE INTERNE BILATERAL HANSITAIRE",
    "EXERESE LIPOME AVANT BRAS GAUCHE",
    "Fracture Ancienne Trochanterienne Droite/Clou Gamma",
    "GENU VARUM GAUCHE",
    "GREFFE CUTANEE K40",
    "INTERVENTION DE VOLKMANN - LAVAGE - DRAINAGE",
    "Iso Précoce sur Matériel d'ostéosynthése/",
    "Kyste de la main",
    "Kyste du poignet gauche",
    "Ligamentoplastie du LCA par DIDT",
    "MONTEE DE SONDE JJ",
    "NEPHROLITHOTOMIE  PERCUTANEE (NLPC) DROITE",
    "NEUROLYSE ULNAIRE ET MEDIAN DROIT",
    "Nephrolithotomie percutanée (NLPC) droite",
    "OSTEOCLASIE   ET DOUBLE PLAQUE VISSEE EN T ET EN \"L\" LCP",
    "OSTEOSYNTHESE DU RADIUS GAUCHE - K80",
    "OSTEOTOMIE DU FEMUR",
    "OSTHEOSYNTHESE",
    "OSTHEOSYNTHESE CLOU GAMMA STANDARD",
    "OSTHEOSYNTHESE ULNA OU CUBITUS",
    "Osteosynthese Humerus proximal",
    "Osteosynthese de l'olecrane par plaque LCP",
    "Osteosynthese de l'olecrane par plaque vissée et Embrochage huméro-radial",
    "Osteosynthese des plateaux tibiaux",
    "Osteosynthese par Clou GAMMA LONG",
    "Osteosynthese par clou GAMMA Standard",
    "Osteosynthese par plaque en crochet",
    "Osthéosynthese du fémur",
    "Ostéosynthese d'une fracture de la cheville",
    "Ostéosynthese d'une fracture récente des deux os",
    "Ostéosynthèse d'une fracture diaphysaire ancienne du tibia  K180",
    "Ostéosynthèse du tibia ou du fémur",
    "Ostéosynthèse par plaque 'L' LPC 4.5/5.0",
    "Ostéosynthèse radius",
    "PROTHESE TOTALE DE LA HANCHE",
    "PROTHESE TOTALE DES DEUX HANCHES",
    "Prothese Totale du Genou Gauche",
    "Prothese intermediaire de la hanche",
    "Prothese intermédiaire hanche",
    "Prothese totale du genou",
    "Prothèse Totale  de la hanche",
    "Prothèse Totale de la Hanche (PTH)",
    "Prothèse Totale du Genou (PTG)",
    "Prothèse intermédiaire Hanche /reprise",
    "Prothèse intermédiaire de hanche",
    "REDUCTION SANGLANTE + EMBROCHAGE",
    "REPRISE OSTEOTOMIE DU FEMUR",
    "REPRISE Ostéosynthèse radius",
    "RESECTION TRANSURETRALE DE LA PROSTATE (RTUP)",
    "Résection Transurétrale de la Prostate (RTUP)",
    "SYNTHESE D'UNE  PSEUDOARTHROSE DU FEMUR DISTAL(K150)",
    "TIBIALE DE VALGISATION K120",
    "Transfert du tibial antérieur sur le 5ᵉ métatarsien + ténotyplasie calcanéen bilatéral",
    "URETEROSCOPLASTIE + SONDE JJ",
    "Ureteroscopie droite + Montée de sonde JJ",
    "VISSAGE DES 2 TÊTES FEMORALES IN SITU BILATERAL"
]

# -------------------------------------------------------------
# 2. FONCTIONS DE NETTOYAGE & NORMALISATION
# -------------------------------------------------------------
def clean_patient_name(name):
    if not name or not isinstance(name, str):
        return ""
    clean = name.strip()
    prefix_pattern = re.compile(
        r'^(?:CERTIFICAT\s+DE\s+MARIAGE|CERTIFICAT\s+DE\s+NON\s+BEGAIEMENT|CERTIFICAT\s+MEDICAL\s+INITIAL\s+DE\s+CONSTATATION\s+DE\s+COUPS\s+ET\s+BLESSURES|CERTIFICAT\s+MEDICAL\s+POUR\s+COUPS\s+ET\s+BLESSURES|CERTIFICAT\s+MEDICAL\s+DE\s+L[’\']ETAT\s+ACTUEL|CERTIFICAT\s+MEDICAL\s+INITIAL|CERTIFICAT\s+MEDICAL|CERTIFICAT\s+MED\s+INITIAL|CERTIFICAT\s+DE\s+REPOS|CERTIFICAT\s+DE\s+REPRISE|CERTIFICAT\s+DE\s+GUERISON|CERTIFICAT\s+DE\s+GUÉRISON|CERTIFICAT\s+DE|CERTIFICAT|RAPPORT\s+MEDICAL|RAPPORT\s+DE\s+MONSIEUR|RAPPORT\s+DE\s+MME|RAPPORT\s+DE|RAPPORT\s+D\'HOSPI|RAPPORT\s+D\'HOSPITALISATION|RAPPORT\s+DE\s+CONSULTATION|RAPPORT|CRO\s+MODELE|CRO|CMI|MEDICAL|MED\s+INITIAL|GUERISON\s+DE\s+MONSIEUR|GUERISON\s+DE\s+MME|GUERISON\s+DE|GUERISON|GUÉRISON|DECES\s+DE\s+MONSIEUR|DECES\s+DE\s+MME|DECES\s+DE|DECES|DÉCÈS|D\'HOSPI\s+TYPE|D\'HOSPI|DHOSPI|ATTESTATION\s+DE\s+GUERISON|ATTESTATION\s+DE\s+GUÉRISON|ATTESTATION\s+DE|ATTESTATION)\s+',
        re.IGNORECASE
    )
    old_clean = ""
    while clean != old_clean:
        old_clean = clean
        clean = prefix_pattern.sub("", clean).strip()
        
    split_pattern = re.compile(
        r'(?:CERTIFICAT|JE\s+SOUSSIGN|JE\s+SOUSSIGNE|RAPPORT|DOSSIER|N°|NO\s+DOSSIER|CMI|CRO|CLINIQUE|MÉDECINE|MEDECINE|DIAGNOSTIC|INTERVENTION|CLIENT|ASSURANCE|AFRICAINE|NSIA|ALLIANZ|SUNU|AROO|SAAR|CORIS|FEDAS|MUTUELLE|SÉJOUR|SEJOUR|DATE|OPÉRATOIRE|OPERATOIRE|CERTFICAT|CERTIF|PATIENT|PATIENTE|COTONOU|RUE\s+PAVILLON|TEL\s*\:|E\-MAIL|EMAIL|E\s+MAIL|GUERISON|GUÉRISON|DECES|DÉCÈS)',
        re.IGNORECASE
    )
    match = split_pattern.search(clean)
    if match:
        clean = clean[:match.start()].strip()
        
    clean = re.sub(r'(?:\s+|:)\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    if clean.upper().endswith("AGE") and len(clean) > 5:
        clean = clean[:-3].strip()
        
    clean = re.sub(r'[\s\-\.\,\:\_]+$', '', clean).strip()
    clean = re.sub(r'\s+', ' ', clean)
    clean = re.sub(r'\s+\(?(?:ok|copie|2|3|4|1)\)?$', '', clean, flags=re.IGNORECASE).strip()
    return clean

def clean_clinical_term(term):
    if not term or not isinstance(term, str):
        return ""
    clean = term.strip()
    clean = re.sub(r'^[\s\-\.\,\:\_\•\*\t\?\/]+', '', clean).strip()
    clean = re.sub(r'[\s\-\.\,\:\_\?\/]+$', '', clean).strip()
    clean = clean.replace('\n', ' ').replace('\r', ' ')
    clean = re.sub(r'\s+', ' ', clean).strip()
    if len(clean) < 5 or len(clean) > 120:
        return ""
    upper = clean.upper()
    forbidden = ["JE SOUSSIGN", "CLINIQUE", "TEL :", "EMAIL", "IFU", "RCCM", "ORABANK", "SEME AGUE", "COMPTE RENDU", "RAPPORT DE"]
    if any(f in upper for f in forbidden):
        return ""
    return clean

def map_insurance(text):
    if not text:
        return "PRIVE"
    t = text.lower().strip()
    if "sanlam" in t or "saham" in t:
        return "SANLAM"
    elif "ascoma" in t:
        return "ASCOMA"
    elif "sunu" in t:
        return "SUNU"
    elif "nsia" in t:
        return "NSIA"
    elif "atlantique" in t:
        return "ATLANTIQUE"
    elif "africaine" in t:
        if "sinistre" in t or "accident" in t:
            return "AFRICAINE_SINISTRE"
        return "AFG"
    elif "lotto" in t or "loto" in t:
        return "LOTTO_FOOTBALL_CLUB"
    elif "coton" in t:
        return "COTON_SPORT"
    elif "nobila" in t:
        return "NOBILA"
    elif "gras" in t or "savoye" in t:
        return "GRAS SAVOYE"
    elif "olea" in t:
        return "OLEA"
    elif "fonds" in t or "garantie" in t or "fga" in t:
        return "FONDS_GARANTIE_AUTO"
    elif "transvie" in t:
        return "TRANSVIE"
    elif "sobremap" in t or "sobemap" in t:
        return "SOBEMAP"
    elif "port" in t or "pac" in t:
        return "PORT_AUTONOME_COTONOU"
    return "PRIVE"

# -------------------------------------------------------------
# 3. EXTRACTION COMPLÈTE PAR PARSAGE DE PATIENTS_DB.JSON
# -------------------------------------------------------------
# Nous chargeons patients_db.json qui contient déjà TOUTES les données d'extraction fusionnées de tous les fichiers docx/xlsx de l'utilisateur !
patients_db_path = os.path.join(app_dir, "patients_db.json")
if not os.path.exists(patients_db_path):
    print("Fichier patients_db.json introuvable. Exécutez d'abord l'extracteur initial.")
    exit(1)

with open(patients_db_path, "r", encoding="utf-8") as f:
    patients = json.load(f)

extracted_diagnoses = set(orig_diagnoses)
extracted_interventions = set(orig_interventions)

for p in patients:
    diag = p.get("diagnosis", "")
    interv = p.get("intervention", "")
    
    diag_c = clean_clinical_term(diag)
    interv_c = clean_clinical_term(interv)
    
    if diag_c:
        extracted_diagnoses.add(diag_c)
    if interv_c:
        extracted_interventions.add(interv_c)

# -------------------------------------------------------------
# 4. PARSER L'EXCEL DE NOUVEAU POUR COMPLÉTER LES DÉSIGNATIONS D'ACTES
# -------------------------------------------------------------
print("Scan complémentaire d'EXEMPLAIRE PROFORMA.xlsx pour les actes...")
xlsx_files = []
for r, d, fs in os.walk(workspace):
    if "MercyFiatMedSuiteDesktop" in r or "node_modules" in r or ".git" in r:
        continue
    for file in fs:
        if file.endswith(".xlsx") and not file.startswith("~$"):
            xlsx_files.append(os.path.join(r, file))

for f_path in xlsx_files:
    try:
        wb = openpyxl.load_workbook(f_path, data_only=True)
        for sheetname in wb.sheetnames:
            if "feuil" in sheetname.lower() or "sheet" in sheetname.lower():
                continue
            sheet = wb[sheetname]
            for r in range(8, 55):
                val = sheet.cell(r, 1).value
                if val and isinstance(val, str) and len(val.strip()) > 5:
                    v_str = val.strip()
                    if not any(kw in v_str.lower() for kw in ["total", "caisse", "option", "chambre", "hébergement", "hebergement", "repas", "soins", "acte"]):
                        clean_acte = clean_clinical_term(v_str)
                        if clean_acte:
                            extracted_interventions.add(clean_acte)
    except Exception as e:
        pass

# -------------------------------------------------------------
# 5. CONSOLIDATION & ÉCRITURE DE DATABASE.JS
# -------------------------------------------------------------
all_diags = sorted(list(extracted_diagnoses))
all_intervs = sorted(list(extracted_interventions))

database_js_path = os.path.join(app_dir, "database.js")
if os.path.exists(database_js_path):
    with open(database_js_path, "r", encoding="utf-8") as f:
        db_js_content = f.read()
        
    # Échapper de manière sûre pour le format JavaScript double-quote
    def escape_for_js_string(s):
        # 1. Échapper les antislashs existants d'abord
        s_clean = s.replace('\\', '\\\\')
        # 2. Échapper les guillemets doubles
        s_clean = s_clean.replace('"', '\\"')
        return s_clean

    escaped_diags = [escape_for_js_string(d) for d in all_diags]
    escaped_intervs = [escape_for_js_string(i) for i in all_intervs]

    new_diag_str = "DIAGNOSES: [\n        " + ",\n        ".join(f'"{d}"' for d in escaped_diags) + "\n    ]"
    new_interv_str = "INTERVENTIONS: [\n        " + ",\n        ".join(f'"{i}"' for i in escaped_intervs) + "\n    ]"
    
    updated_content = re.sub(r'DIAGNOSES\s*:\s*\[.*?\]', new_diag_str, db_js_content, flags=re.DOTALL)
    updated_content = re.sub(r'INTERVENTIONS\s*:\s*\[.*?\]', new_interv_str, updated_content, flags=re.DOTALL)
    
    with open(database_js_path, "w", encoding="utf-8") as f:
        f.write(updated_content)
    print(f"Base de données mise à jour avec succès !")
    print(f"Nombre final de diagnostics écrits : {len(all_diags)}")
    print(f"Nombre final d'interventions écrites : {len(all_intervs)}")
else:
    print("Fichier database.js introuvable !")
