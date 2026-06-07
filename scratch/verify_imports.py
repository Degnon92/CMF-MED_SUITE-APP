import json
import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")
bills_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "bills_db.json")

print(f"Lecture Excel: {excel_path}")
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
print(f"Onglets disponibles: {wb.sheetnames[:30]}")
print()

with open(bills_db_path, encoding='utf-8') as f:
    bills = json.load(f)

# Construire un index rapide par nom patient
bills_by_name = {}
for b in bills:
    key = (b.get('patientNom','') + ' ' + b.get('patientPrenom','')).strip().upper()
    if key not in bills_by_name:
        bills_by_name[key] = []
    bills_by_name[key].append(b)

# Parcourir chaque onglet et extraire les 10 premières lignes de données
IGNORE_SHEETS = {'AIDE', 'CONFIG', 'FORFAITS', 'TARIFS', 'LISTE', 'PARAMETRES', 'PARAM', 'NOMENCLATURE', 'Sheet1', 'Feuil1'}

checked = 0
problems = []

for sheet_name in wb.sheetnames:
    if sheet_name.upper() in IGNORE_SHEETS:
        continue
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True, max_row=80))
        if not rows:
            continue

        # Chercher nom patient, date, diagnostic, intervention dans les 30 premières lignes
        patient_name = None
        date_val = None
        diagnostic = None
        intervention = None
        insurance = None
        total = None

        for row in rows[:40]:
            cells = [str(c).strip() if c is not None else '' for c in row]
            line = ' '.join(cells).strip()
            line_upper = line.upper()

            # Nom patient
            for i, cell in enumerate(cells):
                if not cell:
                    continue
                cu = cell.upper()
                if any(kw in cu for kw in ['NOM', 'PATIENT', 'CLIENT']) and i+1 < len(cells) and cells[i+1]:
                    if len(cells[i+1]) > 3 and not any(x in cells[i+1].upper() for x in ['CLINIQUE', 'MERCY', 'FIAT']):
                        if not patient_name:
                            patient_name = cells[i+1].strip()
                elif i == 0 and cell and len(cell) > 4 and cell.replace(' ','').isalpha() and cell.isupper():
                    # Possible nom patient en colonne A
                    if not patient_name and not any(x in cu for x in ['CLINIQUE', 'MERCY', 'FIAT', 'DATE', 'TOTAL', 'DIAGNOSTIC', 'INTERVENTION']):
                        patient_name = cell

            # Date
            for cell in cells:
                if cell and '/' in cell and len(cell) >= 8:
                    import re
                    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', cell)
                    if m and not date_val:
                        d, mo, y = m.groups()
                        if len(y) == 2:
                            y = '20' + y
                        try:
                            date_val = f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
                        except:
                            pass

            # Diagnostic
            for i, cell in enumerate(cells):
                if not cell:
                    continue
                cu = cell.upper()
                if 'DIAGNOSTIC' in cu and i+1 < len(cells) and cells[i+1] and len(cells[i+1]) > 3:
                    if not diagnostic:
                        diagnostic = cells[i+1].strip()

            # Intervention
            for i, cell in enumerate(cells):
                if not cell:
                    continue
                cu = cell.upper()
                if 'INTERVENTION' in cu and i+1 < len(cells) and cells[i+1] and len(cells[i+1]) > 3:
                    if not intervention:
                        intervention = cells[i+1].strip()

        # Si on a au moins le nom, comparer
        if patient_name and len(patient_name) > 3:
            pn_upper = patient_name.upper().strip()
            # Chercher dans la base
            matched = None
            for key, blist in bills_by_name.items():
                # Match partiel
                parts = pn_upper.split()
                if all(p in key for p in parts[:2]) or pn_upper in key or key in pn_upper:
                    matched = blist
                    break

            if matched:
                bill = matched[0]
                db_nom = bill.get('patientNom','')
                db_prenom = bill.get('patientPrenom','')
                db_date = bill.get('date','')
                db_diag = bill.get('diagnostic','') or ''
                db_interv = bill.get('intervention','') or ''

                # Vérifier les champs
                errs = []

                if date_val and db_date and date_val != db_date:
                    errs.append(f"DATE: source={date_val} | DB={db_date}")

                if diagnostic and db_diag:
                    if diagnostic.upper()[:20] not in db_diag.upper() and db_diag.upper()[:20] not in diagnostic.upper():
                        errs.append(f"DIAG: source='{diagnostic[:50]}' | DB='{db_diag[:50]}'")

                if intervention and db_interv:
                    if intervention.upper()[:20] not in db_interv.upper() and db_interv.upper()[:20] not in intervention.upper():
                        errs.append(f"INTERV: source='{intervention[:50]}' | DB='{db_interv[:50]}'")

                if errs:
                    print(f"[DIFFERENCE] {sheet_name} → {db_nom} {db_prenom}")
                    for e in errs:
                        print(f"   {e}")
                    problems.append(sheet_name)
                else:
                    pass  # OK
                checked += 1
            else:
                print(f"[NON TROUVE en DB] Onglet '{sheet_name}' | Patient: '{patient_name}'")
                problems.append(sheet_name)

    except Exception as e:
        print(f"[ERREUR] Onglet '{sheet_name}': {e}")
        continue

print(f"\n=== RÉSUMÉ ===")
print(f"Onglets vérifiés: {checked}")
print(f"Problèmes détectés: {len(problems)}")
if problems:
    print("Onglets avec problème:")
    for p in problems:
        print(f"  - {p}")
