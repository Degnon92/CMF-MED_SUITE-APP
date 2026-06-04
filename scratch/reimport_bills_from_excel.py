"""
reimport_bills_from_excel.py
===========================
Re-importe les données d'assurance et de split depuis EXEMPLAIRE PROFORMA.xlsx
vers bills_db.json en utilisant une extraction de colonnes dynamique et un matching
robuste basé sur le nom, le type et la date de la facture.
"""
import sys, re, json, copy
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

# --- CHEMINS ---
EXCEL_FILE  = Path(r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx")
BILLS_FILE  = Path(r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\MercyFiatMedSuiteDesktop\bills_db.json")
BACKUP_FILE = Path(r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\MercyFiatMedSuiteDesktop\bills_db_backup_before_reimport.json")
REPORT_FILE = Path(r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\MercyFiatMedSuiteDesktop\scratch\reimport_report.txt")

DRY_RUN = '--dry-run' in sys.argv  # Simulation sans modification
print(f"Mode : {'DRY-RUN (simulation)' if DRY_RUN else 'RÉEL (modification de bills_db.json)'}\n")

months_map = {
    "janvier": "01", "fevrier": "02", "février": "02", "mars": "03", "avril": "04",
    "mai": "05", "juin": "06", "juillet": "07", "aout": "08", "août": "08",
    "septembre": "09", "octobre": "10", "novembre": "11", "decembre": "12", "décembre": "12"
}

def clean_num(val):
    if val is None: return 0
    try:
        f = float(val)
        if f.is_integer():
            return int(f)
        return round(f, 2)
    except:
        return 0


def parse_french_date(text):
    if not text: return None
    m_slashes = re.search(r"(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})", text)
    if m_slashes:
        day = f"{int(m_slashes.group(1)):02d}"
        month = f"{int(m_slashes.group(2)):02d}"
        year = m_slashes.group(3)
        return f"{year}-{month}-{day}"
    m_words = re.search(r"le\s+(\d{1,2})\s+([a-zA-Zéûûôâêîñéèàç]+)\s+(\d{4})", text, re.IGNORECASE)
    if m_words:
        day = f"{int(m_words.group(1)):02d}"
        month_name = m_words.group(2).lower()
        year = m_words.group(3)
        month = months_map.get(month_name, "02")
        return f"{year}-{month}-{day}"
    return None

def date_diff_days(d1_str, d2_str):
    if not d1_str or not d2_str: return 9999
    try:
        dt1 = datetime.strptime(d1_str[:10], '%Y-%m-%d')
        dt2 = datetime.strptime(d2_str[:10], '%Y-%m-%d')
        return abs((dt1 - dt2).days)
    except:
        return 9999

def clean_patient_name(name):
    if not name: return ''
    n = name.upper().strip()
    
    n = re.sub(r'\(\d+\)', '', n).strip()
    
    prefixes = [
        r'^POINT\s+DEF\.?', r'^POINT\s+FINAL', r'^POIN\s+FINAL', r'^PT\s+DEF\.?',
        r'^DETAILS?\s+ASS_?', r'^DETAILS?\s+ASS\.?', r'^DETAILS?_?', r'^DETAIL_?',
        r'^DET\.\s+ASS\._?', r'^DET\.\s+ASS\.?', r'^DET\.\s+ASS?', r'^DET_ASS\.?',
        r'^DET_ASS?', r'^DET_?', r'^DET\.?', r'^POI_?', r'^POINT_?', r'^POINT\.?',
        r'^POIN_?', r'^POIN\.?', r'^PT\.?', r'^PT\s+', r'^PT-?', r'^IMPLANT\s+',
        r'^REPRISE_?', r'^PROLONG_?', r'^PROF\s+DEF\s+', r'^PROF\s+', r'^PROFORMA\s+',
        r'^P_', r'^DETAILS?_', r'^PATIENT\s*:\s*', r'^PATIENTE\s*:\s*', r'^PATIENT\s+',
        r'^PATIENTE\s+', r'^M\.', r'^MME\s+', r'^MR\s+', r'^DR\.', r'^DR\s+',
        r'^PT-MOULERO\s+', r'^POINT-HPT-', r'^POINT\s+'
    ]
    
    changed = True
    while changed:
        old_n = n
        for pfx in prefixes:
            n = re.sub(pfx, '', n, flags=re.I).strip()
        if n == old_n:
            changed = False
            
    suffixes = [
        r'_ASCOMA$', r'_AFG$', r'_FGA$', r'_LCA$', r'_AA$', r'-SANLAM$', r'_SINISTRE$',
        r'_SIN$', r'_SINIS$', r'_SINIS\s+AA$', r'_SINIS\s+AFG$', r'_SINI$', 
        r'_SINI_AFG_ELIEZE$', r'_RTUP$', r'_SONDE_URO$', r'_URO$', r'_URO\s+LABO$',
        r'_URO_AA$', r'_URO_DRAGOUN$', r'_ARTHROS$', r'_ARTH$', r'_ARTHROS_SANLAM$',
        r'_CARDIO$', r'_NEUROCHI$', r'_CHIRPED$', r'_DR\s+DJEDOU$', r'_DR\s+DJED$',
        r'_DRQUENUM$', r'_DRQUENUM\s+HPT$', r'_DRQUE$', r'_DRQUE_AA$', r'_COTON$',
        r'_SPORT$', r'_LOTO$', r'_LOTO\s+SANLAM$', r'_DJEDOU$', r'_SINISTRE\s+AFG$',
        r'_LOTO\s+SANLAM$', r'_Sinistre ok$', r'_Sinistre$', r'_Sinistre\s+ok$',
        r'_ Sinistre$', r'_Sinistre$', r'_Sinistre ok$', r'_AA$', r'_AA$', r'_AA$',
        r'_ASK$', r'_ASK\s+\(\d+\)$', r'_ASK$', r'_ASK$', r'_ASK$', r'_ASK$',
        r'\s+\d+$', r'\s+\d+\)$', r'_1$', r'_2$', r'_3$', r'_4$', r'_5$', r'_6$', r'_7$'
    ]
    
    changed = True
    while changed:
        old_n = n
        for sfx in suffixes:
            n = re.sub(sfx, '', n, flags=re.I).strip()
        if n == old_n:
            changed = False
            
    n = re.sub(r'[^A-Z\s\-]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n

def detect_insurer(text):
    if not text: return None
    t = text.lower().strip()
    if 'sinistre' in t or 'accident' in t or 'sini' in t:
        if 'africaine' in t or ' afg' in t or ' aa' in t or '_aa' in t:
            return 'AFRICAINE_SINISTRE'
        if 'sanlam' in t:
            return 'SANLAM_SINISTRE'
        if 'sunu' in t:
            return 'SUNU_SINISTRE_AUTO'
        if 'generale' in t or 'générale' in t or 'gab' in t:
            return 'GENERAL_ASSURANCE_SINISTRE'
        if 'fonds' in t or 'garantie' in t or 'fga' in t:
            return 'FONDS_GARANTIE_AUTO'
            
    if 'sanlam' in t or 'saham' in t:
        return 'SANLAM'
    if 'ascoma' in t or 'asco' in t:
        return 'ASCOMA'
    if 'sunu' in t:
        return 'SUNU'
    if 'nsia' in t:
        return 'NSIA'
    if 'atlantique' in t:
        return 'ATLANTIQUE'
    if 'dayo' in t:
        return 'DAYO'
    if 'nobila' in t:
        return 'NOBILA'
    if 'olea' in t:
        return 'OLEA'
    if 'transvie' in t:
        return 'TRANSVIE'
    if 'gras savoye' in t or 'gras' in t or 'savoye' in t:
        return 'GRAS SAVOYE'
    if 'coton' in t:
        return 'COTON_SPORT'
    if 'lotto' in t or 'loto' in t:
        return 'LOTTO_FOOTBALL_CLUB'
    if 'sobemap' in t:
        return 'SOBEMAP'
    if 'port' in t or 'pac' in t:
        return 'PORT_AUTONOME_COTONOU'
    if 'energie' in t:
        return 'ENERGIE_BASKET_BALL'
    if 'gab' in t or 'générale des assurances' in t or 'generale des assurances' in t:
        return 'GENERAL_ASSURANCE_SINISTRE'
    if 'africaine' in t or 'afg' in t or 'aa' in t:
        return 'AFG'
    return None

def detect_bill_type_from_sheet(sheet_name, rows_text):
    sn = sheet_name.upper()
    rt = rows_text.upper()
    if any(k in sn for k in ['DETAILS', 'DETAIL ASS', 'DET ASS', 'DETAILS ASS']):
        return 'DETAIL_ASSUR'
    if any(k in sn for k in ['POIN', 'POINT', 'DEF ', 'FINAL', 'DEFINITIF']):
        return 'DEFINITIF'
    if any(k in rt for k in ['POINT DEFINITIF', 'POINT FINAL', 'POINT DEF', 'DEFINITIF']):
        return 'DEFINITIF'
    if any(k in rt for k in ['DETAIL ASSURANCE', 'DETAILS ASSURANCE', 'DETAIL ASSU']):
        return 'DETAIL_ASSUR'
    return 'PROFORMA'

def safe(v, maxlen=100):
    if v is None: return ''
    return str(v).strip()[:maxlen]

def map_header_cols(header_row):
    cols = [str(c).strip().lower() for c in header_row]
    mapping = {
        'desig': 0, 'qty': 1, 'price': 2, 'subtotal': 3,
        'part_assurance': [], 'part_patient': None
    }
    
    for idx, c in enumerate(cols):
        if 'designation' in c or 'désignation' in c or 'actes' in c or 'prest' in c:
            mapping['desig'] = idx
            break
    for idx, c in enumerate(cols):
        if 'qt' in c or 'nbre' in c:
            mapping['qty'] = idx
            break
    for idx, c in enumerate(cols):
        if 'unit' in c or 'prix' in c or 'unitaire' in c:
            mapping['price'] = idx
            break
    for idx, c in enumerate(cols):
        if 'montant' in c or 'total' in c or 'clinique' in c:
            if 'assur' not in c and 'pat' not in c and idx > mapping['price']:
                mapping['subtotal'] = idx
                break

    insurance_keywords = [
        'assur', 'tiers', 'fga', 'sanlam', 'saham', 'nsia', 'ascoma', 'sunu', 'olea', 
        'transvie', 'atlantique', 'afg', 'lotto', 'loto', 'coton', 'sobemap', 'port', 'aa', 'gab'
    ]
    for idx, c in enumerate(cols):
        if 'part' in c and any(kw in c for kw in insurance_keywords) and 'pat' not in c:
            mapping['part_assurance'].append(idx)
            
    if not mapping['part_assurance']:
        for idx, c in enumerate(cols):
            if any(kw in c for kw in insurance_keywords) and 'pat' not in c and 'limit' not in c and 'total' not in c:
                mapping['part_assurance'].append(idx)
                
    if not mapping['part_assurance']:
        for idx, c in enumerate(cols):
            if 'assur' in c and 'pat' not in c and 'limit' not in c and 'total' not in c:
                mapping['part_assurance'].append(idx)

    for idx, c in enumerate(cols):
        if 'pat' in c and any(kw in c for kw in ['final', 'apr', 'tot', 'net']):
            mapping['part_patient'] = idx
            break
    if mapping['part_patient'] is None:
        for idx, c in enumerate(cols):
            if 'pat' in c and ('part' in c or 'tot' in c or 'net' in c):
                mapping['part_patient'] = idx
                break
    if mapping['part_patient'] is None:
        mapping['part_patient'] = len(cols) - 1
        
    if not mapping['part_assurance'] and mapping['part_patient'] is not None:
        for idx, c in enumerate(cols):
            if ('limit' in c or 'tot' in c) and 'assur' in c and idx != mapping['part_patient']:
                mapping['part_assurance'].append(idx)
                break
                
    return mapping

def name_score(n1, n2):
    n1 = n1.upper().strip()
    n2 = n2.upper().strip()
    if n1 == n2: return 100
    if n1 in n2 or n2 in n1: return 90
    w1 = set(n1.split())
    w2 = set(n2.split())
    common = w1 & w2
    if len(common) >= 2: return 70 + len(common) * 5
    if len(common) == 1 and len(list(common)[0]) > 4: return 50
    return 0

# 1. Charger les fichiers
print("Chargement de l'EXCEL...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"Total onglets : {len(wb.sheetnames)}")

print("Chargement de bills_db.json...")
with open(BILLS_FILE, 'r', encoding='utf-8') as f:
    bills = json.load(f)
print(f"Factures en DB : {len(bills)}")

excel_sheets = []

# 2. Extraction des données de chaque onglet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_data = {
        'sheet_name': sheet_name,
        'patient_norm': clean_patient_name(sheet_name),
        'patient_raw': sheet_name,
        'insurer': detect_insurer(sheet_name),
        'coverage': 0,
        'bill_type': 'PROFORMA',
        'gross_total': 0,
        'part_assurance_total': 0,
        'part_patient_total': 0,
        'items': [],
        'has_split': False,
        'rows_text': '',
        'date': None
    }
    
    all_rows = list(ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=True))
    rows_text = ' '.join(safe(v) for row in all_rows for v in row if v)
    sheet_data['rows_text'] = rows_text
    sheet_data['bill_type'] = detect_bill_type_from_sheet(sheet_name, rows_text)
    
    # Détecter assureur dans les lignes si pas trouvé dans le nom d'onglet
    if not sheet_data['insurer']:
        for row in all_rows:
            row_str = ' '.join(safe(v) for v in row if v)
            ins = detect_insurer(row_str)
            if ins:
                sheet_data['insurer'] = ins
                break
                
    item_start_row = None
    for r_idx, row in enumerate(all_rows, 1):
        row_str = ' '.join(safe(v) for v in row if v)
        
        # Nom du patient (strict regex avec colon pour éviter les notes de bas de page)
        for v in row:
            if v and isinstance(v, str):
                m_pat = re.search(r'patient\s*:\s*(.+)$', v, re.I)
                if m_pat:
                    extracted = m_pat.group(1).strip()
                    if len(extracted) > 2:
                        sheet_data['patient_norm'] = clean_patient_name(extracted)
                        sheet_data['patient_raw'] = extracted
                        
        # Date
        if not sheet_data['date']:
            if "cotonou" in row_str.lower():
                for v in row:
                    if v and isinstance(v, str) and "cotonou" in v.lower():
                        date_parsed = parse_french_date(v)
                        if date_parsed:
                            sheet_data['date'] = date_parsed
                            
        # Couverture
        m = re.search(r'(\d{2,3})\s*%', row_str)
        if m and not sheet_data['coverage']:
            cov = int(m.group(1))
            if 1 <= cov <= 100:
                sheet_data['coverage'] = cov
                
        # Ligne d'en-tête (au moins 4 colonnes pour pouvoir gérer les layouts simples sans split)
        non_empty = [v for v in row if v is not None]
        if len(non_empty) >= 4 and item_start_row is None:
            if any(k in row_str.lower() for k in ['désignation', 'designation', 'actes', 'prestations']):
                item_start_row = r_idx + 1
                col5 = safe(row[4]) if len(row) > 4 else ''
                col6 = safe(row[5]) if len(row) > 5 else ''
                if any(k in (col5+col6).lower() for k in ['assur', 'patient', 'part', 'tiers']):
                    sheet_data['has_split'] = True
                    
    # Déterminer la couverture par défaut si non trouvée mais assureur connu
    if not sheet_data['coverage'] and sheet_data['insurer']:
        default_rates = {
            'AFRICAINE_SINISTRE': 100, 'FONDS_GARANTIE_AUTO': 100,
            'GENERAL_ASSURANCE_SINISTRE': 100, 'SANLAM_SINISTRE': 100,
            'SUNU_SINISTRE_AUTO': 100, 'COTON_SPORT': 100,
            'ENERGIE_BASKET_BALL': 100, 'LOTTO_FOOTBALL_CLUB': 100,
            'PORT_AUTONOME_COTONOU': 100, 'SOBEMAP': 100,
        }
        sheet_data['coverage'] = default_rates.get(sheet_data['insurer'], 80)

    # Parse des items
    if (sheet_data['has_split'] or sheet_data['bill_type'] in ['DETAIL_ASSUR', 'DEFINITIF']) and item_start_row:
        header_row_idx = item_start_row - 2
        header_row_vals = all_rows[header_row_idx] if 0 <= header_row_idx < len(all_rows) else []
        col_map = map_header_cols(header_row_vals)
        
        for r_idx, row in enumerate(all_rows, 1):
            if r_idx < item_start_row: continue
            if not row or row[0] is None: continue
            
            name = safe(row[0])
            if not name or 'total' in name.lower() or 'caisse' in name.lower() or 'reste a pay' in name.lower():
                continue
                
            try:
                qty_raw = row[col_map['qty']] if col_map['qty'] < len(row) else 1
                qty = float(qty_raw) if qty_raw is not None else 1
                
                price_raw = row[col_map['price']] if col_map['price'] < len(row) else 0
                price = float(price_raw) if price_raw is not None else 0
                
                subtot_raw = row[col_map['subtotal']] if col_map['subtotal'] < len(row) else None
                subtotal = float(subtot_raw) if subtot_raw is not None else 0
                
                part_ass = 0
                for idx in col_map['part_assurance']:
                    if idx < len(row) and row[idx] is not None:
                        part_ass += float(row[idx])
                
                part_pat_raw = row[col_map['part_patient']] if col_map['part_patient'] < len(row) else None
                part_pat = float(part_pat_raw) if part_pat_raw is not None else 0
                
                # Robust computation of missing subtotal (like for BILAN item)
                if subtotal == 0 and part_ass + part_pat > 0:
                    subtotal = part_ass + part_pat
                elif subtotal == 0:
                    subtotal = qty * price
                
                # Auto-split pour patients assurés dans des onglets sans colonnes de split
                if not sheet_data['has_split'] and sheet_data['insurer']:
                    cov = sheet_data['coverage'] or 80
                    part_ass = round(subtotal * (cov / 100))
                    part_pat = subtotal - part_ass
                else:
                    if part_pat == 0 and part_pat_raw is None:
                        part_pat = subtotal - part_ass
                    
                if qty > 0:
                    price = subtotal / qty
                else:
                    price = 0
                
                if subtotal > 0 or price > 0:
                    sheet_data['items'].append({
                        'name': name, 
                        'qty': clean_num(qty), 
                        'price': clean_num(price), 
                        'subtotal': clean_num(subtotal),
                        'partAssurance': clean_num(part_ass), 
                        'partPatient': clean_num(part_pat),
                        'splitLimit': clean_num(subtotal), 
                        'splitRate': clean_num(round(part_ass / subtotal * 100)) if subtotal > 0 else 0
                    })
            except Exception as e:
                pass
                
    # Calculer le total à partir de la somme des items pour garantir la consistance mathématique
    if sheet_data['items']:
        sheet_data['gross_total'] = clean_num(sum(it['subtotal'] for it in sheet_data['items']))
        sheet_data['part_assurance_total'] = clean_num(sum(it['partAssurance'] for it in sheet_data['items']))
        sheet_data['part_patient_total'] = clean_num(sum(it['partPatient'] for it in sheet_data['items']))
    elif sheet_data['insurer'] and sheet_data['gross_total'] > 0:
        cov = sheet_data['coverage'] or 80
        sheet_data['part_assurance_total'] = round(sheet_data['gross_total'] * cov / 100)
        sheet_data['part_patient_total'] = sheet_data['gross_total'] - sheet_data['part_assurance_total']
        
    excel_sheets.append(sheet_data)

# 3. Propagation intelligente de l'assureur
patient_to_insurers = {}
for sheet in excel_sheets:
    pat = sheet['patient_norm']
    ins = sheet['insurer']
    if pat and ins:
        patient_to_insurers.setdefault(pat, set()).add(ins)
for sheet in excel_sheets:
    if not sheet['insurer']:
        pat = sheet['patient_norm']
        if pat in patient_to_insurers and len(patient_to_insurers[pat]) == 1:
            sheet['insurer'] = list(patient_to_insurers[pat])[0]

# 4. Regroupement des factures en base et matching bipartite glouton
db_patients = {}
for idx, b in enumerate(bills):
    fullname = f"{b.get('patientNom','') or ''} {b.get('patientPrenom','') or ''}".upper().strip()
    fullname = re.sub(r'\s+', ' ', fullname)
    if not fullname: continue
    cleaned = clean_patient_name(fullname)
    if cleaned:
        db_patients.setdefault(cleaned, []).append((idx, b))

db_patient_keys = list(db_patients.keys())

# Grouper les onglets par clé de patient en base de données
patient_sheets = {}
unmatched_sheets = []

for sheet in excel_sheets:
    # Traiter uniquement les onglets utiles
    if not sheet['items'] and not sheet['insurer']:
        continue
        
    best_pat_key = None
    best_score = 0
    for db_key in db_patient_keys:
        score = name_score(sheet['patient_norm'], db_key)
        if score > best_score:
            best_score = score
            best_pat_key = db_key
            
    if best_score >= 75:
        patient_sheets.setdefault(best_pat_key, []).append(sheet)
    else:
        unmatched_sheets.append(sheet)

# Effectuer l'appariement optimal glouton unique pour chaque patient
unique_updates = {}
matched_count = 0
skipped_count = 0

for pat_key, group_sheets in patient_sheets.items():
    group_bills = db_patients[pat_key]
    
    # Calculer tous les scores de paires possibles dans ce groupe
    pairs = []
    for sheet in group_sheets:
        for idx_db, b in group_bills:
            b_name = f"{b.get('patientNom','') or ''} {b.get('patientPrenom','') or ''}".upper().strip()
            b_clean = clean_patient_name(b_name)
            n_score = name_score(sheet['patient_norm'], b_clean)
            
            # Score de type
            t_score = 200 if b.get('type') == sheet['bill_type'] else -200
            
            # Score de montant
            diff_total = abs(sheet['gross_total'] - b.get('grossTotal', 0))
            if diff_total <= 100:
                tot_score = 150
            elif diff_total <= 10000:
                tot_score = 100
            elif diff_total <= 50000:
                tot_score = 50
            else:
                tot_score = -50
                
            # Score de date (pas de pénalité sur date placeholder 2026-06-01)
            b_date = b.get('date')
            s_date = sheet['date']
            if b_date == '2026-06-01' or not s_date or not b_date:
                date_score = 0
            else:
                diff_days = date_diff_days(b_date, s_date)
                if diff_days <= 7:
                    date_score = 100
                elif diff_days <= 30:
                    date_score = 75
                elif diff_days <= 60:
                    date_score = 50
                elif diff_days > 180:
                    date_score = -150
                else:
                    date_score = 0
                    
            total_score = n_score + t_score + tot_score + date_score
            pairs.append({
                'sheet': sheet,
                'bill_idx': idx_db,
                'bill_id': b['id'],
                'bill': b,
                'score': total_score,
                'details': f"name={n_score}, type={t_score}, total={tot_score}, date={date_score}"
            })
            
    # Assignation 1-to-1 gloutonne
    pairs.sort(key=lambda x: -x['score'])
    assigned_sheets = set()
    assigned_bills = set()
    
    for p in pairs:
        s_name = p['sheet']['sheet_name']
        b_idx = p['bill_idx']
        if s_name not in assigned_sheets and b_idx not in assigned_bills:
            if p['score'] >= 50:
                assigned_sheets.add(s_name)
                assigned_bills.add(b_idx)
                
                sheet = p['sheet']
                bill = p['bill']
                
                changes = {}
                
                # Assureur & Type
                if sheet['insurer']:
                    changes['insurance'] = sheet['insurer']
                    changes['patientType'] = 'SINISTRE' if ('SINISTRE' in sheet['insurer'] or 'GARANTIE' in sheet['insurer']) else 'MALADIE'
                
                # Couverture
                if sheet['coverage'] > 0:
                    changes['coverage'] = sheet['coverage']
                elif (not sheet['coverage'] or sheet['coverage'] == 0) and bill.get('coverage', 0) == 0:
                    if sheet['insurer']:
                        default_rates = {
                            'AFRICAINE_SINISTRE': 100, 'FONDS_GARANTIE_AUTO': 100,
                            'GENERAL_ASSURANCE_SINISTRE': 100, 'SANLAM_SINISTRE': 100,
                            'SUNU_SINISTRE_AUTO': 100, 'COTON_SPORT': 100,
                            'ENERGIE_BASKET_BALL': 100, 'LOTTO_FOOTBALL_CLUB': 100,
                            'PORT_AUTONOME_COTONOU': 100, 'SOBEMAP': 100,
                        }
                        changes['coverage'] = default_rates.get(sheet['insurer'], 80)
                        
                # Split et totaux d'items réels
                if sheet['items']:
                    changes['useSplit'] = True
                    changes['partAssurance'] = sheet['part_assurance_total']
                    changes['partPatient'] = sheet['part_patient_total']
                    changes['grossTotal'] = sheet['gross_total']
                    changes['amount'] = sheet['gross_total']
                    changes['items'] = sheet['items']
                    
                    # Solde / Règlement
                    part_pat = sheet['part_patient_total']
                    if bill.get('status') == 'RÉGLÉ':
                        changes['amountPaidPatient'] = part_pat
                        changes['balancePatient'] = 0
                    else:
                        paid = bill.get('amountPaidPatient', 0)
                        changes['balancePatient'] = max(0, part_pat - paid)
                        
                # Mettre à jour la date si c'est un placeholder
                if bill.get('date') == '2026-06-01' and sheet['date']:
                    changes['date'] = sheet['date']
                    
                if changes:
                    unique_updates[bill['id']] = {
                        'bill_idx': b_idx,
                        'bill_id': bill['id'],
                        'patient': f"{bill['patientNom']} {bill['patientPrenom']}",
                        'sheet': sheet['sheet_name'],
                        'changes': changes
                    }
                    matched_count += 1
            else:
                skipped_count += 1

print(f"\nFactures uniques à mettre à jour : {len(unique_updates)}")

# 5. Application réelle si pas dry-run
if not DRY_RUN and unique_updates:
    # Création du Backup
    with open(BACKUP_FILE, 'w', encoding='utf-8') as f:
        json.dump(bills, f, indent=2, ensure_ascii=False)
    print(f"Backup créé dans : {BACKUP_FILE.name}")
    
    # Appliquer les modifications
    updated_bills = copy.deepcopy(bills)
    for bid, u in unique_updates.items():
        bill_idx = u['bill_idx']
        for k, v in u['changes'].items():
            updated_bills[bill_idx][k] = v
            
    # Sauvegarder
    with open(BILLS_FILE, 'w', encoding='utf-8') as f:
        json.dump(updated_bills, f, indent=2, ensure_ascii=False)
    print(f"Base de données {BILLS_FILE.name} mise à jour avec succès !")
    
else:
    if DRY_RUN:
        print("\n[DRY-RUN] Mode simulation. Aucune modification de bills_db.json.")
    else:
        print("\nAucune modification à appliquer.")

# 6. Écriture du rapport
with open(REPORT_FILE, 'w', encoding='utf-8') as f:
    f.write(f"Rapport de re-import enrichi - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write(f"{'='*60}\n")
    f.write(f"Mode de lancement      : {'SIMULATION (DRY-RUN)' if DRY_RUN else 'RÉEL'}\n")
    f.write(f"Total onglets analysés : {len(excel_sheets)}\n")
    f.write(f"Factures mises à jour  : {len(unique_updates)}\n")
    
    f.write("\nDÉTAIL DES FACTURES ENRICHIES:\n")
    for bid, u in sorted(unique_updates.items()):
        f.write(f"\n  [{bid}] Patient: {u['patient']} (Onglet: {u['sheet']})\n")
        for k, v in u['changes'].items():
            if k == 'items':
                f.write(f"    - items : {len(v)} items avec split d'assurance\n")
                for it in v:
                    f.write(f"      * {it['name'][:30]:30} | PU={str(it['price']):8} | Qty={str(it['qty']):3} | Sub={str(it['subtotal']):8} | Ass={str(it['partAssurance']):8} | Pat={str(it['partPatient']):8}\n")
            else:
                f.write(f"    - {k} : {v}\n")

print(f"\nRapport écrit dans : scratch/reimport_report.txt")
