"""
detect_split_sheets.py
Scanne tous les onglets de EXEMPLAIRE PROFORMA.xlsx pour identifier :
1. Les onglets avec colonnes SPLIT (Part Assurance / Part Patient)
2. Les onglets DEFINITIF (Point d'hospitalisation)
3. Les onglets avec données assurance (Société/Assurance mentionnée)
4. La structure exacte de chaque type d'onglet
"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from pathlib import Path
import re

EXCEL_FILE = Path(r"C:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx")

def safe(v, maxlen=50):
    if v is None: return ''
    s = str(v).strip().replace('\n', ' ').replace('\r', '')
    s = s.encode('utf-8', errors='replace').decode('utf-8')
    return s[:maxlen] if len(s) > maxlen else s

print("Chargement du fichier Excel...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"Total onglets : {len(wb.sheetnames)}\n")

# Catégories à détecter
split_sheets = []       # Onglets avec colonnes Part Assurance / Part Patient
definitif_sheets = []   # Onglets avec "Point Définitif" ou "DEFINITIF"
detail_assur_sheets = []  # Onglets avec "DETAIL" + "assurance"
insured_sheets = []     # Onglets avec une compagnie d'assurance identifiée

INSURERS = ['sanlam', 'nsia', 'ascoma', 'sunu', 'atlantique', 'africaine', 'afg', 
            'loto', 'lotto', 'coton', 'sobemap', 'port autonome', 'fonds', 'sinistre',
            'assurance', 'mutuelle', 'compagnie', 'société', 'societe']

for i, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    
    sheet_info = {
        'index': i+1,
        'name': sheet_name,
        'patient': '',
        'insurer': '',
        'coverage': 0,
        'bill_type': 'PROFORMA',
        'total': 0,
        'num_cols': ws.max_column,
        'sample_rows': [],
        'has_split_cols': False,
        'has_definitif': False,
        'has_detail_assur': False,
    }
    
    # Lire les 40 premières lignes
    all_rows = list(ws.iter_rows(min_row=1, max_row=min(40, ws.max_row), values_only=True))
    
    for row_idx, row in enumerate(all_rows, 1):
        row_str = ' '.join(safe(v) for v in row if v is not None).lower()
        
        # Détecter le type de document
        if any(k in row_str for k in ['point définitif', 'point definitif', 'définitif', 'definitif']):
            sheet_info['has_definitif'] = True
            sheet_info['bill_type'] = 'DEFINITIF'
        
        if any(k in row_str for k in ['détail assurance', 'detail assurance', 'détail assu', 'detail assu']):
            sheet_info['has_detail_assur'] = True
            sheet_info['bill_type'] = 'DETAIL_ASSUR'
        
        # Détecter les colonnes split (chercher une ligne d'en-tête avec part assur + part patient)
        if any(k in row_str for k in ['part assur', 'part patient', 'tiers payant', 'ticket modér']):
            sheet_info['has_split_cols'] = True
        
        # Détecter l'assureur
        for ins in INSURERS:
            if ins in row_str and not sheet_info['insurer']:
                # Trouver la ligne complète
                for v in row:
                    if v and ins in safe(v).lower():
                        sheet_info['insurer'] = safe(v, 60)
                        break
        
        # Chercher le nom du patient
        if not sheet_info['patient']:
            for v in row:
                if v and isinstance(v, str) and ('patient' in v.lower() or 'patient:' in v.lower()):
                    p = re.sub(r'patient\s*:?\s*', '', v, flags=re.IGNORECASE).strip()
                    if p and len(p) > 2:
                        sheet_info['patient'] = p[:50]
                        break
        
        # Chercher le total
        for c_idx, v in enumerate(row):
            if v and isinstance(v, str) and 'total' in v.lower():
                # Chercher la valeur numérique sur la même ligne
                for nv in row[c_idx:]:
                    if isinstance(nv, (int, float)) and nv > 10000:
                        sheet_info['total'] = int(nv)
                        break
        
        # Capturer le % de couverture
        m = re.search(r'(\d+)\s*%', row_str)
        if m:
            sheet_info['coverage'] = int(m.group(1))
    
    # Si pas de patient dans le contenu, utiliser le nom d'onglet
    if not sheet_info['patient']:
        sheet_info['patient'] = sheet_name
    
    # Classifier
    if sheet_info['has_split_cols']:
        split_sheets.append(sheet_info)
    if sheet_info['has_definitif']:
        definitif_sheets.append(sheet_info)
    if sheet_info['has_detail_assur']:
        detail_assur_sheets.append(sheet_info)
    if sheet_info['insurer']:
        insured_sheets.append(sheet_info)

# Afficher les résultats
print(f"\n{'='*70}")
print(f"RÉSULTATS DE L'ANALYSE")
print(f"{'='*70}")
print(f"Onglets avec colonnes SPLIT (Part Assur / Part Patient) : {len(split_sheets)}")
print(f"Onglets DEFINITIF (Point d'hospitalisation)             : {len(definitif_sheets)}")
print(f"Onglets DETAIL_ASSUR                                    : {len(detail_assur_sheets)}")
print(f"Onglets avec assureur identifié                         : {len(insured_sheets)}")

print(f"\n\n--- EXEMPLES ONGLETS AVEC SPLIT (premiers 5) ---")
for s in split_sheets[:5]:
    print(f"  [{s['index']:3d}] {s['name'][:40]:40s} | Patient: {s['patient'][:30]} | Assureur: {s['insurer'][:25]} | Total: {s['total']:>10,}")

print(f"\n--- EXEMPLES ONGLETS DEFINITIF (premiers 10) ---")
for s in definitif_sheets[:10]:
    print(f"  [{s['index']:3d}] {s['name'][:40]:40s} | Patient: {s['patient'][:30]} | Total: {s['total']:>10,} | Assureur: {s['insurer'][:20]}")

print(f"\n--- EXEMPLES ONGLETS ASSURÉS (premiers 10) ---")
for s in insured_sheets[:10]:
    print(f"  [{s['index']:3d}] {s['name'][:40]:40s} | {s['insurer'][:40]} | Taux: {s['coverage']}%")

# Chercher un exemple avec split pour voir la structure exacte
print(f"\n\n--- STRUCTURE DÉTAILLÉE D'UN ONGLET SPLIT ---")
if split_sheets:
    ex = split_sheets[0]
    ws = wb[ex['name']]
    print(f"Onglet: {ex['name']}")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), 1):
        if any(v is not None for v in row):
            vals = [safe(v, 25) for v in row]
            print(f"  L{row_idx:2d} [{len([v for v in row if v is not None])} vals]: {' | '.join(v for v in vals if v)}")

print("\nAnalyse terminée!")
