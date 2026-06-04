import openpyxl
import os
import re

workspace_dir = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_main = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")

wb = openpyxl.load_workbook(excel_main, data_only=True)
feuil_sheets = [s for s in wb.sheetnames if re.match(r'^feuil\d+(\s*\(.*\))?$', s.lower().strip())]

target_sheets = [
    # Those that failed to find names
    "Feuil6", "Feuil4", "Feuil7", "Feuil3",
    # Some of those that returned CLINIQUE MERCY FIAT
    "Feuil2 (17)", "Feuil2 (20)", "Feuil2 (28)", "Feuil2 (54)", "Feuil2 (55)", "Feuil2 (56)", "Feuil2 (60)"
]

# If they are not exactly in the workbook, find case-insensitive/approximate matches
matching_sheets = []
for target in target_sheets:
    for s in wb.sheetnames:
        if s.lower().strip() == target.lower().strip():
            matching_sheets.append(s)
            break

print(f"Inspecting {len(matching_sheets)} sheets...\n")

for sheetname in matching_sheets:
    print("=" * 80)
    print(f"SHEET: {sheetname}")
    print("=" * 80)
    ws = wb[sheetname]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=20), 1):
        # Format the row to show column index and value for non-empty cells
        row_cells = []
        for c_idx, val in enumerate(row, 1):
            if val is not None:
                row_cells.append(f"Col {c_idx} ({openpyxl.utils.get_column_letter(c_idx)}): {repr(val)}")
        if row_cells:
            print(f"Row {r_idx:02d}: " + " | ".join(row_cells))
    print()
