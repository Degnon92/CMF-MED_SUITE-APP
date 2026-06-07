import openpyxl, re, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

EXCEL_FILE = r"C:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx"
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=True))
    
    # Check if there is an acts/designation header row
    header_row = None
    for r_idx, row in enumerate(all_rows, 1):
        row_str = ' '.join(str(v) for v in row if v is not None).lower()
        if any(k in row_str for k in ['actes/designations', 'actes/désignations', 'désignations', 'designation', 'actes', 'prestations']):
            header_row = (r_idx, [v for v in row if v is not None])
            break
            
    if header_row:
        r_idx, cols = header_row
        # Check if it has split (contains 'assur', 'patient', 'part', 'tiers')
        cols_str = ' | '.join(str(v) for v in cols).lower()
        if any(k in cols_str for k in ['assur', 'patient', 'part', 'tiers']):
            print(f"Sheet: '{sheet_name}' (type: {len(cols)} cols)")
            print(f"  Header L{r_idx}: {cols}")
            # print first data row
            first_data = None
            for row in all_rows[r_idx:]:
                if any(row):
                    first_data = [v for v in row if v is not None]
                    break
            if first_data:
                print(f"  First data row: {first_data}")
