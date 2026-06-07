"""
CORRECTION COMPLÈTE DE LA BASE DE DONNÉES
1. Corrige les 35 montants erronés
2. Importe les 393 onglets manquants depuis EXEMPLAIRE PROFORMA.xlsx
"""
import json
import sys
import os
import re
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_main = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")
bills_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "bills_db.json")

print("="*80)
print("CORRECTION COMPLÈTE — MONTANTS + FACTURES MANQUANTES")
print("="*80)

with open(bills_db_path, encoding='utf-8') as f:
    bills_db = json.load(f)

print(f"Factures actuelles en DB: {len(bills_db)}")

# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────────────────────────────────────
def normalize(s):
    if not s: return ""
    return re.sub(r'\s+', ' ', str(s).strip().upper())

def cell_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return int(val)
    try:
        return int(float(str(val).replace(' ', '').replace('\xa0','').replace(',','.')))
    except: return 0

def parse_date_val(val):
    if not val: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2: y = '20' + y
        try: return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except: return None
    return None

def clean_term(s):
    if not s or not isinstance(s, str): return ""
    s = s.strip().strip('-').strip('.').strip()
    if len(s) < 2: return ""
    return s

def map_insurance(text):
    if not text: return "PRIVE"
    t = text.lower().strip()
    if "sanlam" in t: return "SANLAM"
    if "ascoma" in t: return "ASCOMA"
    if "sunu" in t: return "SUNU"
    if "nsia" in t: return "NSIA"
    if "atlantique" in t: return "ATLANTIQUE"
    if "africaine" in t or "afg" in t or " aa " in t or t.endswith(" aa"): return "AFG"
    if "loto" in t or "lotto" in t: return "LOTTO_FOOTBALL_CLUB"
    if "coton" in t: return "COTON_SPORT"
    if "olea" in t: return "OLEA"
    if "gras" in t or "savoye" in t: return "GRAS SAVOYE"
    if "sobemap" in t or "sobremap" in t: return "SOBEMAP"
    if "fga" in t or "garantie" in t or "fonds" in t: return "FONDS_GARANTIE_AUTO"
    if "transvie" in t: return "TRANSVIE"
    if "nobila" in t: return "NOBILA"
    return "PRIVE"

def split_nom_prenom(full):
    full = re.sub(r'\s+', ' ', full.strip())
    parts = full.split(' ')
    if len(parts) == 1:
        return full.upper(), ""
    nom_parts, prenom_parts = [], []
    for p in parts:
        if p.isupper() and len(p) > 1:
            nom_parts.append(p)
        else:
            prenom_parts.append(p)
    if not nom_parts:
        return parts[0].upper(), ' '.join(parts[1:])
    if not prenom_parts and len(parts) > 1:
        return parts[0].upper(), ' '.join(parts[1:])
    return ' '.join(nom_parts).upper(), ' '.join(prenom_parts)

IGNORE_SHEETS = {'feuil1','feuil2','feuil3','feuil4','feuil5','feuil6','feuil7',
                 'winsu sports','sheet1','config','param','aide','tarifs',
                 'forfaits','liste','nomenclature',''}

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION COMPLÈTE D'UN ONGLET
# ─────────────────────────────────────────────────────────────────────────────
def extract_sheet(ws, sheetname):
    """Extrait toutes les données d'un onglet. Retourne None si onglet vide/invalide."""
    all_rows = list(ws.iter_rows(values_only=True, max_row=120))
    if not all_rows:
        return None

    patient_name = ""
    date_source = None
    items = []
    total_source = 0
    intervention = ""
    diagnostic = ""
    insurance_txt = ""
    k_code = ""
    bill_type = "PROFORMA"
    header_row = -1

    for ri, row in enumerate(all_rows[:50]):
        for ci, val in enumerate(row):
            if val is None: continue
            if isinstance(val, datetime):
                if not date_source:
                    date_source = val.strftime('%Y-%m-%d')
                continue
            if not isinstance(val, str): continue
            s = val.strip()
            sl = s.lower()

            # Patient
            if re.match(r'patient\s*:', sl, re.I):
                v = s.split(':', 1)[1].strip()
                if v and len(v) > 2: patient_name = v
            # Assurance / Client
            elif re.match(r'(client|assurance|assureur)\s*:', sl, re.I):
                insurance_txt = s.split(':', 1)[1].strip()
            # Intervention
            elif re.match(r'intervention\s*:', sl, re.I):
                v = s.split(':', 1)[1].strip()
                if v: intervention = v
            # Diagnostic
            elif re.match(r'diagnostic\s*:', sl, re.I):
                v = s.split(':', 1)[1].strip()
                if v: diagnostic = v
            # Date
            if ('cotonou' in sl or 'le ' in sl) and not date_source:
                d = parse_date_val(s)
                if d: date_source = d
            # Type de facture
            if any(k in sl for k in ['définitif', 'definitif', 'point d\'hospit', 'point hospit']):
                bill_type = "DEFINITIF"
            elif any(k in sl for k in ['détail assur', 'detail assur', 'détail facture assur']):
                bill_type = "DETAIL_ASSUR"
            # Code K
            km = re.search(r'\b(KC?\s*\d+)\b', s)
            if km: k_code = km.group(1)
            # En-tête tableau
            if any(k in sl for k in ['désignation','designation','actes','libellé','libelle','prestation']):
                if header_row == -1:
                    header_row = ri

    # Chercher aussi la date dans les cellules datetime de tout l'onglet
    if not date_source:
        for row in all_rows[:60]:
            for val in row:
                if isinstance(val, datetime):
                    date_source = val.strftime('%Y-%m-%d')
                    break
            if date_source: break

    # Si toujours pas de patient → utiliser le nom de l'onglet
    if not patient_name:
        if re.search(r'[A-Za-z]{3,}', sheetname) and sheetname.lower() not in IGNORE_SHEETS:
            patient_name = sheetname
        else:
            return None

    # ── Items du tableau ──
    if header_row >= 0:
        for row in all_rows[header_row+1: header_row+50]:
            if not row or row[0] is None: continue
            d0 = str(row[0]).strip() if row[0] else ''
            if not d0: continue
            d0_low = d0.lower()
            # Ligne total → stop
            if any(k in d0_low for k in ['total','caisse','à payer','a payer','montant total','net à','reste','solde','arrêté']):
                # Capturer le total de cette ligne
                for v in row:
                    n = cell_num(v)
                    if n > 10000:
                        total_source = max(total_source, n)
                break
            price = cell_num(row[2] if len(row) > 2 else None)
            qty = cell_num(row[1] if len(row) > 1 else None) or 1
            subtot = cell_num(row[3] if len(row) > 3 else None) or (price * qty)
            if price > 0 and len(d0) > 2 and d0_low not in ['quantité','qté','prix','désignation','designation','montant','total']:
                items.append({'name': clean_term(d0), 'price': price, 'qty': qty, 'subtotal': subtot})
                if subtot: total_source += subtot

    # Chercher le total global dans toute la feuille (ligne TOTAL)
    for row in all_rows:
        for ci, val in enumerate(row):
            if not val or not isinstance(val, str): continue
            vl = val.lower().strip()
            if any(k in vl for k in ['total général','total general','total g','à payer','a payer','net à payer']):
                for v in row[ci:]:
                    n = cell_num(v)
                    if n > 10000:
                        total_source = n  # Prendre le total officiel
                        break

    if not items and total_source == 0:
        return None  # Onglet vraiment vide

    insurance = map_insurance(insurance_txt)
    coverage = 80 if insurance != "PRIVE" else 0
    part_assurance = int(total_source * coverage / 100) if insurance != "PRIVE" else 0
    part_patient = total_source - part_assurance

    nom, prenom = split_nom_prenom(patient_name)
    if not nom:
        return None

    return {
        'patientNom': nom,
        'patientPrenom': prenom,
        'patient_full': f"{nom} {prenom}".strip(),
        'date': date_source or '2026-06-01',
        'date_invented': not bool(date_source),
        'items': items,
        'grossTotal': total_source,
        'insurance': insurance,
        'coverage': coverage,
        'partAssurance': part_assurance,
        'partPatient': part_patient,
        'intervention': intervention,
        'diagnostic': diagnostic,
        'kCode': k_code,
        'billType': bill_type,
        'sheetname': sheetname
    }

# ─────────────────────────────────────────────────────────────────────────────
# INDEX EXISTANT EN DB
# ─────────────────────────────────────────────────────────────────────────────
# Clé = (patientNom_upper, date) → list de bills
db_index_by_nom_date = {}
for b in bills_db:
    nom = normalize(b.get('patientNom', ''))
    prenom = normalize(b.get('patientPrenom', ''))
    date = b.get('date', '')
    key = (nom, date)
    db_index_by_nom_date.setdefault(key, []).append(b)

db_index_by_nom = {}
for b in bills_db:
    nom = normalize(b.get('patientNom', ''))
    prenom = normalize(b.get('patientPrenom', ''))
    full = f"{nom} {prenom}".strip()
    db_index_by_nom.setdefault(full, []).append(b)

def find_existing(nom, prenom, date, sheetname):
    """Cherche si cette facture existe déjà en DB."""
    n = normalize(nom)
    p = normalize(prenom)
    full = f"{n} {p}".strip()
    # Clé exacte nom+date
    if (n, date) in db_index_by_nom_date:
        return db_index_by_nom_date[(n, date)]
    # Clé nom seul
    if full in db_index_by_nom:
        return db_index_by_nom[full]
    # Partiel
    n_parts = [w for w in n.split() if len(w) > 2]
    for key, blist in db_index_by_nom.items():
        if len(n_parts) >= 1 and all(w in key for w in n_parts[:2]):
            return blist
    return None

# ─────────────────────────────────────────────────────────────────────────────
# CHARGER L'EXCEL ET TRAITER CHAQUE ONGLET
# ─────────────────────────────────────────────────────────────────────────────
print(f"\nOuverture de: {excel_main}")
wb = openpyxl.load_workbook(excel_main, data_only=True)
print(f"603 onglets à traiter...\n")

corrected = 0
added = 0
skipped_empty = 0
skipped_dupe = 0
errors = []

# Générer un ID unique pour les nouvelles factures
max_id = 0
for b in bills_db:
    bid = b.get('id', '')
    m = re.search(r'(\d+)$', bid)
    if m:
        max_id = max(max_id, int(m.group(1)))

next_id = max_id + 1

new_bills = []

for sheetname in wb.sheetnames:
    sn_low = sheetname.lower().strip()
    if sn_low in IGNORE_SHEETS:
        continue

    try:
        ws = wb[sheetname]
        data = extract_sheet(ws, sheetname)
    except Exception as e:
        errors.append(f"[ERREUR] {sheetname}: {e}")
        continue

    if data is None:
        skipped_empty += 1
        continue

    nom = data['patientNom']
    prenom = data['patientPrenom']
    date = data['date']
    total = data['grossTotal']

    # Chercher si déjà en DB
    existing = find_existing(nom, prenom, date, sheetname)

    if existing:
        # Vérifier si le montant est différent
        bill = existing[0]
        # Choisir la facture la plus proche par date
        for b in existing:
            if b.get('date') == date:
                bill = b
                break

        db_total = bill.get('grossTotal', 0)

        if total > 0 and abs(db_total - total) > 5000:
            # CORRIGER LE MONTANT
            old_total = db_total
            bill['grossTotal'] = total
            bill['discountedTotal'] = total
            if bill.get('insurance', 'PRIVE') != 'PRIVE':
                cov = bill.get('coverage', 80)
                bill['partAssurance'] = int(total * cov / 100)
                bill['partPatient'] = total - bill['partAssurance']
                bill['amountPaidPatient'] = bill['partPatient']
            else:
                bill['partPatient'] = total
                bill['partAssurance'] = 0
                bill['amountPaidPatient'] = total
            bill['balancePatient'] = 0
            # Mettre à jour les items si on en a
            if data['items']:
                bill['items'] = data['items']
            # Mettre à jour date si elle était inventée
            if bill.get('date') in ('2026-06-01', '') and not data['date_invented']:
                bill['date'] = date
            corrected += 1
            print(f"✅ CORRIGÉ: {nom} {prenom} | {old_total:,} → {total:,} FCFA | Date: {date} | [{sheetname}]")
        else:
            skipped_dupe += 1
    else:
        # NOUVELLE FACTURE À AJOUTER
        if total == 0 and not data['items']:
            skipped_empty += 1
            continue

        # Créer la facture
        items = data['items'] if data['items'] else [{'name': 'Frais de prise en charge', 'price': total, 'qty': 1, 'subtotal': total}]
        insurance = data['insurance']
        coverage = data['coverage']
        part_assurance = data['partAssurance']
        part_patient = data['partPatient']
        bill_type = data['billType']

        new_bill = {
            "id": f"BILL-CMF-IMP-{next_id:04d}",
            "reference": f"MF-{bill_type[:3]}-IMP-{next_id:04d}",
            "patientNom": nom,
            "patientPrenom": prenom,
            "type": bill_type,
            "customTitle": "Point Définitif d'Hospitalisation" if bill_type == "DEFINITIF" else "",
            "insurance": insurance,
            "coverage": coverage,
            "matricule": "",
            "diagnostic": data['diagnostic'],
            "intervention": data['intervention'],
            "kCode": data['kCode'],
            "showDiag": bool(data['diagnostic']),
            "showInterv": bool(data['intervention']),
            "showSig": True,
            "showCachet": True,
            "useSplit": insurance != "PRIVE",
            "items": items,
            "grossTotal": total,
            "discountPct": 0,
            "reductionAmount": 0,
            "discountedTotal": total,
            "discountType": "PERCENT",
            "discountValue": 0,
            "partAssurance": part_assurance,
            "partPatient": part_patient,
            "paymentMethod": "TIERS_PAYANT" if insurance != "PRIVE" else "CASH",
            "amountPaidPatient": part_patient,
            "balancePatient": 0,
            "status": "RÉGLÉ",
            "date": date,
            "patientAge": "",
            "diagnosis": data['diagnostic']
        }
        new_bills.append(new_bill)
        next_id += 1
        added += 1
        print(f"➕ AJOUTÉ: {nom} {prenom} | {total:,} FCFA | Date: {date} | [{sheetname}]")

# ─────────────────────────────────────────────────────────────────────────────
# SAUVEGARDER
# ─────────────────────────────────────────────────────────────────────────────
all_bills = bills_db + new_bills

with open(bills_db_path, 'w', encoding='utf-8') as f:
    json.dump(all_bills, f, ensure_ascii=False, indent=2)

print()
print("="*80)
print("RÉSULTAT FINAL")
print("="*80)
print(f"✅ Montants corrigés     : {corrected}")
print(f"➕ Factures ajoutées     : {added}")
print(f"⏭️  Doublons ignorés      : {skipped_dupe}")
print(f"📭 Onglets vides ignorés : {skipped_empty}")
print(f"❌ Erreurs               : {len(errors)}")
print(f"\nDB avant: {len(bills_db)} → DB après: {len(all_bills)} factures")
print(f"\nBase de données sauvegardée ✓")

if errors:
    print("\nErreurs rencontrées:")
    for e in errors[:20]:
        print(f"  {e}")
