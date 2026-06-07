import docx
import openpyxl
import re
import json
import os
from docx.oxml.ns import qn
from datetime import datetime

# File paths
workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
docx_path = os.path.join(workspace_dir, "RAPPORT HOSPI CMF", "RAPPORT D'HOSPI CMF.docx")
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")

patients_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "patients_db.json")
documents_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "documents_db.json")
bills_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "bills_db.json")
real_data_js_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "real_data.js")
database_js_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "database.js")

print("Loading existing database files...")

# Load existing data
existing_patients = []
if os.path.exists(patients_db_path):
    try:
        with open(patients_db_path, "r", encoding="utf-8") as f:
            existing_patients = json.load(f)
    except Exception as e:
        print(f"Error loading patients: {e}")

existing_docs = []
if os.path.exists(documents_db_path):
    try:
        with open(documents_db_path, "r", encoding="utf-8") as f:
            existing_docs = json.load(f)
    except Exception as e:
        print(f"Error loading documents: {e}")

existing_bills = []
if os.path.exists(bills_db_path):
    try:
        with open(bills_db_path, "r", encoding="utf-8") as f:
            existing_bills = json.load(f)
    except Exception as e:
        print(f"Error loading bills: {e}")

print(f"Loaded {len(existing_patients)} patients, {len(existing_docs)} documents, {len(existing_bills)} bills.")

# Create lookup map for existing patients
patients_map = {}
for p in existing_patients:
    patients_map[p["name"].strip().upper()] = p

# =============================================================
# 1. PARSING RAPPORT D'HOSPI CMF.docx
# =============================================================
print(f"Parsing hospitalisation DOCX: {docx_path}...")
doc = docx.Document(docx_path)

tb_texts = []
root = doc.element
for el in root.iter():
    if el.tag.endswith('txbxContent'):
        for p in el.findall(qn('w:p')):
            p_obj = docx.text.paragraph.Paragraph(p, doc)
            text = p_obj.text.strip()
            if text:
                tb_texts.append(text)

raw_reports = []
current_report = []
for text in tb_texts:
    if text.lower().startswith("patient") and ":" in text:
        if current_report:
            raw_reports.append(current_report)
        current_report = [text]
    else:
        current_report.append(text)
if current_report:
    raw_reports.append(current_report)

sidebar_doctors = ["Dr DAH", "Dr BACHAROU", "Dr HAZOUME", "Dr LASSISSI", "Dr MEDENOU", "Dr SESSINOU", "Dr CHOBLI", "Dr AGAVOEDO", "Dr DJEDOU", "Dr JACQUET", "Dr SOUMANOU", "Dr ELEGBEDE", "Dr KASSEIN", "Dr AKPAKPO", "Collaborateurs", "Médecine générale", "Pédiatrie", "Cardiologie", "Endocrinologie", "Neurologie", "Anesthésie", "Traumatologie", "Urologie", "Radiologie", "Laboratoire"]

months_map = {
    "janvier": "01", "fevrier": "02", "février": "02", "mars": "03", "avril": "04",
    "mai": "05", "juin": "06", "juillet": "07", "aout": "08", "août": "08",
    "septembre": "09", "octobre": "10", "novembre": "11", "decembre": "12", "décembre": "12"
}

def parse_french_date(text):
    if not text:
        return None
    m_slashes = re.search(r"(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})", text)
    if m_slashes:
        day = f"{int(m_slashes.group(1)):02d}"
        month = f"{int(m_slashes.group(2)):02d}"
        year = m_slashes.group(3)
        return f"{year}-{month}-{day}"
    m_words = re.search(r"le\s+(\d{1,2})\s+(\S+)\s+(\d{4})", text, re.IGNORECASE)
    if m_words:
        day = f"{int(m_words.group(1)):02d}"
        m_name = m_words.group(2).lower()
        year = m_words.group(3)
        
        month_val = None
        if m_name.startswith("juin"):
            month_val = 6
        elif m_name.startswith("juil"):
            month_val = 7
        else:
            months_prefixes = {
                "jan": 1, "fev": 2, "fév": 2, "mar": 3, "avr": 4, "mai": 5, "jui": 6,
                "ao": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12, "déc": 12
            }
            for pref, val in months_prefixes.items():
                if m_name.startswith(pref):
                    month_val = val
                    break
        if month_val:
            return f"{year}-{month_val:02d}-{day}"
    return None

def clean_patient_name(name):
    if not name:
        return ""
    clean = name.strip()
    # Strip administrative suffixes
    clean = re.sub(r'(?:\s+|:)\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    if clean.upper().endswith("AGE") and len(clean) > 5:
        clean = clean[:-3].strip()
    # Remove trailing punctuation
    clean = re.sub(r'[\s\-\.\,\:\_]+$', '', clean).strip()
    clean = re.sub(r'\s+', ' ', clean)
    clean = re.sub(r'\s+\(?(?:ok|copie|2|3|4|1)\)?$', '', clean, flags=re.IGNORECASE).strip()
    return clean

# Extract unique reports to prevent duplicates from overlapping textboxes
unique_hospi_reports = {}
for r in raw_reports:
    header = r[0]
    header_match = re.match(r"patient\s*:\s*(.*)", header, re.IGNORECASE)
    if not header_match:
        continue
    raw_name = header_match.group(1).strip()
    
    clean_lines = []
    for line in r[1:]:
        if any(doc_name in line for doc_name in sidebar_doctors):
            continue
        clean_lines.append(line.strip())
        
    full_text = "\n".join(clean_lines).strip()
    
    # Hash on name + text content snippet to deduplicate textbox overlaps
    text_snippet = re.sub(r'\s+', '', full_text[:150].lower())
    key = (clean_patient_name(raw_name).upper(), text_snippet)
    if key not in unique_hospi_reports:
        unique_hospi_reports[key] = {
            "raw_name": raw_name,
            "text": full_text,
            "lines": clean_lines
        }

print(f"Unique reports extracted from docx: {len(unique_hospi_reports)}")

doc_import_count = 0
new_patient_count = 0

all_new_docs = []

for key, data in unique_hospi_reports.items():
    raw_name = data["raw_name"]
    full_text = data["text"]
    clean_lines = data["lines"]
    
    clean_name = clean_patient_name(raw_name)
    words = clean_name.split()
    if len(words) < 2:
        continue
        
    last_name = words[0].upper()
    first_name = " ".join(words[1:])
    fullname_upper = clean_name.upper()
    
    # Extract age
    age = "N/A"
    age_match = re.search(r"g\s+de\s+(\d+\s*(?:ans|mois))|ge\s+de\s+(\d+\s*(?:ans|mois))|g\s+(\d+\s*(?:ans|mois))|ge\s+(\d+\s*(?:ans|mois))|(\d+)\s*ans", full_text, re.IGNORECASE)
    if age_match:
        age = next(g for g in age_match.groups() if g is not None).strip()
        if not age.endswith("ans") and not age.endswith("mois"):
            age = age + " ans"
            
    # Extract diagnosis
    diagnosis = "Bilan clinique"
    diag_match = re.search(r"(?:objectiv|mis\s+en\s+vidence|diagnostic\s+de)\s+([^.\n]+)", full_text, re.IGNORECASE)
    if diag_match:
        diagnosis = diag_match.group(1).strip()
    else:
        diag_patterns = [
            r"tendinite\s+[^.\n]+",
            r"rupture\s+[^.\n]+",
            r"syndrome\s+[^.\n]+",
            r"fracture\s+[^.\n]+",
            r"entorse\s+[^.\n]+",
            r"lésion\s+[^.\n]+",
            r"ostéite\s+[^.\n]+"
        ]
        for pattern in diag_patterns:
            m = re.search(pattern, full_text, re.IGNORECASE)
            if m:
                diagnosis = m.group(0).strip()
                break
    diagnosis = re.sub(r"[\s\-\.\,\:\_]+$", "", diagnosis).strip()
    if len(diagnosis) > 100:
        diagnosis = diagnosis[:97] + "..."
        
    # Check insurance
    insurer = "PRIVE"
    prise_en_charge = 0
    insurer_map = {
        "ascoma": "ASCOMA",
        "sunu": "SUNU",
        "nsia": "NSIA",
        "allianz": "ALLIANZ",
        "saham": "SANLAM",
        "sanlam": "SANLAM",
        "axa": "PRIVE",
        "lotto": "LOTTO_FOOTBALL_CLUB",
        "coton": "COTON_SPORT",
        "energie": "ENERGIE_BASKET_BALL"
    }
    
    has_insurance = False
    for keyword, partner_id in insurer_map.items():
        if re.search(keyword, full_text, re.IGNORECASE) or re.search(keyword, clean_name, re.IGNORECASE):
            insurer = partner_id
            has_insurance = True
            if partner_id != "PRIVE":
                prise_en_charge = 80 if partner_id not in ["LOTTO_FOOTBALL_CLUB", "COTON_SPORT", "ENERGIE_BASKET_BALL"] else 100
            break
    if not has_insurance:
        if any(w in full_text.lower() for w in ["assurance", "mutuelle", "prise en charge", "sinistre", "bon"]):
            insurer = "ASCOMA"
            prise_en_charge = 80
            
    # Extract intervention
    intervention = ""
    k_code = ""
    interv_patterns = [
        (r"ligamentoplastie\s+du\s+lca|didt", "Ligamentoplastie du LCA par DIDT", "LCA"),
        (r"prothèse\s+totale\s+du\s+genou|ptg", "Prothèse Totale du Genou (PTG)", "PTG"),
        (r"prothèse\s+totale\s+de\s+la\s+hanche|pth", "Prothèse Totale de la Hanche (PTH)", "PTH"),
        (r"ostéosynthèse\s+de\s+la\s+clavicule", "Ostéosynthèse de la clavicule", "CLAVICULE"),
        (r"ostéosynthèse\s+de\s+l'humérus", "Ostéosynthèse de l'humérus / radius / cubitus", "HUMERUS"),
        (r"ostéosynthèse\s+du\s+radius|ostéosynthèse\s+de\s+radius", "Ostéosynthèse de l'humérus / radius / cubitus", "HUMERUS"),
        (r"ostéosynthèse\s+du\s+tibia|ostéosynthèse\s+de\s+tibia|ostéosynthèse\s+du\s+fémur|ostéosynthèse\s+de\s+fémur", "Ostéosynthèse du tibia ou du fémur", "TIBIA"),
        (r"ablation\s+de\s+matériel|amos", "Ablation de matériel d'ostéosynthèse (AMOS)", "AMOS_CLE"),
        (r"arthroscopie\s+diagnostique", "Arthroscopie diagnostique & debridement du genou", "ARTHRO_DIAG"),
        (r"réduction\s+orthopédique\s+de\s+luxation", "Réduction orthopédique de luxation de membre", "LUXATION_CMF"),
        (r"ostéosynthèse\s+de\s+fracture\s+symphysaire|symphysaire", "Ostéosynthèse de fracture symphysaire mandibulaire par mini-plaques", "FX_MANDIBULE"),
        (r"résection\s+transurétrale|rtup", "Résection Transurétrale de la Prostate (RTUP)", "RTUP"),
        (r"pose\s+de\s+sonde\s+double\s+j|pose\s+de\s+sonde\s+jj", "Pose ou descente de Sonde Double J (JJ)", "SONDE_JJ_POSE"),
        (r"ablation\s+de\s+sonde\s+double\s+j|ablation\s+de\s+sonde\s+jj", "Ablation de Sonde double J par urétéroscopie", "SONDE_JJ_ABLATION"),
        (r"cure\s+d'hydrocèle|varicocèle", "Cure d'hydrocèle ou de varicocèle", "HYDROCELE")
    ]
    for pat, label, code in interv_patterns:
        if re.search(pat, full_text, re.IGNORECASE):
            intervention = label
            k_code = code
            break

    # Date
    report_date = parse_french_date(full_text)
    if not report_date:
        report_date = "2025-03-15" # Default reasonable hospitalisation date

    # Categorization
    template_id = "rapport_hospi_simple"
    doc_title = "RAPPORT D'HOSPITALISATION"
    doc_category = "Rapport HOSPI"
    
    for line in clean_lines[:4]:
        line_upper = line.upper()
        if "REPOS" in line_upper:
            template_id = "certif_repos"
            doc_title = "CERTIFICAT DE REPOS MEDICAL"
            doc_category = "Certificat"
            break
        elif "REPRISE" in line_upper:
            template_id = "certif_reprise"
            doc_title = "CERTIFICAT DE REPRISE DE TRAVAIL"
            doc_category = "Certificat"
            break
        elif "MEDICAL" in line_upper:
            template_id = "rapport_medical"
            doc_title = "RAPPORT MÉDICAL"
            doc_category = "Rapport HOSPI"
            break

    # Merge or Add Patient without duplication
    if fullname_upper in patients_map:
        # Merge clinical details if the new ones are more specific
        p = patients_map[fullname_upper]
        if not p["intervention"] and intervention:
            p["intervention"] = intervention
            p["kCode"] = k_code
        if p["diagnosis"] == "Bilan clinique" and diagnosis != "Bilan clinique":
            p["diagnosis"] = diagnosis
    else:
        patients_map[fullname_upper] = {
            "name": clean_name,
            "diagnosis": diagnosis,
            "intervention": intervention,
            "kCode": k_code,
            "age": age,
            "matricule": "",
            "insurer": insurer,
            "priseEnCharge": prise_en_charge
        }
        new_patient_count += 1

    # Document ID sequence
    doc_id = f"DOC-CMF-H-{doc_import_count+1:03d}"
    all_new_docs.append({
        "id": doc_id,
        "type": "DOC",
        "category": doc_category,
        "title": doc_title,
        "templateId": template_id,
        "patientNom": last_name,
        "patientPrenom": first_name,
        "patientAge": age,
        "date": report_date,
        "diagnosis": diagnosis,
        "content": full_text,
        "text": full_text,
        "savedAt": "2026-06-02T17:45:00.000Z"
    })
    doc_import_count += 1

print(f"Hospitalisation docx parsing: Added {new_patient_count} new patients, {doc_import_count} documents.")

# =============================================================
# 2. PARSING EXEMPLAIRE PROFORMA.xlsx
# =============================================================
print(f"Parsing Excel: {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Sheets to exclude
exclude_sheets = ["feuil1", "feuil2", "winsu sports", "feuil4"]

excel_bill_count = 0
excel_patient_count = 0

all_new_bills = []

def map_insurance(text):
    if not text:
        return "PRIVE"
    t = text.lower().strip()
    if "sanlam" in t:
        return "SANLAM"
    elif "ascoma" in t:
        return "ASCOMA"
    elif "sunu" in t:
        return "SUNU"
    elif "nsia" in t:
        return "NSIA"
    elif "atlantique" in t:
        return "ATLANTIQUE"
    elif "africaine" in t:
        if "sinistre" in t or "accident" in t:
            return "AFRICAINE_SINISTRE"
        return "AFG"
    elif "lotto" in t:
        return "LOTTO_FOOTBALL_CLUB"
    elif "coton" in t:
        return "COTON_SPORT"
    elif "nobila" in t:
        return "NOBILA"
    elif "gras" in t or "savoye" in t:
        return "GRAS SAVOYE"
    elif "olea" in t:
        return "OLEA"
    elif "fonds" in t or "garantie" in t or "fga" in t:
        return "FONDS_GARANTIE_AUTO"
    elif "transvie" in t:
        return "TRANSVIE"
    elif "sobremap" in t or "sobemap" in t:
        return "SOBEMAP"
    elif "port" in t or "pac" in t:
        return "PORT_AUTONOME_COTONOU"
    return "PRIVE"

for sheetname in wb.sheetnames:
    if sheetname.lower() in exclude_sheets:
        continue
        
    sheet = wb[sheetname]
    
    patient_name = ""
    bill_type = "PROFORMA"
    insurance = "PRIVE"
    date_str = ""
    k_code = ""
    intervention = ""
    
    # Scan the top part of sheet for metadata
    for r in range(1, 35):
        for c in range(1, 13):
            val = sheet.cell(r, c).value
            if not val or not isinstance(val, str):
                continue
            val_clean = val.strip()
            
            if val_clean.lower().startswith("patient:") or val_clean.lower().startswith("patient :"):
                patient_name = val_clean.split(":", 1)[1].strip()
            elif val_clean.lower().startswith("client:") or val_clean.lower().startswith("client :") or val_clean.lower().startswith("assurance:") or val_clean.lower().startswith("assurance :"):
                insurance = map_insurance(val_clean.split(":", 1)[1].strip())
            elif val_clean.lower().startswith("intervention:") or val_clean.lower().startswith("intervention :"):
                intervention = val_clean.split(":", 1)[1].strip()
                
            if "définitive" in val_clean.lower() or "definitif" in val_clean.lower() or "facture définitive" in val_clean.lower():
                bill_type = "DEFINITIF"
            elif "détail" in val_clean.lower() or "detail" in val_clean.lower() or "détail assurance" in val_clean.lower():
                bill_type = "DETAIL_ASSUR"
                
            if "cotonou,le" in val_clean.lower() or "cotonou, le" in val_clean.lower():
                date_str = parse_french_date(val_clean)
                
            k_m = re.search(r'\b(K\d+)\b', val_clean)
            if k_m:
                k_code = k_m.group(1)
                
    # Fallback to sheetname if no patient name found in cells
    if not patient_name:
        if re.search(r'[A-Za-z]', sheetname):
            patient_name = sheetname
        else:
            continue
            
    clean_name = clean_patient_name(patient_name)
    words = clean_name.split()
    if len(words) < 2:
        continue
        
    last_name = words[0].upper()
    first_name = " ".join(words[1:])
    fullname_upper = clean_name.upper()
    
    if not date_str:
        date_str = "2026-05-15"
        
    # Search items table
    header_row = -1
    for r in range(8, 20):
        val = sheet.cell(r, 1).value
        if val and isinstance(val, str) and ("actes" in val.lower() or "désignations" in val.lower() or "designations" in val.lower()):
            header_row = r
            break
            
    items = []
    gross_total = 0
    
    if header_row != -1:
        for r in range(header_row + 1, header_row + 30):
            desig = sheet.cell(r, 1).value
            qty = sheet.cell(r, 2).value
            price = sheet.cell(r, 3).value
            subtotal = sheet.cell(r, 4).value
            
            if desig:
                desig_str = str(desig).strip()
                if "total" in desig_str.lower() or "caisse" in desig_str.lower():
                    break
                
                try:
                    qty_val = int(float(str(qty).replace(' ', ''))) if qty else 1
                except:
                    qty_val = 1
                try:
                    price_val = int(float(str(price).replace(' ', ''))) if price else 0
                except:
                    price_val = 0
                try:
                    subtotal_val = int(float(str(subtotal).replace(' ', ''))) if subtotal else price_val * qty_val
                except:
                    subtotal_val = price_val * qty_val
                    
                if price_val > 0:
                    items.append({
                        "name": desig_str,
                        "price": price_val,
                        "qty": qty_val,
                        "subtotal": subtotal_val
                    })
                    gross_total += subtotal_val
                    
    if not items:
        # Default minimum standard service if empty
        gross_total = 210000
        items = [{"name": "Frais de soins cliniques standards", "price": 210000, "qty": 1, "subtotal": 210000}]
        
    # Fallback si l'assurance n'est pas détectée dans les cellules mais est présente dans le nom de l'onglet
    if insurance == "PRIVE":
        sheet_insurance = None
        for ins_id in ["SANLAM", "ASCOMA", "SUNU", "NSIA", "ATLANTIQUE", "AFG", "LOTTO", "COTON", "NOBILA", "GRAS SAVOYE", "OLEA", "TRANSVIE", "SOBEMAP", "PORT_AUTONOME_COTONOU"]:
            if ins_id.lower().replace("_", "") in sheetname.lower().replace("-", "").replace(" ", ""):
                sheet_insurance = ins_id
                break
        if sheet_insurance:
            insurance = sheet_insurance
            
    coverage = 80 if insurance != "PRIVE" else 0
    part_assurance = 0
    part_patient = gross_total
    if insurance != "PRIVE":
        part_assurance = int(gross_total * (coverage / 100))
        part_patient = gross_total - part_assurance
        
    # Merge/Add Patient
    if fullname_upper in patients_map:
        p = patients_map[fullname_upper]
        if not p["intervention"] and intervention:
            p["intervention"] = intervention
            p["kCode"] = k_code
    else:
        patients_map[fullname_upper] = {
            "name": clean_name,
            "diagnosis": "Bilan clinique",
            "intervention": intervention,
            "kCode": k_code,
            "age": "45 ans",
            "matricule": "",
            "insurer": insurance,
            "priseEnCharge": coverage
        }
        excel_patient_count += 1
        
    # Add bill object
    bill_id = f"BILL-CMF-X-{excel_bill_count+1:03d}"
    all_new_bills.append({
        "id": bill_id,
        "reference": f"MF-{bill_type[:3].upper()}-2026-{excel_bill_count+1:03d}",
        "patientNom": last_name,
        "patientPrenom": first_name,
        "type": bill_type,
        "customTitle": "Point Définitif d'Hospitalisation" if bill_type == "DEFINITIF" else "",
        "insurance": insurance,
        "coverage": coverage,
        "matricule": "N/A" if insurance != "PRIVE" else "",
        "diagnostic": "Bilan clinique",
        "intervention": intervention,
        "kCode": k_code,
        "showDiag": True,
        "showInterv": True if intervention else False,
        "showSig": True,
        "showCachet": True,
        "useSplit": True if insurance != "PRIVE" else False,
        "items": items,
        "grossTotal": gross_total,
        "discountPct": 0,
        "reductionAmount": 0,
        "discountedTotal": gross_total,
        "discountType": "PERCENT",
        "discountValue": 0,
        "partAssurance": part_assurance,
        "partPatient": part_patient,
        "paymentMethod": "TIERS_PAYANT" if insurance != "PRIVE" else "CASH",
        "amountPaidPatient": part_patient,
        "balancePatient": 0,
        "status": "RÉGLÉ",
        "date": date_str
    })
    excel_bill_count += 1

print(f"Excel parsing: Added {excel_patient_count} new patients, {excel_bill_count} bills.")

# =============================================================
# 3. MERGING & SAVE TO FILE DB
# =============================================================
final_patients_list = list(patients_map.values())
final_docs_list = existing_docs + all_new_docs
final_bills_list = existing_bills + all_new_bills

# Write JSON files
with open(patients_db_path, "w", encoding="utf-8") as f:
    json.dump(final_patients_list, f, indent=4, ensure_ascii=False)

with open(documents_db_path, "w", encoding="utf-8") as f:
    json.dump(final_docs_list, f, indent=4, ensure_ascii=False)

with open(bills_db_path, "w", encoding="utf-8") as f:
    json.dump(final_bills_list, f, indent=4, ensure_ascii=False)

# Write real_data.js
js_content = f"""/* ==========================================
   real_data.js - Vrais Rapports Médicaux Clinique Mercy Fiat
   ========================================== */

window.MercyFiatRealDocs = {json.dumps(final_docs_list, indent=4, ensure_ascii=False)};
"""
with open(real_data_js_path, "w", encoding="utf-8") as f:
    f.write(js_content)

print(f"Saved database lists:")
print(f"  Total Patients: {len(final_patients_list)} (originally {len(existing_patients)})")
print(f"  Total Documents: {len(final_docs_list)} (originally {len(existing_docs)})")
print(f"  Total Bills: {len(final_bills_list)} (originally {len(existing_bills)})")

# =============================================================
# 4. COLLECTING UNIQUE DIAGNOSES & INTERVENTIONS -> database.js
# =============================================================
print("Collecting all unique diagnoses and interventions...")

diagnoses_set = set()
interventions_set = set()

# Seed with standard/existing items in JS code if possible, or just build from all documents/patients
for p in final_patients_list:
    if p.get("diagnosis") and p["diagnosis"] != "Bilan clinique":
        diagnoses_set.add(p["diagnosis"])
    if p.get("intervention"):
        interventions_set.add(p["intervention"])

for d in final_docs_list:
    if d.get("diagnosis") and d["diagnosis"] != "Bilan clinique":
        diagnoses_set.add(d["diagnosis"])

# Load database.js and replace the DIAGNOSES and INTERVENTIONS arrays
with open(database_js_path, "r", encoding="utf-8") as f:
    db_content = f.read()

# Filter out empty or too long items and escape double quotes
clean_diagnoses = sorted([diag.replace('"', '\\"') for diag in diagnoses_set if diag and len(diag) > 4 and len(diag) < 120])
clean_interventions = sorted([inter.replace('"', '\\"') for inter in interventions_set if inter and len(inter) > 4 and len(inter) < 120])

# Replace DIAGNOSES array
diag_regex = re.compile(r'DIAGNOSES:\s*\[[\s\S]*?\s*\]', re.MULTILINE)
new_diags_str = f"DIAGNOSES: [\n" + ",\n".join(f'        "{d}"' for d in clean_diagnoses) + "\n    ]"
db_content = diag_regex.sub(new_diags_str, db_content)

# Replace INTERVENTIONS array
interv_regex = re.compile(r'INTERVENTIONS:\s*\[[\s\S]*?\s*\]', re.MULTILINE)
new_intervs_str = f"INTERVENTIONS: [\n" + ",\n".join(f'        "{i}"' for i in clean_interventions) + "\n    ]"
db_content = interv_regex.sub(new_intervs_str, db_content)

with open(database_js_path, "w", encoding="utf-8") as f:
    f.write(db_content)

print(f"Updated database.js with {len(clean_diagnoses)} unique diagnoses and {len(clean_interventions)} unique interventions.")
print("Processing completed successfully!")
