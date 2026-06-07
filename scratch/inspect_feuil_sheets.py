import openpyxl
import os
import re

excel_path = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\EXEMPLAIRE PROFORMA.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Find sheets like Feuil2 (86) and Feuil2 (7)
target_sheets = [s for s in wb.sheetnames if "feuil" in s.lower() and s.lower() not in ["feuil1", "feuil2", "feuil4"]]

print(f"Total feuil sheets found: {len(target_sheets)}")
print("Sample sheet names:", target_sheets[:10])

# Let's inspect some of these sheets
for sname in ["Feuil2 (86)", "Feuil2 (7)", "Feuil2 (98)"]:
    if sname not in wb.sheetnames:
        print(f"Sheet {sname} not in workbook.")
        continue
        
    print(f"\n--- INSPECTION OF SHEET: {sname} ---")
    sheet = wb[sname]
    for r in range(1, 15):
        row_vals = [sheet.cell(r, c).value for c in range(1, 10)]
        if any(row_vals):
            print(f"Row {r:2d}: {row_vals}")
