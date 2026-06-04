"""
VÉRIFICATION COMPLÈTE : Source Excel → Base de données
Lit chaque onglet de EXEMPLAIRE PROFORMA.xlsx, extrait les vraies données,
et compare avec ce qui est dans bills_db.json
"""
import json
import sys
import os
import re
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_main = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")
bills_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "bills_db.json")

print("="*80)
print("VÉRIFICATION DES MONTANTS IMPORTÉS vs SOURCE EXCEL")
print("="*80)
print()

# Charger la base
with open(bills_db_path, encoding='utf-8') as f:
    bills_db = json.load(f)

# Index par nom normalisé
def normalize(s):
    if not s:
        return ""
    return re.sub(r'\s+', ' ', s.strip().upper())

bills_index = {}
for b in bills_db:
    nom = normalize(b.get('patientNom',''))
    prenom = normalize(b.get('patientPrenom',''))
    full = f"{nom} {prenom}".strip()
    if full not in bills_index:
        bills_index[full] = []
    bills_index[full].append(b)

def find_bill(sheet_name, patient_from_sheet):
    """Cherche une facture dans la DB"""
    pn = normalize(patient_from_sheet)
    sn = normalize(sheet_name)
    
    # Essai 1: correspondance exacte
    for key, blist in bills_index.items():
        if pn == key or sn == key:
            return blist
    
    # Essai 2: correspondance partielle (tous les mots du nom de l'onglet dans la clé)
    sn_parts = [w for w in sn.split() if len(w) > 2]
    for key, blist in bills_index.items():
        if len(sn_parts) >= 2 and all(w in key for w in sn_parts[:2]):
            return blist
    
    # Essai 3: correspondance partielle sur les 2 premiers mots du patient
    pn_parts = [w for w in pn.split() if len(w) > 2]
    for key, blist in bills_index.items():
        if len(pn_parts) >= 2 and all(w in key for w in pn_parts[:2]):
            return blist
    
    return None

def parse_date_cell(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = '20' + y
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except:
            return None
    return None

def cell_num(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(str(val).replace(' ', '').replace('\xa0', '').replace(',', '.')))
    except:
        return 0

# Charger l'Excel principal
print(f"Lecture: {excel_main}")
wb = openpyxl.load_workbook(excel_main, data_only=True)
print(f"Onglets: {len(wb.sheetnames)} feuilles\n")

IGNORE = {'feuil1','feuil2','feuil3','feuil4','winsu sports','sheet1','config','param','aide','tarifs','forfaits','liste','nomenclature'}

results = {
    'ok': [],
    'montant_different': [],
    'date_inventee': [],
    'items_inventes': [],
    'non_trouve_en_db': [],
    'onglet_vide': []
}

for sheetname in wb.sheetnames:
    if sheetname.lower().strip() in IGNORE:
        continue
    
    ws = wb[sheetname]
    
    # === EXTRACTION DES DONNÉES SOURCE ===
    patient_name = ""
    date_source = None
    items_source = []
    total_source = 0
    intervention_source = ""
    diagnostic_source = ""
    insurance_source = "PRIVE"
    header_row = -1
    
    all_rows = list(ws.iter_rows(values_only=True, max_row=100))
    if not all_rows:
        results['onglet_vide'].append(sheetname)
        continue
    
    for ri, row in enumerate(all_rows[:40]):
        for ci, val in enumerate(row):
            if val is None:
                continue
            if isinstance(val, (int, float)):
                continue
            s = str(val).strip()
            sl = s.lower()
            
            if sl.startswith("patient") and ':' in s:
                patient_name = s.split(':', 1)[1].strip()
            elif sl.startswith("client") and ':' in s:
                insurance_source = s.split(':', 1)[1].strip()
            elif sl.startswith("intervention") and ':' in s:
                intervention_source = s.split(':', 1)[1].strip()
            elif sl.startswith("diagnostic") and ':' in s:
                diagnostic_source = s.split(':', 1)[1].strip()
            elif 'cotonou' in sl and ('le' in sl or ',' in s):
                d = parse_date_cell(s)
                if d:
                    date_source = d
            
            # Détecter la ligne d'en-tête du tableau
            if any(kw in sl for kw in ['désignation', 'designation', 'actes', 'libellé', 'libelle']):
                if header_row == -1:
                    header_row = ri
    
    # Si pas de date → regarder si une cellule datetime existe dans la feuille
    if not date_source:
        for row in all_rows[:40]:
            for val in row:
                if isinstance(val, datetime):
                    date_source = val.strftime('%Y-%m-%d')
                    break
            if date_source:
                break
    
    # Si toujours pas de patient → utiliser le nom de l'onglet
    if not patient_name:
        if re.search(r'[A-Za-z]', sheetname):
            patient_name = sheetname
        else:
            results['onglet_vide'].append(sheetname)
            continue
    
    # === EXTRAIRE LES ITEMS DU TABLEAU ===
    if header_row >= 0:
        for row in all_rows[header_row+1:header_row+40]:
            if not row or row[0] is None:
                continue
            d0 = str(row[0]).strip() if row[0] else ''
            if not d0:
                continue
            d0_low = d0.lower()
            if any(kw in d0_low for kw in ['total', 'caisse', 'à payer', 'montant total', 'reste']):
                # Capturer le total final
                for v in row:
                    n = cell_num(v)
                    if n > 50000:
                        total_source = max(total_source, n)
                break
            
            # Ligne d'item
            price = cell_num(row[2] if len(row) > 2 else None)
            qty = cell_num(row[1] if len(row) > 1 else None) or 1
            subtotal = cell_num(row[3] if len(row) > 3 else None) or (price * qty)
            
            if price > 0 and len(d0) > 2:
                items_source.append({
                    'name': d0,
                    'price': price,
                    'qty': qty,
                    'subtotal': subtotal
                })
                total_source += subtotal
    
    # Chercher aussi le total directement (ligne "TOTAL" ou "A PAYER")
    for row in all_rows:
        for ci, val in enumerate(row):
            if val and isinstance(val, str) and any(kw in val.lower() for kw in ['total général', 'total general', 'a payer', 'à payer']):
                # Le montant est souvent à droite
                for v in row[ci:ci+5]:
                    n = cell_num(v)
                    if n > 50000:
                        total_source = n
                        break
    
    # === CHERCHER DANS LA BASE ===
    db_matches = find_bill(sheetname, patient_name)
    
    if not db_matches:
        results['non_trouve_en_db'].append({
            'onglet': sheetname,
            'patient': patient_name,
            'total_source': total_source,
            'date_source': date_source
        })
        continue
    
    # Prendre la facture la plus pertinente (même date si possible)
    db_bill = db_matches[0]
    for b in db_matches:
        if b.get('date') == date_source:
            db_bill = b
            break
    
    db_total = db_bill.get('grossTotal', 0)
    db_date = db_bill.get('date', '')
    db_items = db_bill.get('items', [])
    
    errors = []
    
    # Vérifier les items inventés (forfait 210 000)
    if len(db_items) == 1 and db_items[0].get('price') == 210000 and db_items[0].get('name') == 'Frais de soins cliniques standards':
        if items_source:
            errors.append(f"ITEMS INVENTÉS: DB a forfait 210,000 mais source a {len(items_source)} ligne(s) réelle(s)")
        elif total_source == 0:
            errors.append("ITEMS INVENTÉS: forfait 210,000 par défaut (source vide/illisible)")
    
    # Vérifier les montants
    if total_source > 0 and abs(db_total - total_source) > 5000:
        errors.append(f"MONTANT: source={total_source:,} FCFA | DB={db_total:,} FCFA | Écart={abs(db_total-total_source):,}")
    
    # Vérifier la date inventée
    if db_date == '2026-06-01' and date_source and date_source != '2026-06-01':
        errors.append(f"DATE INVENTÉE: DB=2026-06-01 | source={date_source}")
    elif not date_source and db_date == '2026-06-01':
        errors.append("DATE INVENTÉE: 2026-06-01 (pas de date dans la source)")
    
    nom_display = f"{db_bill.get('patientNom','')} {db_bill.get('patientPrenom','')}".strip()
    
    if errors:
        entry = {
            'onglet': sheetname,
            'patient_db': nom_display,
            'patient_source': patient_name,
            'date_source': date_source,
            'date_db': db_date,
            'total_source': total_source,
            'total_db': db_total,
            'items_source': len(items_source),
            'items_db': len(db_items),
            'errors': errors
        }
        if 'MONTANT' in str(errors):
            results['montant_different'].append(entry)
        if 'INVENTÉS' in str(errors):
            results['items_inventes'].append(entry)
        if 'DATE' in str(errors):
            results['date_inventee'].append(entry)
    else:
        results['ok'].append(nom_display)

# === RAPPORT FINAL ===
print(f"\n{'='*80}")
print(f"RÉSUMÉ DE LA VÉRIFICATION")
print(f"{'='*80}")
print(f"✅ Factures OK (montants cohérents) : {len(results['ok'])}")
print(f"❌ Montants différents              : {len(results['montant_different'])}")
print(f"⚠️  Items/forfaits inventés          : {len(results['items_inventes'])}")
print(f"📅 Dates inventées (2026-06-01)     : {len(results['date_inventee'])}")
print(f"🔍 Non trouvés en DB                : {len(results['non_trouve_en_db'])}")
print(f"📭 Onglets vides/ignorés            : {len(results['onglet_vide'])}")

if results['montant_different']:
    print(f"\n{'='*80}")
    print("DÉTAIL — MONTANTS DIFFÉRENTS")
    print(f"{'='*80}")
    for e in results['montant_different']:
        print(f"\n🔴 [{e['onglet']}] → {e['patient_db']}")
        for err in e['errors']:
            if 'MONTANT' in err:
                print(f"   {err}")

if results['items_inventes']:
    print(f"\n{'='*80}")
    print("DÉTAIL — ITEMS/FORFAITS INVENTÉS")
    print(f"{'='*80}")
    for e in results['items_inventes']:
        print(f"\n⚠️  [{e['onglet']}] → {e['patient_db']}")
        for err in e['errors']:
            if 'INVENT' in err:
                print(f"   {err}")

if results['date_inventee']:
    print(f"\n{'='*80}")
    print("DÉTAIL — DATES INVENTÉES")
    print(f"{'='*80}")
    for e in results['date_inventee']:
        print(f"\n📅 [{e['onglet']}] → {e['patient_db']}")
        for err in e['errors']:
            if 'DATE' in err:
                print(f"   {err}")

if results['non_trouve_en_db']:
    print(f"\n{'='*80}")
    print("ONGLETS NON TROUVÉS EN BASE")
    print(f"{'='*80}")
    for e in results['non_trouve_en_db']:
        print(f"  [{e['onglet']}] | Patient source: {e['patient']} | Total: {e['total_source']:,} FCFA | Date: {e['date_source']}")

print(f"\n{'='*80}")
print("FIN DE VÉRIFICATION")
print(f"{'='*80}")
