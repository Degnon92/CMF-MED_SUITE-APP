"""
INSPECTION PROFONDE DES FEUILLES Feuil2 (X)
Lit le contenu cellule par cellule pour trouver le vrai nom du patient
et importe correctement les données.
"""
import json
import sys
import os
import re
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_main = os.path.join(workspace_dir, "PROFORMA CHIRURGIE", "EXEMPLAIRE PROFORMA.xlsx")
bills_db_path = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop", "bills_db.json")

with open(bills_db_path, encoding='utf-8') as f:
    bills_db = json.load(f)

print(f"DB actuelle: {len(bills_db)} factures\n")

# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────────────────────────────────────
def cell_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return int(val)
    try:
        return int(float(str(val).replace(' ','').replace('\xa0','').replace(',','.')))
    except: return 0

def parse_date_val(val):
    if not val: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2: y = '20' + y
        try: return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except: return None
    return None

def normalize(s):
    if not s: return ""
    return re.sub(r'\s+', ' ', str(s).strip().upper())

def map_insurance(text):
    if not text: return "PRIVE"
    t = text.lower().strip()
    if "sanlam" in t: return "SANLAM"
    if "ascoma" in t: return "ASCOMA"
    if "sunu" in t: return "SUNU"
    if "nsia" in t: return "NSIA"
    if "atlantique" in t: return "ATLANTIQUE"
    if "africaine" in t or " afg" in t or t.endswith(" aa") or " aa " in t: return "AFG"
    if "loto" in t or "lotto" in t: return "LOTTO_FOOTBALL_CLUB"
    if "coton" in t: return "COTON_SPORT"
    if "olea" in t: return "OLEA"
    if "gras" in t or "savoye" in t: return "GRAS SAVOYE"
    if "sobemap" in t: return "SOBEMAP"
    if "fga" in t or "garantie" in t: return "FONDS_GARANTIE_AUTO"
    return "PRIVE"

def split_nom_prenom(full):
    full = re.sub(r'\s+', ' ', full.strip())
    parts = full.split(' ')
    if len(parts) == 1: return full.upper(), ""
    nom_parts, prenom_parts = [], []
    for p in parts:
        if p.isupper() and len(p) > 1: nom_parts.append(p)
        else: prenom_parts.append(p)
    if not nom_parts: return parts[0].upper(), ' '.join(parts[1:])
    if not prenom_parts and len(parts) > 1: return parts[0].upper(), ' '.join(parts[1:])
    return ' '.join(nom_parts).upper(), ' '.join(prenom_parts)

# Mots à rejeter (pas des noms de patients)
REJECT_WORDS = {
    'clinique','mercy','fiat','cotonou','patient','nom','prenom','date',
    'total','montant','diagnostic','intervention','assurance','client',
    'facture','proforma','designation','actes','quantite','prix',
    'benin','republique','rue','tel','email','bp','boite',
    'docteur','medecin','chirurgie','orthopedique','traumatologie',
    'honoraires','frais','soins','hospitalisation','consultation',
    'point','definitif','detail','accord','assur','prive',
    'hebergement','salle','bloc','anesthesie','infirmier',
    'medicaments','materiel','labo','radio','echo','scanner',
    'feuil','sheet','page','copie'
}

def is_likely_patient_name(s):
    """Détermine si une chaîne ressemble à un nom de patient."""
    if not s or len(s) < 4 or len(s) > 60: return False
    s_up = s.upper().strip()
    # Doit avoir au moins 2 mots
    words = s_up.split()
    if len(words) < 2: return False
    # Doit être principalement alphabétique
    alpha_chars = sum(1 for c in s if c.isalpha() or c == ' ' or c == '-')
    if alpha_chars / len(s) < 0.7: return False
    # Ne doit pas contenir de mots rejetés
    words_low = [w.lower() for w in words]
    if any(w in REJECT_WORDS for w in words_low): return False
    # Ne doit pas commencer par des mots-clés cliniques
    first_word = words_low[0]
    if first_word in {'dr','docteur','mr','mme','melle','m','le','la','de','du','des','une','un'} and len(words) == 2:
        return False
    # Doit avoir des lettres en majuscule (noms souvent en caps)
    if not any(c.isupper() for c in s): return False
    return True

# Patterns pour trouver un nom après une étiquette
PATIENT_LABEL_PATTERNS = [
    r'(?:nom\s+(?:du\s+)?patient|patient|nom\s+(?:et\s+)?pr[eé]nom|nom\s*:)\s*[:\-]?\s*(.{4,50})',
    r'(?:patiente?|nom)\s*[:\-]\s*(.{4,50})',
    r'(?:assuré|assure|bénéficiaire|beneficiaire)\s*[:\-]\s*(.{4,50})',
    r'(?:monsieur|madame|mademoiselle|m\.|mme\.?|mr\.?)\s+([A-Z][A-Za-zÀ-ÿ\s\-]{3,40})',
]

def find_patient_name_in_sheet(ws):
    """
    Cherche le nom du patient dans tous les contenus de la feuille.
    Stratégie multi-niveaux.
    """
    all_rows = list(ws.iter_rows(values_only=True, max_row=100))
    
    # NIVEAU 1: Chercher les patterns "Patient:", "Nom:", etc.
    for row in all_rows[:50]:
        for ci, val in enumerate(row):
            if not val or not isinstance(val, str): continue
            s = val.strip()
            sl = s.lower()
            
            # Pattern direct: la cellule contient "Patient: NOM PRENOM"
            for pattern in PATIENT_LABEL_PATTERNS:
                m = re.search(pattern, s, re.IGNORECASE)
                if m:
                    candidate = m.group(1).strip().split('\n')[0].strip()
                    candidate = re.sub(r'[:\-_]+$', '', candidate).strip()
                    if is_likely_patient_name(candidate):
                        return candidate
            
            # Pattern: la cellule est un label, et la cellule SUIVANTE est le nom
            if re.match(r'^(?:patient|nom|pr[eé]nom|nom\s+(?:du\s+)?patient)\s*:?\s*$', sl):
                # Chercher dans la même ligne, colonne suivante
                for next_val in row[ci+1:ci+4]:
                    if next_val and isinstance(next_val, str) and is_likely_patient_name(next_val.strip()):
                        return next_val.strip()
                # Chercher dans la ligne suivante, même colonne
                row_idx = all_rows.index(row)
                for next_row in all_rows[row_idx+1:row_idx+3]:
                    if len(next_row) > ci:
                        nv = next_row[ci]
                        if nv and isinstance(nv, str) and is_likely_patient_name(nv.strip()):
                            return nv.strip()
    
    # NIVEAU 2: Chercher les cellules en majuscules dans les 15 premières lignes
    # (souvent les noms de patients sont écrits en MAJUSCULES dans les en-têtes)
    candidates = []
    for row in all_rows[:20]:
        for val in row:
            if not val or not isinstance(val, str): continue
            s = val.strip()
            # Cellule entièrement en majuscules avec 2-4 mots
            words = s.split()
            if (2 <= len(words) <= 5 and 
                all(w.isupper() or w[0].isupper() for w in words if w) and
                is_likely_patient_name(s)):
                candidates.append((s, len(s)))
    
    if candidates:
        # Prendre le candidat le plus long (plus probable d'être un vrai nom complet)
        candidates.sort(key=lambda x: -x[1])
        return candidates[0][0]
    
    # NIVEAU 3: Chercher "CLINIQUE MERCY FIAT" et prendre le texte qui suit
    for ri, row in enumerate(all_rows[:30]):
        for ci, val in enumerate(row):
            if not val or not isinstance(val, str): continue
            if 'mercy' in val.lower() or 'clinique' in val.lower():
                # Les lignes juste après l'en-tête de la clinique contiennent souvent le patient
                for next_row in all_rows[ri+1:ri+8]:
                    for nv in next_row:
                        if nv and isinstance(nv, str) and is_likely_patient_name(nv.strip()):
                            return nv.strip()
    
    return None

def extract_full_sheet(ws, sheetname, patient_override=None):
    """Extraction complète d'un onglet."""
    all_rows = list(ws.iter_rows(values_only=True, max_row=120))
    if not all_rows: return None

    patient_name = patient_override or ""
    date_source = None
    items = []
    total_source = 0
    intervention = ""
    diagnostic = ""
    insurance_txt = ""
    k_code = ""
    bill_type = "PROFORMA"
    header_row = -1

    for ri, row in enumerate(all_rows[:60]):
        for ci, val in enumerate(row):
            if val is None: continue
            if isinstance(val, datetime):
                if not date_source: date_source = val.strftime('%Y-%m-%d')
                continue
            if not isinstance(val, str): continue
            s = val.strip()
            sl = s.lower()

            if not patient_name:
                if re.match(r'patient\s*:', sl, re.I):
                    v = s.split(':', 1)[1].strip()
                    if v and len(v) > 2: patient_name = v
            if re.match(r'(client|assurance)\s*:', sl, re.I):
                insurance_txt = s.split(':', 1)[1].strip()
            if re.match(r'intervention\s*:', sl, re.I):
                v = s.split(':', 1)[1].strip()
                if v: intervention = v
            if re.match(r'diagnostic\s*:', sl, re.I):
                v = s.split(':', 1)[1].strip()
                if v: diagnostic = v
            if ('cotonou' in sl) and not date_source:
                d = parse_date_val(s)
                if d: date_source = d
            if any(k in sl for k in ['définitif','definitif','point d\'hospit']):
                bill_type = "DEFINITIF"
            elif any(k in sl for k in ['détail assur','detail assur']):
                bill_type = "DETAIL_ASSUR"
            km = re.search(r'\b(KC?\s*\d+)\b', s)
            if km: k_code = km.group(1)
            if any(k in sl for k in ['désignation','designation','actes','libellé','libelle','prestation']):
                if header_row == -1: header_row = ri

    # Date depuis datetime cells
    if not date_source:
        for row in all_rows[:60]:
            for val in row:
                if isinstance(val, datetime):
                    date_source = val.strftime('%Y-%m-%d')
                    break
            if date_source: break

    if not patient_name: return None

    # Items
    if header_row >= 0:
        for row in all_rows[header_row+1: header_row+60]:
            if not row or row[0] is None: continue
            d0 = str(row[0]).strip() if row[0] else ''
            if not d0: continue
            d0_low = d0.lower()
            if any(k in d0_low for k in ['total','caisse','à payer','a payer','net à','reste','solde','arrêté','sous-total']):
                for v in row:
                    n = cell_num(v)
                    if n > 10000: total_source = max(total_source, n)
                break
            price = cell_num(row[2] if len(row) > 2 else None)
            qty = cell_num(row[1] if len(row) > 1 else None) or 1
            subtot = cell_num(row[3] if len(row) > 3 else None) or (price * qty)
            if price > 0 and len(d0) > 2:
                items.append({'name': d0[:100], 'price': price, 'qty': qty, 'subtotal': subtot})
                if subtot: total_source += subtot

    # Total global
    for row in all_rows:
        for ci, val in enumerate(row):
            if not val or not isinstance(val, str): continue
            if any(k in val.lower() for k in ['total général','total general','à payer','a payer','net à payer']):
                for v in row[ci:]:
                    n = cell_num(v)
                    if n > 10000: total_source = n; break

    if not items and total_source == 0: return None

    insurance = map_insurance(insurance_txt)
    coverage = 80 if insurance != "PRIVE" else 0
    part_assurance = int(total_source * coverage / 100) if insurance != "PRIVE" else 0
    part_patient = total_source - part_assurance
    nom, prenom = split_nom_prenom(patient_name)
    if not nom: return None

    return {
        'patientNom': nom, 'patientPrenom': prenom,
        'date': date_source or '?',
        'items': items, 'grossTotal': total_source,
        'insurance': insurance, 'coverage': coverage,
        'partAssurance': part_assurance, 'partPatient': part_patient,
        'intervention': intervention, 'diagnostic': diagnostic,
        'kCode': k_code, 'billType': bill_type,
        'sheetname': sheetname, 'date_invented': not bool(date_source)
    }

# ─────────────────────────────────────────────────────────────────────────────
# INDEX DB EXISTANT
# ─────────────────────────────────────────────────────────────────────────────
db_by_nom = {}
for b in bills_db:
    n = normalize(b.get('patientNom',''))
    p = normalize(b.get('patientPrenom',''))
    full = f"{n} {p}".strip()
    db_by_nom.setdefault(full, []).append(b)

def already_in_db(nom, prenom, date):
    n = normalize(nom); p = normalize(prenom)
    full = f"{n} {p}".strip()
    # Exact match
    if full in db_by_nom:
        for b in db_by_nom[full]:
            if b.get('date') == date: return True
    # Partiel
    parts = [w for w in n.split() if len(w) > 2]
    for key in db_by_nom:
        if len(parts) >= 2 and all(w in key for w in parts[:2]):
            for b in db_by_nom[key]:
                if b.get('date') == date: return True
    return False

# ─────────────────────────────────────────────────────────────────────────────
# TRAITER LES ONGLETS Feuil2 (X)
# ─────────────────────────────────────────────────────────────────────────────
print("Ouverture Excel...")
wb = openpyxl.load_workbook(excel_main, data_only=True)

feuil2_sheets = [s for s in wb.sheetnames if re.match(r'^feuil\d+(\s*\(.*\))?$', s.lower().strip())]
print(f"Onglets Feuil(X) trouvés: {len(feuil2_sheets)}")
print()

found_names = []
not_found = []
already_exists = []
added_bills = []

max_id = 0
for b in bills_db:
    bid = b.get('id', '')
    m = re.search(r'(\d+)$', bid)
    if m: max_id = max(max_id, int(m.group(1)))
next_id = max_id + 1

for sheetname in feuil2_sheets:
    try:
        ws = wb[sheetname]
        
        # CHERCHER LE NOM DU PATIENT DANS LE CONTENU
        patient_name = find_patient_name_in_sheet(ws)
        
        if not patient_name:
            not_found.append(sheetname)
            print(f"❓ [{sheetname}] → Nom non trouvé")
            # Afficher les 5 premières cellules non-vides pour aide manuelle
            rows = list(ws.iter_rows(values_only=True, max_row=20))
            shown = 0
            for row in rows:
                for val in row:
                    if val and isinstance(val, str) and len(val.strip()) > 3:
                        print(f"   Cellule: '{val.strip()[:80]}'")
                        shown += 1
                        if shown >= 6: break
                if shown >= 6: break
            print()
            continue
        
        # Extraire les données complètes
        data = extract_full_sheet(ws, sheetname, patient_override=patient_name)
        
        if not data:
            not_found.append(sheetname)
            print(f"⚠️  [{sheetname}] → Patient trouvé '{patient_name}' mais données vides")
            continue
        
        nom = data['patientNom']
        prenom = data['patientPrenom']
        date = data['date']
        total = data['grossTotal']
        
        # Vérifier si déjà en DB
        if already_in_db(nom, prenom, date):
            already_exists.append(sheetname)
            print(f"⏭️  [{sheetname}] → {nom} {prenom} ({date}) déjà en DB")
            continue
        
        found_names.append((sheetname, patient_name, total, date))
        print(f"✅ [{sheetname}] → {nom} {prenom} | {total:,} FCFA | {date}")
        
        # Créer la facture
        items = data['items'] if data['items'] else [{'name': 'Frais cliniques', 'price': total, 'qty': 1, 'subtotal': total}]
        
        new_bill = {
            "id": f"BILL-CMF-IMP-{next_id:04d}",
            "reference": f"MF-{data['billType'][:3]}-IMP-{next_id:04d}",
            "patientNom": nom,
            "patientPrenom": prenom,
            "type": data['billType'],
            "customTitle": "Point Définitif d'Hospitalisation" if data['billType'] == "DEFINITIF" else "",
            "insurance": data['insurance'],
            "coverage": data['coverage'],
            "matricule": "",
            "diagnostic": data['diagnostic'],
            "intervention": data['intervention'],
            "kCode": data['kCode'],
            "showDiag": bool(data['diagnostic']),
            "showInterv": bool(data['intervention']),
            "showSig": True,
            "showCachet": True,
            "useSplit": data['insurance'] != "PRIVE",
            "items": items,
            "grossTotal": total,
            "discountPct": 0,
            "reductionAmount": 0,
            "discountedTotal": total,
            "discountType": "PERCENT",
            "discountValue": 0,
            "partAssurance": data['partAssurance'],
            "partPatient": data['partPatient'],
            "paymentMethod": "TIERS_PAYANT" if data['insurance'] != "PRIVE" else "CASH",
            "amountPaidPatient": data['partPatient'],
            "balancePatient": 0,
            "status": "RÉGLÉ",
            "date": date if not data['date_invented'] else "",
            "patientAge": "",
            "diagnosis": data['diagnostic']
        }
        added_bills.append(new_bill)
        next_id += 1

    except Exception as e:
        not_found.append(sheetname)
        print(f"❌ [{sheetname}] → ERREUR: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# SAUVEGARDER
# ─────────────────────────────────────────────────────────────────────────────
if added_bills:
    all_bills = bills_db + added_bills
    with open(bills_db_path, 'w', encoding='utf-8') as f:
        json.dump(all_bills, f, ensure_ascii=False, indent=2)
    print(f"\n✅ {len(added_bills)} factures ajoutées. DB: {len(bills_db)} → {len(all_bills)}")
else:
    print("\nAucune nouvelle facture à ajouter.")

print()
print("="*70)
print(f"RÉSUMÉ — Feuilles Feuil(X)")
print("="*70)
print(f"✅ Nom trouvé + importé  : {len(found_names)}")
print(f"⏭️  Déjà en DB            : {len(already_exists)}")
print(f"❓ Nom introuvable        : {len(not_found)}")
print()
if not_found:
    print("Onglets sans nom trouvable (à traiter manuellement si important):")
    for s in not_found:
        print(f"  - {s}")
