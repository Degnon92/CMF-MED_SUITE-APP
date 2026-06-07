import openpyxl
import os

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")

wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Searching sheets...")
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    found = False
    for r in range(1, 40):
        for c in range(1, 10):
            val = sheet.cell(r, c).value
            if val and "NOUKPOZOUNKOU" in str(val).upper():
                found = True
                break
        if found:
            break
            
    if found:
        print(f"\n===== Found NOUKPOZOUNKOU in Sheet: {sheetname} =====")
        for r in range(1, 20):
            row_vals = []
            for c in range(1, 10):
                val = sheet.cell(r, c).value
                if val is not None:
                    row_vals.append(f"C{c}: {repr(val)}")
            if row_vals:
                print(f"Row {r:02d} -> " + " | ".join(row_vals))
