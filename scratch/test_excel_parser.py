import openpyxl
import re

wb_path = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\EXEMPLAIRE PROFORMA.xlsx"
wb = openpyxl.load_workbook(wb_path, data_only=True)

def parse_sheet(sheet):
    data = {
        "patient_name": "",
        "bill_type": "PROFORMA",
        "insurance": "PRIVE",
        "date": "",
        "k_code": "",
        "intervention": "",
        "items": [],
        "total": 0
    }
    
    # 1. Scan headers (rows 1-15)
    for r in range(1, 16):
        for c in range(1, 6):
            cell_val = sheet.cell(r, c).value
            if not cell_val or not isinstance(cell_val, str):
                continue
            val = cell_val.strip()
            
            # Patient Name
            if val.lower().startswith("patient:"):
                data["patient_name"] = val.replace("Patient:", "").replace("patient:", "").strip()
            elif val.lower().startswith("patient :"):
                data["patient_name"] = val.replace("Patient :", "").replace("patient :", "").strip()
                
            # Client / Insurance
            elif val.lower().startswith("client:"):
                data["insurance"] = val.replace("Client:", "").replace("client:", "").strip()
            elif val.lower().startswith("client :"):
                data["insurance"] = val.replace("Client :", "").replace("client :", "").strip()
                
            # Bill Type
            if "proforma" in val.lower():
                data["bill_type"] = "PROFORMA"
            elif "définitive" in val.lower() or "definitif" in val_clean.lower() if 'val_clean' in locals() else False:
                data["bill_type"] = "DEFINITIF"
                
            # Date
            date_match = re.search(r'(?:le\s+)?(\d{1,2}/\d{1,2}/\d{2,4})', val, re.IGNORECASE)
            if date_match:
                data["date"] = date_match.group(1)
                
            # K Code
            k_match = re.search(r'\b(K\d+)\b', val)
            if k_match:
                data["k_code"] = k_match.group(1)
                
            # Intervention
            if val.lower().startswith("intervention:"):
                data["intervention"] = val.replace("Intervention:", "").replace("intervention:", "").strip()
            elif val.lower().startswith("intervention :"):
                data["intervention"] = val.replace("Intervention :", "").replace("intervention :", "").strip()
                
    # If no patient name found, skip
    if not data["patient_name"]:
        return None
        
    # 2. Find items table
    # Scan for header row (e.g. containing ACTES or DESIGNATIONS)
    header_row = -1
    for r in range(8, 20):
        val = sheet.cell(r, 1).value
        if val and isinstance(val, str) and ("actes" in val.lower() or "désignations" in val.lower() or "designations" in val.lower()):
            header_row = r
            break
            
    if header_row != -1:
        # Scan rows below header until "TOTAL"
        for r in range(header_row + 1, header_row + 25):
            desig = sheet.cell(r, 1).value
            qty = sheet.cell(r, 2).value
            price = sheet.cell(r, 3).value
            subtotal = sheet.cell(r, 4).value
            
            if desig:
                desig_str = str(desig).strip()
                if "total" in desig_str.lower():
                    data["total"] = subtotal or price or 0
                    break
                
                # Clean qty and price
                try:
                    qty = int(float(str(qty).replace(' ', ''))) if qty else 1
                except:
                    qty = 1
                try:
                    price = int(float(str(price).replace(' ', ''))) if price else 0
                except:
                    price = 0
                try:
                    subtotal = int(float(str(subtotal).replace(' ', ''))) if subtotal else price * qty
                except:
                    subtotal = price * qty
                    
                data["items"].append({
                    "name": desig_str,
                    "qty": qty,
                    "price": price,
                    "subtotal": subtotal
                })
                
    return data

for sheetname in ["KOUSSA JACQUES", "WINSU SPORTS", "Feuil2 (10)"]:
    res = parse_sheet(wb[sheetname])
    print(f"\nSheet {sheetname}:")
    print(res)
