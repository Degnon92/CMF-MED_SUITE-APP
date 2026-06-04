"""
inspect_insurance_files.py
Inspecte les fichiers qui semblent contenir des données assurance:
- FACTURE AFG MALADIE 1.xlsx
- CREANCES NSIA.xlsx
- EXEMPLAIRE PROFORMA. LAWANIxlsx.xlsx
"""
import sys
import os

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from pathlib import Path

files_to_check = [
    r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\FACTURE AFG MALADIE 1.xlsx",
    r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\CREANCES NSIA.xlsx",
    r"C:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA. LAWANIxlsx.xlsx",
]

for filepath in files_to_check:
    p = Path(filepath)
    if not p.exists():
        print(f"[NON TROUVE] {p.name}\n")
        continue
    
    print(f"\n{'#'*70}")
    print(f"FICHIER : {p.name}  ({p.stat().st_size/1024:.1f} Ko)")
    print('#'*70)
    
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
        print(f"Onglets : {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  --- Onglet : '{sheet_name}' ({ws.max_row}L x {ws.max_column}C) ---")
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(40, ws.max_row), values_only=True), 1):
                if any(v is not None and str(v).strip() for v in row):
                    vals = []
                    for v in row:
                        s = str(v)[:32] if v is not None else ''
                        vals.append(s)
                    line = ' | '.join(v for v in vals)
                    # Remplacer les chars problematiques
                    line = line.encode('ascii', errors='replace').decode('ascii')
                    print(f"    L{row_idx:2d}: {line[:130]}")
    except Exception as e:
        print(f"  ERREUR: {e}")
