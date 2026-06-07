import openpyxl
import os

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")
wb = openpyxl.load_workbook(excel_path, read_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print("Sheet names:")
for name in wb.sheetnames:
    print(name)
