import openpyxl
import os
import re

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")

wb = openpyxl.load_workbook(excel_path, data_only=True)

feuil_sheets = [s for s in wb.sheetnames if "feuil" in s.lower() and s.lower() not in ["feuil1", "feuil2", "feuil4"]]
print(f"Total Feuil sheets: {len(feuil_sheets)}")

def extract_name_from_value(val):
    if not val or not isinstance(val, str):
        return None
    val_clean = val.strip()
    
    # Strip Patient: / Patient : / Client: / Client :
    for prefix in ["patient:", "patient :", "client:", "client :", "diagnostic:", "diagnostic :", "intervention:", "intervention :"]:
        if val_clean.lower().startswith(prefix):
            return None # Ignore prefixes that don't hold the name itself, or strip them if it holds the name
            
    # Check if it has Patient: embedded
    m = re.search(r'(?:patient|client)\s*:\s*(.*)', val_clean, re.IGNORECASE)
    if m:
        return m.group(1).strip()
        
    # If cell is just uppercase words (patient name typically looks like "LASTNAME Firstname" or "LASTNAME FIRSTNAME")
    # Let's filter out standard layout text like "CLINIQUE MERCY FIAT", "FACTURE PROFORMA", "ACTES/DESIGNATIONS", "TOTAL", etc.
    upper = val_clean.upper()
    junk = ["CLINIQUE", "MEDECINE", "FACTURE", "PROFORMA", "ACTES", "DESIGNATIONS", "TOTAL", "COTINOU", "COTONOU", "CAISSE", "OPTION", "CHAMBRE"]
    if any(j in upper for j in junk):
        return None
        
    # Check if the text contains letters and is likely a name
    if re.search(r'[A-Za-z]', val_clean) and len(val_clean) > 3 and len(val_clean) < 50:
        return val_clean
        
    return None

resolved = 0
unresolved = []

for sheetname in feuil_sheets:
    sheet = wb[sheetname]
    patient_name = None
    diag = None
    interv = None
    
    # Let's scan rows 3, 4, 5 in column 1 (c=1)
    for r in [3, 4, 5]:
        val = sheet.cell(r, 1).value
        if not val or not isinstance(val, str):
            continue
        val_clean = val.strip()
        
        # Check if it specifies Patient
        m_pat = re.search(r'(?:patient|patiente)\s*:\s*(.*)', val_clean, re.IGNORECASE)
        if m_pat:
            patient_name = m_pat.group(1).strip()
            continue
            
        m_diag = re.search(r'diagnostic\s*:\s*(.*)', val_clean, re.IGNORECASE)
        if m_diag:
            diag = m_diag.group(1).strip()
            continue
            
        m_interv = re.search(r'intervention\s*:\s*(.*)', val_clean, re.IGNORECASE)
        if m_interv:
            interv = m_interv.group(1).strip()
            continue
            
        # If it's none of the above, check if it looks like a patient name (usually Row 3 or 4)
        extracted = extract_name_from_value(val_clean)
        if extracted and not patient_name:
            patient_name = extracted
            
    if patient_name:
        resolved += 1
        # Clean any trailing insurer in the name cell e.g., "TADOGBE CLAUDE(ASSUR SANLAM)"
        clean_pat_name = re.sub(r'\(.*?\)', '', patient_name).strip()
        print(f"Sheet: {sheetname} -> Patient: '{clean_pat_name}' | Diag: '{diag}' | Interv: '{interv}'")
    else:
        unresolved.append(sheetname)

print(f"\nSummary: Resolved {resolved}/{len(feuil_sheets)} sheet names.")
if unresolved:
    print(f"Unresolved sheets: {unresolved}")
