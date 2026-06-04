import openpyxl, os

workspace_dir = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheets matching 'KPEK':")
for name in wb.sheetnames:
    if 'KPEK' in name.upper():
        print(f"Sheet name: {name}")
        ws = wb[name]
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        # Print first 20 rows
        for r in range(1, 25):
            row_vals = [ws.cell(r, c).value for c in range(1, 12)]
            if any(row_vals):
                print(f"  Row {r:02d}: {row_vals}")
        print("-" * 50)
