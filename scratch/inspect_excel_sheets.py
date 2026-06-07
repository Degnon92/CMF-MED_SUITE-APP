import openpyxl
import re
import os

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")

wb = openpyxl.load_workbook(excel_path, data_only=True)
print(f"Total sheet names: {len(wb.sheetnames)}")

sheet_details = []
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    patient_name = ""
    client = ""
    total = 0
    
    # Check some cells for patient name
    for r in range(1, 35):
        for c in range(1, 6):
            val = sheet.cell(r, c).value
            if not val or not isinstance(val, str):
                continue
            val_clean = val.strip()
            if val_clean.lower().startswith("patient:") or val_clean.lower().startswith("patient :"):
                patient_name = val_clean.split(":", 1)[1].strip()
            elif val_clean.lower().startswith("client:") or val_clean.lower().startswith("client :"):
                client = val_clean.split(":", 1)[1].strip()
                
    sheet_details.append({
        "sheetname": sheetname,
        "patient_name": patient_name,
        "client": client
    })

print("\nSheet Details (first 100):")
for idx, sd in enumerate(sheet_details):
    print(f"{idx+1:02d}. Sheet: '{sd['sheetname']}' | Cell Patient: '{sd['patient_name']}' | Cell Client: '{sd['client']}'")
