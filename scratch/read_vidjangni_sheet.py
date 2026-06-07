import openpyxl
import os

excel_file = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx"
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    found = False
    for sheetname in wb.sheetnames:
        if "VIDJANGNI" in sheetname.upper():
            print(f"Found sheet '{sheetname}' in '{excel_file}'")
            sheet = wb[sheetname]
            found = True
            for r in range(1, 20):
                row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
                if any(row_vals):
                    print(f"Row {r:02d}: {row_vals}")
    if not found:
        print("VIDJANGNI sheet not found in EXEMPLAIRE PROFORMA.xlsx")
else:
    print(f"File not found: {excel_file}")
