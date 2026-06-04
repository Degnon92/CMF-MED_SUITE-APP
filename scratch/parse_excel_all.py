import openpyxl
import re

wb_path = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\EXEMPLAIRE PROFORMA.xlsx"
wb = openpyxl.load_workbook(wb_path, data_only=True)

valid_sheets = []
for name in wb.sheetnames:
    sheet = wb[name]
    
    # Try to find a cell containing "patient:"
    patient_name = None
    bill_type = "PROFORMA"
    date_str = None
    k_code = ""
    interventions = []
    
    # Scan the first 10 rows
    for r in range(1, 12):
        for c in range(1, 4):
            val = sheet.cell(r, c).value
            if val and isinstance(val, str):
                val_clean = val.strip()
                if val_clean.lower().startswith("patient:"):
                    patient_name = val_clean.replace("Patient:", "").replace("patient:", "").strip()
                elif "proforma" in val_clean.lower():
                    bill_type = "PROFORMA"
                elif "définitive" in val_clean.lower() or "definitif" in val_clean.lower():
                    bill_type = "DEFINITIF"
                elif "cotonou,le" in val_clean.lower() or "cotonou, le" in val_clean.lower():
                    date_str = val_clean
                elif re.match(r'^K\d+$', val_clean):
                    k_code = val_clean
                elif val_clean.startswith("-") or val_clean.startswith("•"):
                    interventions.append(val_clean.strip("- •"))
                    
    if patient_name:
        valid_sheets.append((name, patient_name, bill_type, k_code, len(interventions)))

print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Valid patient sheets found: {len(valid_sheets)}")
print("First 15 sheets details:")
for item in valid_sheets[:15]:
    print(item)
