import openpyxl
import os

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)

found = False
for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val and isinstance(val, str) and "KELLY" in val.upper():
                print(f"Match found in sheet '{name}', cell ({r},{c}): {val}")
                found = True

if not found:
    print("KELLY not found in any sheet.")
