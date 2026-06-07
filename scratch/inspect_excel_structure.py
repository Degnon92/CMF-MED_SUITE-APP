"""
inspect_excel_structure.py
Inspecte la structure de l'exemplaire proforma Excel KPEKPASSI
pour comprendre toutes les colonnes disponibles dans chaque onglet.
"""
import sys
try:
    import openpyxl
except ImportError:
    print("Installation d'openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from pathlib import Path

# Fichier de test KPEKPASSI
EXCEL_FILE = Path(r"C:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\MercyFiatMedSuiteDesktop\Facture_PARAISO_PROFORMA.xlsx")

if not EXCEL_FILE.exists():
    # Chercher dans le dossier parent
    parent = Path(r"C:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE")
    files = list(parent.rglob("EXEMPLAIRE*.xlsx")) + list(parent.rglob("KPEK*.xlsx")) + list(parent.rglob("*.xlsx"))
    if files:
        EXCEL_FILE = files[0]
        print(f"Fichier trouvé : {EXCEL_FILE}")
    else:
        print("Aucun fichier Excel trouvé!")
        sys.exit(1)

print(f"\n=== Inspection de : {EXCEL_FILE.name} ===\n")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"Onglets disponibles : {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"--- Onglet : '{sheet_name}' ({ws.max_row} lignes x {ws.max_column} colonnes) ---")
    
    # Afficher les 25 premières lignes
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=True), 1):
        # Ne pas afficher les lignes entièrement vides
        if any(v is not None for v in row):
            vals = [str(v)[:30] if v is not None else '' for v in row]
            print(f"  L{row_idx:2d}: {' | '.join(v for v in vals if v)[:120]}")
    print()
