import openpyxl, os, re

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)

from reimport_bills_from_excel import parse_french_date, detect_insurer, detect_bill_type_from_sheet

for name in wb.sheetnames:
    if 'KPEK' not in name.upper():
        continue
    ws = wb[name]
    print(f"Sheet: '{name}'")
    all_rows = list(ws.iter_rows(values_only=True))
    rows_text = ' '.join(str(v) for row in all_rows for v in row if v)
    bill_type = detect_bill_type_from_sheet(name, rows_text)
    
    # Let's search for Cotonou / Date
    date_val = None
    for row in all_rows:
        row_str = ' '.join(str(v) for v in row if v)
        if "cotonou" in row_str.lower():
            for v in row:
                if v and isinstance(v, str) and "cotonou" in v.lower():
                    date_val = parse_french_date(v)
                    
    # Find total
    total_val = None
    for row in all_rows:
        if row and row[0] and 'total' in str(row[0]).lower():
            # Print non-empty items in row
            total_val = [v for v in row if v is not None]
            break
            
    print(f"  Detected Type : {bill_type}")
    print(f"  Detected Date : {date_val}")
    print(f"  Total Row     : {total_val}")
    print(f"  First 3 items :")
    # find where acts start
    started = False
    count = 0
    for row in all_rows:
        row_str = ' '.join(str(v) for v in row if v)
        if any(k in row_str.lower() for k in ['désignation', 'designation', 'actes', 'prestations']):
            started = True
            continue
        if started:
            if row[0] is None:
                continue
            if 'total' in str(row[0]).lower() or 'caisse' in str(row[0]).lower():
                break
            print(f"    - {row[0]} | Qty={row[1]} | Price={row[2]} | Total={row[3]} | RowVals={[v for v in row if v is not None]}")
            count += 1
            if count >= 3:
                break
    print("-" * 60)
