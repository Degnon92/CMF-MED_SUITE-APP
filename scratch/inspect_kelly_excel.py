import openpyxl
import os

workspace_dir = r"c:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheets containing KELLY:")
for name in wb.sheetnames:
    if 'KELLY' in name.upper():
        print(f"\nSheet name: {name}")
        ws = wb[name]
        for r in range(1, 35):
            row_vals = [ws.cell(r, c).value for c in range(1, 13)]
            if any(row_vals):
                print(f"  Row {r:02d}: {row_vals}")
