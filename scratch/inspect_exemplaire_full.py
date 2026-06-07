"""
inspect_exemplaire_full.py
Inspection complète de EXEMPLAIRE PROFORMA.xlsx avec encodage safe
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from pathlib import Path

EXCEL_FILE = Path(r"C:\Users\Farus\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy\PROFORMA CHIRURGIE\EXEMPLAIRE PROFORMA.xlsx")

print(f"=== {EXCEL_FILE.name} ===\n")

def safe(v, maxlen=40):
    if v is None: return ''
    s = str(v).strip()
    s = s.replace('\n', ' ').replace('\r', '')
    s = s.encode('utf-8', errors='replace').decode('utf-8')
    if len(s) > maxlen: s = s[:maxlen] + '...'
    return s

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"Nombre d'onglets : {len(wb.sheetnames)}")
print(f"Onglets : {wb.sheetnames}\n")

# Pour chaque onglet, afficher le nombre de colonnes et les 5 premières lignes non vides
for i, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    
    # Compter les colonnes réelles
    max_col = ws.max_column
    max_row = ws.max_row
    
    print(f"\n{'='*70}")
    print(f"[{i+1:02d}] ONGLET : '{safe(sheet_name)}' ({max_row}L x {max_col}C)")
    print('='*70)
    
    rows_shown = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(45, max_row), values_only=True), 1):
        non_empty = [v for v in row if v is not None and str(v).strip()]
        if not non_empty:
            continue
        vals = [safe(v, 35) for v in row]
        line = ' | '.join(v for v in vals if v)
        print(f"  L{row_idx:2d}: {line[:140]}")
        rows_shown += 1
    
    if rows_shown == 0:
        print("  (Onglet vide ou données masquées)")
    
    # Identifier si l'onglet a des colonnes assurance (split)
    has_split = False
    for row in ws.iter_rows(min_row=1, max_row=min(15, max_row), values_only=True):
        for v in row:
            if v and isinstance(v, str):
                vl = v.lower()
                if 'assur' in vl or 'patient' in vl or 'part' in vl or 'tiers' in vl:
                    has_split = True
    if has_split:
        print(f"  *** CET ONGLET CONTIENT DES DONNÉES ASSURANCE ***")

print(f"\n\nRÉSUMÉ : {len(wb.sheetnames)} onglets analysés.")
