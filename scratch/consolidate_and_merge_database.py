import json
import os
import re
import openpyxl
import difflib
import unicodedata

workspace_dir = r"c:\Users\Degnon\Documents\2.MERCY FIAT CLINIQUE\2. Dr Gipsy"
app_dir = os.path.join(workspace_dir, "MercyFiatMedSuiteDesktop")

patients_db_path = os.path.join(app_dir, "patients_db.json")
documents_db_path = os.path.join(app_dir, "documents_db.json")
bills_db_path = os.path.join(app_dir, "bills_db.json")
real_data_js_path = os.path.join(app_dir, "real_data.js")
database_js_path = os.path.join(app_dir, "database.js")
excel_path = os.path.join(workspace_dir, "EXEMPLAIRE PROFORMA.xlsx")

print("Starting Database Consolidation & De-duplication Process...")

# =====================================================================
# 1. RESOLVE FEUIL2 SHEET NAMES FROM EXCEL
# =====================================================================
print(f"Loading Excel workbook: {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

feuil_sheets = [s for s in wb.sheetnames if "feuil" in s.lower() and s.lower() not in ["feuil1", "feuil2", "feuil4"]]
print(f"Found {len(feuil_sheets)} sheets matching 'Feuil2...' pattern.")

def extract_name_from_value(val):
    if not val or not isinstance(val, str):
        return None
    val_clean = val.strip()
    
    # Strip prefixes that don't hold the name itself
    for prefix in ["patient:", "patient :", "client:", "client :", "diagnostic:", "diagnostic :", "intervention:", "intervention :"]:
        if val_clean.lower().startswith(prefix):
            return None
            
    m = re.search(r'(?:patient|client)\s*:\s*(.*)', val_clean, re.IGNORECASE)
    if m:
        return m.group(1).strip()
        
    upper = val_clean.upper()
    junk = ["CLINIQUE", "MEDECINE", "FACTURE", "PROFORMA", "ACTES", "DESIGNATIONS", "TOTAL", "COTINOU", "COTONOU", "CAISSE", "OPTION", "CHAMBRE"]
    if any(j in upper for j in junk):
        return None
        
    if re.search(r'[A-Za-z]', val_clean) and len(val_clean) > 3 and len(val_clean) < 50:
        return val_clean
        
    return None

feuil_map = {} # Maps normalized sheet name (e.g. 'FEUIL2(60)') to (resolved_name, diagnosis, intervention)

for sheetname in wb.sheetnames:
    # Normalize sheet name for matching
    sheet_key = re.sub(r'\s+', '', sheetname.upper())
    
    if sheetname.lower() in ["feuil1", "feuil2", "feuil4"]:
        continue
        
    sheet = wb[sheetname]
    patient_name = None
    diag = None
    interv = None
    
    # Scan rows 3, 4, 5 in column 1 for name, diagnosis, intervention
    for r in [3, 4, 5]:
        val = sheet.cell(r, 1).value
        if not val or not isinstance(val, str):
            continue
        val_clean = val.strip()
        
        m_pat = re.search(r'(?:patient|patiente)\s*:\s*(.*)', val_clean, re.IGNORECASE)
        if m_pat:
            patient_name = m_pat.group(1).strip()
            continue
            
        m_diag = re.search(r'diagnostic\s*:\s*(.*)', val_clean, re.IGNORECASE)
        if m_diag:
            diag = m_diag.group(1).strip()
            continue
            
        m_interv = re.search(r'intervention\s*:\s*(.*)', val_clean, re.IGNORECASE)
        if m_interv:
            interv = m_interv.group(1).strip()
            continue
            
        extracted = extract_name_from_value(val_clean)
        if extracted and not patient_name:
            patient_name = extracted

    # Clean the patient name from trailing insurer annotations like "TADOGBE CLAUDE(ASSUR SANLAM)"
    if patient_name:
        patient_name = re.sub(r'\(.*?\)', '', patient_name).strip()
        
    if patient_name:
        feuil_map[sheet_key] = (patient_name, diag or "Bilan clinique", interv or "")
    else:
        # Fallback to sheetname if absolutely unresolved
        feuil_map[sheet_key] = (sheetname, "Bilan clinique", "")

print(f"Resolved {len(feuil_map)} sheet names to patient identities.")

# =====================================================================
# 2. LOAD CURRENT JSON DATABASES
# =====================================================================
print("Loading existing database files...")
with open(patients_db_path, "r", encoding="utf-8") as f:
    patients = json.load(f)
with open(documents_db_path, "r", encoding="utf-8") as f:
    docs = json.load(f)
with open(bills_db_path, "r", encoding="utf-8") as f:
    bills = json.load(f)

print(f"Loaded {len(patients)} patients, {len(docs)} documents, {len(bills)} bills.")

# =====================================================================
# 3. APPLY FEUIL2 MAPPING TO DATABASES (FIRST PASS RESOLUTION)
# =====================================================================
print("Resolving Feuil2 anonymous records to real names...")

def resolve_feuil_name(name):
    if not name:
        return name
    # Check if this name matches Feuil pattern
    key = re.sub(r'\s+', '', name.upper())
    # Try sheetname format
    if key in feuil_map:
        return feuil_map[key][0]
    return name

# Map patients
for p in patients:
    p["name"] = resolve_feuil_name(p["name"])

# Map bills
for b in bills:
    fullname = f"{b['patientNom']} {b['patientPrenom']}".strip()
    resolved_full = resolve_feuil_name(fullname)
    if resolved_full != fullname:
        # Re-split resolved name into Nom/Prenom
        words = resolved_full.split()
        if len(words) >= 2:
            b["patientNom"] = words[0].upper()
            b["patientPrenom"] = " ".join(words[1:])
        else:
            b["patientNom"] = resolved_full.upper()
            b["patientPrenom"] = ""
            
    # Also resolve diagnostic and intervention if they were missing
    key = re.sub(r'\s+', '', fullname.upper())
    if key in feuil_map:
        _, diag, interv = feuil_map[key]
        if (not b.get("diagnostic") or b["diagnostic"] == "Bilan clinique") and diag != "Bilan clinique":
            b["diagnostic"] = diag
        if not b.get("intervention") and interv:
            b["intervention"] = interv
            b["showInterv"] = True

# Map documents
for d in docs:
    fullname = f"{d['patientNom']} {d['patientPrenom']}".strip()
    resolved_full = resolve_feuil_name(fullname)
    if resolved_full != fullname:
        words = resolved_full.split()
        if len(words) >= 2:
            d["patientNom"] = words[0].upper()
            d["patientPrenom"] = " ".join(words[1:])
        else:
            d["patientNom"] = resolved_full.upper()
            d["patientPrenom"] = ""

# =====================================================================
# 4. FUZZY DE-DUPLICATION & MERGING ALGORITHM
# =====================================================================
print("Grouping patients using fuzzy sequence similarity matching...")

def strip_accents(text):
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore')
    return text.decode("utf-8")

def normalize_name_for_merging(name):
    if not name:
        return ""
    # Convert to uppercase & strip accents
    s = strip_accents(name.upper())
    # Remove administrative words/parentheses
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'\b(?:OK|COPIE|POINT|DETAILS|ASSUR|URO|CHIPED|FGA|AVOIR|POINT DEF|POINT JR|DETAILS ASSUR)\b', '', s)
    # Remove middle initials like " J. ", " T. ", " P ", " S. ", " T "
    s = re.sub(r'\b[A-Z]\.?\b', ' ', s)
    # Keep letters
    s = re.sub(r'[^A-Z]', ' ', s)
    # Normalize spaces
    s = " ".join(s.split())
    return s

def clean_patient_name(name):
    if not name or not isinstance(name, str):
        return ""
    clean = name.strip()
    prefix_pattern = re.compile(
        r'^(?:CERTIFICAT\s+DE\s+MARIAGE|CERTIFICAT\s+DE\s+NON\s+BEGAIEMENT|CERTIFICAT\s+MEDICAL\s+INITIAL\s+DE\s+CONSTATATION\s+DE\s+COUPS\s+ET\s+BLESSURES|CERTIFICAT\s+MEDICAL\s+POUR\s+COUPS\s+ET\s+BLESSURES|CERTIFICAT\s+MEDICAL\s+DE\s+L[’\']ETAT\s+ACTUEL|CERTIFICAT\s+MEDICAL\s+INITIAL|CERTIFICAT\s+MEDICAL|CERTIFICAT\s+MED\s+INITIAL|CERTIFICAT\s+DE\s+REPOS|CERTIFICAT\s+DE\s+REPRISE|CERTIFICAT\s+DE\s+GUERISON|CERTIFICAT\s+DE\s+GUÉRISON|CERTIFICAT\s+DE|CERTIFICAT|RAPPORT\s+MEDICAL|RAPPORT\s+DE\s+MONSIEUR|RAPPORT\s+DE\s+MME|RAPPORT\s+DE|RAPPORT\s+D\'HOSPI|RAPPORT\s+D\'HOSPITALISATION|RAPPORT\s+DE\s+CONSULTATION|RAPPORT|CRO\s+MODELE|CRO|CMI|MEDICAL|MED\s+INITIAL|GUERISON\s+DE\s+MONSIEUR|GUERISON\s+DE\s+MME|GUERISON\s+DE|GUERISON|GUÉRISON|DECES\s+DE\s+MONSIEUR|DECES\s+DE\s+MME|DECES\s+DE|DECES|DÉCÈS|D\'HOSPI\s+TYPE|D\'HOSPI|DHOSPI|ATTESTATION\s+DE\s+GUERISON|ATTESTATION\s+DE\s+GUÉRISON|ATTESTATION\s+DE|ATTESTATION)\s+',
        re.IGNORECASE
    )
    old_clean = ""
    while clean != old_clean:
        old_clean = clean
        clean = prefix_pattern.sub("", clean).strip()
        
    split_pattern = re.compile(
        r'(?:CERTIFICAT|JE\s+SOUSSIGN|JE\s+SOUSSIGNE|RAPPORT|DOSSIER|N°|NO\s+DOSSIER|CMI|CRO|CLINIQUE|MÉDECINE|MEDECINE|DIAGNOSTIC|INTERVENTION|CLIENT|ASSURANCE|AFRICAINE|NSIA|ALLIANZ|SUNU|AROO|SAAR|CORIS|FEDAS|MUTUELLE|SÉJOUR|SEJOUR|DATE|OPÉRATOIRE|OPERATOIRE|CERTFICAT|CERTIF|PATIENT|PATIENTE|COTONOU|RUE\s+PAVILLON|TEL\s*\:|E\-MAIL|EMAIL|E\s+MAIL|GUERISON|GUÉRISON|DECES|DÉCÈS)',
        re.IGNORECASE
    )
    match = split_pattern.search(clean)
    if match:
        clean = clean[:match.start()].strip()
        
    clean = re.sub(r'(?:\s+|:)\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\b(?:ANS|AGE|ÂGE|ANS\s+D[\’\']ÂGE)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    if clean.upper().endswith("AGE") and len(clean) > 5:
        clean = clean[:-3].strip()
        
    clean = re.sub(r'[\s\-\.\,\:\_]+$', '', clean).strip()
    clean = re.sub(r'\s+', ' ', clean)
    clean = re.sub(r'\s+\(?(?:ok|copie|2|3|4|1)\)?$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s+fev\s+\d{2}$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s+nov\s+\d{2}$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s+aout\s+\d{2}$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s+sept\s+\d{2}$', '', clean, flags=re.IGNORECASE).strip()
    
    return clean

def split_nom_prenom(full_name):
    cleaned = clean_patient_name(full_name)
    if not cleaned:
        return "", ""
        
    parts = cleaned.split(' ')
    if len(parts) == 1:
        return cleaned.upper(), ""
        
    nom_parts = []
    prenom_parts = []
    
    for part in parts:
        if part.isupper() and len(part) > 1:
            nom_parts.append(part)
        else:
            prenom_parts.append(part)
            
    if not nom_parts:
        nom = parts[0].upper()
        prenom = " ".join(parts[1:])
    else:
        nom = " ".join(nom_parts).upper()
        prenom = " ".join(prenom_parts)
        
    if not prenom and len(parts) > 1:
        nom = parts[0].upper()
        prenom = " ".join(parts[1:])
        
    return nom.strip(), prenom.strip()

def get_name_score(name):
    if not name:
        return -100
    score = 0
    # Prefer mixed case
    if not name.isupper() and not name.islower():
        score += 10
    # Prefer letters only
    if re.match(r'^[A-Za-z\s]+$', name):
        score += 5
    # Penalize administrative elements
    if '(' in name or ')' in name:
        score -= 10
    if '.' in name or ',' in name:
        score -= 5
    score += min(len(name), 30)
    return score

def format_canonical_name(name):
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(r'\b(?:OK|COPIE|POINT|DETAILS|ASSUR|URO|CHIPED|FGA|AVOIR)\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'[\s\-\.\,\:\_]+$', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

groups = []
visited = set()

# Index patients by normalized name
for i, p1 in enumerate(patients):
    if i in visited:
        continue
    
    group = [p1]
    visited.add(i)
    norm1 = normalize_name_for_merging(p1["name"])
    if not norm1:
        continue
        
    for j, p2 in enumerate(patients):
        if j in visited:
            continue
        norm2 = normalize_name_for_merging(p2["name"])
        if not norm2:
            continue
            
        matched = False
        if norm1 == norm2:
            matched = True
        else:
            # Fuzzy check
            ratio = difflib.SequenceMatcher(None, norm1.replace(" ", ""), norm2.replace(" ", "")).ratio()
            if ratio >= 0.95:
                matched = True
            elif ratio >= 0.90:
                w1 = norm1.split()[0] if norm1.split() else ""
                w2 = norm2.split()[0] if norm2.split() else ""
                w_ratio = difflib.SequenceMatcher(None, w1, w2).ratio()
                if w_ratio >= 0.85:
                    matched = True
                    
        if matched:
            group.append(p2)
            visited.add(j)
            
    groups.append(group)

print(f"Consolidated patients list from {len(patients)} into {len(groups)} unique patient profiles.")

# Build mapping of all names in a group to their canonical name
name_mapping = {} # Old uppercase name -> Canonical name
canonical_patients = []

for group in groups:
    # 1. Determine canonical name by scoring names
    best_member = max(group, key=lambda x: get_name_score(x["name"]))
    canonical_name = format_canonical_name(best_member["name"])
    
    # 2. Extract best attributes across all duplicates
    ages = [p.get("age") for p in group]
    specific_ages = [a for a in ages if a and a != "N/A" and a != "45 ans" and a != "35 ans"]
    best_age = specific_ages[0] if specific_ages else (ages[0] if ages else "N/A")
    if not best_age or best_age == "N/A":
        # Fallback to any valid age
        valid_ages = [a for a in ages if a and a != "N/A"]
        best_age = valid_ages[0] if valid_ages else "N/A"
        
    diagnoses = [p.get("diagnosis") for p in group]
    specific_diag = [d for d in diagnoses if d and d != "Bilan clinique" and len(d) > 3]
    best_diag = max(specific_diag, key=len) if specific_diag else "Bilan clinique"
    
    interventions = [(p.get("intervention"), p.get("kCode")) for p in group]
    specific_interv = [t for t in interventions if t[0] and len(t[0]) > 3]
    best_interv, best_kcode = max(specific_interv, key=lambda x: len(x[0])) if specific_interv else ("", "")
    
    insurers = [(p.get("insurer"), p.get("priseEnCharge", 0)) for p in group]
    specific_insurer = [t for t in insurers if t[0] and t[0] != "PRIVE"]
    best_insurer, best_coverage = specific_insurer[0] if specific_insurer else ("PRIVE", 0)
    
    matricules = [p.get("matricule") for p in group]
    specific_mat = [m for m in matricules if m and m != "N/A" and len(m) > 1]
    best_matricule = specific_mat[0] if specific_mat else ""
    
    canonical_pat = {
        "name": canonical_name,
        "diagnosis": best_diag,
        "intervention": best_interv,
        "kCode": best_kcode,
        "age": best_age,
        "matricule": best_matricule,
        "insurer": best_insurer,
        "priseEnCharge": best_coverage
    }
    canonical_patients.append(canonical_pat)
    
    # Map every duplicate name in the group to this canonical name
    for member in group:
        name_mapping[member["name"].strip().upper()] = canonical_name

# =====================================================================
# 5. REWRITE RELATIONS IN DOCUMENTS & BILLS
# =====================================================================
print("Rewriting patient references in bills and documents to match canonical names...")

def update_record_name(record):
    nom = record.get("patientNom", "")
    prenom = record.get("patientPrenom", "")
    full_old = f"{nom} {prenom}".strip().upper()
    
    # 1. Match against mapping
    if full_old in name_mapping:
        canon = name_mapping[full_old]
        record_nom, record_prenom = split_nom_prenom(canon)
        record["patientNom"] = record_nom
        record["patientPrenom"] = record_prenom
    else:
        # Fallback to fuzzy match mapping if exact upper string didn't hit
        matched = False
        norm_old = normalize_name_for_merging(full_old)
        for old_mapped, canon in name_mapping.items():
            if normalize_name_for_merging(old_mapped) == norm_old:
                record_nom, record_prenom = split_nom_prenom(canon)
                record["patientNom"] = record_nom
                record["patientPrenom"] = record_prenom
                matched = True
                break
        if not matched:
            record["patientNom"] = nom.upper()
            record["patientPrenom"] = prenom

# Update all documents
for d in docs:
    update_record_name(d)
    
# Update all bills
for b in bills:
    update_record_name(b)

# Clean documents and bills list of any duplicate files/bills
unique_docs = {}
for d in docs:
    # Hash based on content snippet to deduplicate overlaps
    text_snippet = re.sub(r'\s+', '', (d.get("content") or d.get("text") or "")[:200].lower())
    key = (d["patientNom"], d["patientPrenom"], text_snippet)
    if key not in unique_docs:
        unique_docs[key] = d

final_docs = list(unique_docs.values())
for i, d in enumerate(final_docs):
    # Keep their sequence prefix intact
    if (d["id"] or "").startswith("DOC-REAL-"):
        d["id"] = f"DOC-REAL-AUTO-{i+1}"
    elif (d["id"] or "").startswith("DOC-CMF-H-"):
        d["id"] = f"DOC-CMF-H-{i+1:03d}"

unique_bills = {}
for b in bills:
    # Deduplicate bills of same patient, total amount, and date
    key = (b["patientNom"], b["patientPrenom"], b["grossTotal"], b["date"])
    if key not in unique_bills:
        unique_bills[key] = b
        
final_bills = list(unique_bills.values())
for i, b in enumerate(final_bills):
    if (b["id"] or "").startswith("BILL-REAL-"):
        b["id"] = f"BILL-REAL-AUTO-{i+1}"
    elif (b["id"] or "").startswith("BILL-CMF-X-"):
        b["id"] = f"BILL-CMF-X-{i+1:03d}"

print(f"Database statistics after consolidation:")
print(f"  Patients count: {len(canonical_patients)}")
print(f"  Documents count: {len(final_docs)}")
print(f"  Bills count: {len(final_bills)}")

# =====================================================================
# 6. WRITE BACK UPDATED DATABASES
# =====================================================================
print("Writing database files to disk...")

with open(patients_db_path, "w", encoding="utf-8") as f:
    json.dump(canonical_patients, f, indent=4, ensure_ascii=False)

with open(documents_db_path, "w", encoding="utf-8") as f:
    json.dump(final_docs, f, indent=4, ensure_ascii=False)

with open(bills_db_path, "w", encoding="utf-8") as f:
    json.dump(final_bills, f, indent=4, ensure_ascii=False)

# Update real_data.js
js_content = f"""/* ==========================================
   real_data.js - Vrais Rapports Médicaux Clinique Mercy Fiat
   ========================================== */

window.MercyFiatRealDocs = {json.dumps(final_docs, indent=4, ensure_ascii=False)};
"""
with open(real_data_js_path, "w", encoding="utf-8") as f:
    f.write(js_content)

print("Database files saved successfully.")

# =====================================================================
# 7. RE-EXTRACT UNIQUE DIAGNOSES & INTERVENTIONS -> database.js
# =====================================================================
print("Updating autocomplete list in database.js...")

diagnoses_set = set()
interventions_set = set()

for p in canonical_patients:
    if p.get("diagnosis") and p["diagnosis"] != "Bilan clinique":
        diagnoses_set.add(p["diagnosis"])
    if p.get("intervention"):
        interventions_set.add(p["intervention"])

for d in final_docs:
    if d.get("diagnosis") and d["diagnosis"] != "Bilan clinique":
        diagnoses_set.add(d["diagnosis"])
    if d.get("intervention"):
        interventions_set.add(d["intervention"])

for b in final_bills:
    if b.get("diagnostic") and b["diagnostic"] != "Bilan clinique":
        diagnoses_set.add(b["diagnostic"])
    if b.get("intervention"):
        interventions_set.add(b["intervention"])

# Load database.js and replace the DIAGNOSES and INTERVENTIONS arrays
with open(database_js_path, "r", encoding="utf-8") as f:
    db_content = f.read()

# Filter out empty or too long items and escape double quotes
clean_diagnoses = sorted([diag.replace('"', '\\"') for diag in diagnoses_set if diag and len(diag) > 4 and len(diag) < 120])
clean_interventions = sorted([inter.replace('"', '\\"') for inter in interventions_set if inter and len(inter) > 4 and len(inter) < 120])

# Replace DIAGNOSES array
diag_regex = re.compile(r'DIAGNOSES:\s*\[[\s\S]*?\s*\]', re.MULTILINE)
new_diags_str = f"DIAGNOSES: [\n" + ",\n".join(f'        "{d}"' for d in clean_diagnoses) + "\n    ]"
db_content = diag_regex.sub(new_diags_str, db_content)

# Replace INTERVENTIONS array
interv_regex = re.compile(r'INTERVENTIONS:\s*\[[\s\S]*?\s*\]', re.MULTILINE)
new_intervs_str = f"INTERVENTIONS: [\n" + ",\n".join(f'        "{i}"' for i in clean_interventions) + "\n    ]"
db_content = interv_regex.sub(new_intervs_str, db_content)

with open(database_js_path, "w", encoding="utf-8") as f:
    f.write(db_content)

print(f"Updated database.js with {len(clean_diagnoses)} unique diagnoses and {len(clean_interventions)} unique interventions.")
print("Consolidation process completed successfully!")
